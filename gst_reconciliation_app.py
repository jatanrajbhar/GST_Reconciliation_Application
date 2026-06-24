"""
GST Reconciliation Tool - Desktop Application
A portable desktop application for GST reconciliation processing.
"""

import os
import sys
import threading
import sqlite3
import json
from datetime import datetime
import customtkinter as ctk
from tkinter import filedialog, messagebox
import tkinter as tk
from tkinter import ttk
import pandas as pd
import numpy as np
import math 
from decimal import Decimal, ROUND_HALF_UP


def round_rupee(x):
    """Round a numeric value to nearest rupee using ROUND_HALF_UP behavior.
    Returns 0 for NaN or non-finite values."""
    try:
        if pd.isna(x) or not np.isfinite(x):
            return 0
        # Use Decimal for consistent HALF_UP rounding on .5
        d = Decimal(str(float(x)))
        return int(d.quantize(Decimal('1'), rounding=ROUND_HALF_UP))
    except Exception:
        return 0
from io import BytesIO
from PIL import Image
import re
from difflib import SequenceMatcher
import license_manager

# Set appearance mode and default color theme
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")  # Will override with custom magenta/pink colors

# Custom magenta/pink theme colors
THEME_PRIMARY = "#E91E63"  # Magenta Pink
THEME_HOVER = "#C2185B"    # Darker Pink
THEME_LIGHT = "#FCE4EC"    # Light Pink background
THEME_DARK = "#880E4F"     # Dark Pink/Maroon

# Default folder where all YTD databases are stored (auto-created)
GST_DB_DIR = os.path.join(os.path.expanduser("~"), "Documents", "GST_YTD_Databases")

# Increase Pandas Styler limit for large dataframes
pd.set_option("styler.render.max_elements", 5000000)


def get_resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)


def fix_sci_notation(val):
    """Convert '2.70E+14' / '270000000000000.0' strings to plain integer strings.
    Needed because Excel numeric cells read via openpyxl can produce float strings."""
    if not isinstance(val, str):
        return val
    v = val.strip()
    if not v:
        return v
    # Scientific notation: e.g. '2.70E+14', '2.70e+09'
    if re.match(r'^-?\d+\.?\d*[eE][+-]?\d+$', v):
        try:
            return str(int(float(v)))
        except (ValueError, OverflowError):
            return v
    # Float-like integer: '270000000.0' → '270000000'
    if re.match(r'^-?\d+\.0+$', v):
        return v[:v.index('.')]
    return v


def _build_period_pickers(parent_frame, period_var):
    """Render Month + Year dropdowns inside parent_frame, updating period_var on change."""
    _months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
               'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    _cur_yr = datetime.now().year
    _years  = [str(y) for y in range(2018, _cur_yr + 4)]

    # Parse any pre-existing value like "Apr 2025"
    _existing = (period_var.get() or '').strip().split()
    _pre_mo = _existing[0] if _existing and _existing[0] in _months else _months[0]
    _pre_yr = _existing[1] if len(_existing) > 1 and _existing[1] in _years else str(_cur_yr)

    _mo_var = tk.StringVar(value=_pre_mo)
    _yr_var = tk.StringVar(value=_pre_yr)

    def _sync(*_):
        period_var.set(f"{_mo_var.get()} {_yr_var.get()}")

    _row = ctk.CTkFrame(parent_frame, fg_color="transparent")
    _row.pack(fill="x", pady=(0, 10))
    ctk.CTkOptionMenu(_row, variable=_mo_var, values=_months,
                      width=100, height=34, command=lambda _: _sync()).pack(side="left", padx=(0, 8))
    ctk.CTkOptionMenu(_row, variable=_yr_var, values=_years,
                      width=100, height=34, command=lambda _: _sync()).pack(side="left")
    _sync()  # initialise period_var immediately


def fix_sci_notation_in_df(df, cols=None):
    """Apply fix_sci_notation to invoice-number-like columns in a DataFrame."""
    if df is None or df.empty:
        return df
    target_cols = cols or [c for c in df.columns
                           if any(k in c.lower() for k in
                                  ('invoice', 'doc no', 'boe', 'note no', 'ref'))]
    for col in target_cols:
        if col in df.columns and df[col].dtype == object:
            df[col] = df[col].apply(fix_sci_notation)
    return df


def safe_numeric_conversion(value):
    """Safely convert value to numeric, handling various formats"""
    if pd.isna(value):
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        # Remove common formatting characters
        value = value.replace(',', '').replace('\u20b9', '').strip()
        if value in ['', '-', 'NA', 'N/A', 'null', 'NULL', 'None']:
            return 0.0
        try:
            return float(value)
        except:
            return 0.0
    return 0.0


# Helper: normalize invoice strings for consistent matching
def normalize_invoice(s):
    """Return a normalized invoice string: uppercase, alphanumeric only, strip leading zeros.
    Also expands scientific notation (e.g. '2.7183E+13') to full integer string for consistent
    matching between CSV and Excel loaded data."""
    if pd.isna(s):
        return ''
    s = str(s).strip()
    # Expand scientific notation to full integer string (e.g. '2.7183E+13' -> '27183000000000')
    if re.match(r'^-?[\d.]+[eE][+\-]?\d+$', s):
        try:
            s = str(int(float(s)))
        except (ValueError, OverflowError):
            pass
    s = s.upper()
    # Remove date-like parts, whitespace, and non-alphanumeric characters
    s = re.sub(r'[^A-Z0-9]', '', s)
    s = s.lstrip('0')
    return s


# Helper: normalize GSTIN to canonical 15-char alphanum uppercase without spaces
def normalize_gstin(s):
    if pd.isna(s):
        return ''
    s = str(s).upper().replace(' ', '')
    s = re.sub(r'[^A-Z0-9]', '', s)
    return s


# Helper: fuzzy similarity between two strings
def similarity(a, b):
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a, b).ratio()


def find_tax_amount_columns(df):
    """Find the correct CGST, SGST, IGST amount columns, preferring 'Amount' columns over '%' columns.
    Returns (cgst_col, sgst_col, igst_col) tuple with column names or None if not found."""
    cgst_col = None
    sgst_col = None
    igst_col = None

    # First pass: look specifically for 'Amount' columns
    for col in df.columns:
        col_lower = col.lower().strip()
        if 'cgst' in col_lower and 'amount' in col_lower and not cgst_col:
            cgst_col = col
        elif 'sgst' in col_lower and 'amount' in col_lower and not sgst_col:
            sgst_col = col
        elif 'igst' in col_lower and 'amount' in col_lower and not igst_col:
            igst_col = col

    # Second pass: look for exact 'cgst', 'sgst', 'igst' columns (without % or amount suffix)
    for col in df.columns:
        col_lower = col.lower().strip()
        if col_lower == 'cgst' and not cgst_col:
            cgst_col = col
        elif col_lower == 'sgst' and not sgst_col:
            sgst_col = col
        elif col_lower == 'igst' and not igst_col:
            igst_col = col

    # Third pass: fall back to any column containing cgst/sgst/igst but NOT containing '%'
    for col in df.columns:
        col_lower = col.lower().strip()
        if 'cgst' in col_lower and '%' not in col_lower and not cgst_col:
            cgst_col = col
        elif 'sgst' in col_lower and '%' not in col_lower and not sgst_col:
            sgst_col = col
        elif 'igst' in col_lower and '%' not in col_lower and not igst_col:
            igst_col = col

    return cgst_col, sgst_col, igst_col


def find_booking_month_column(df):
    """Find the booking month / period column in a DataFrame.
    Priority: '2B MONTH' → 'Booking Month' → 'Month' → 'Period'.
    Returns column name or None."""
    if df is None or df.empty:
        return None
    for col in df.columns:
        cl = col.lower().strip()
        if cl == '2b month':
            return col
    for col in df.columns:
        cl = col.lower().strip()
        if cl == 'booking month':
            return col
    for col in df.columns:
        cl = col.lower().strip()
        if cl == 'month':
            return col
    for col in df.columns:
        cl = col.lower().strip()
        if cl == 'period':
            return col
    return None


def create_template_excel():
    """Return the bytes of Template updated.xlsx.
    Priority: (1) embedded bytes in template_data.py, (2) file next to exe, (3) minimal fallback.
    Using embedded data means the template is always available regardless of spec/build config."""
    # 1. Embedded data (always available — no spec file dependency)
    try:
        import template_data as _td
        return _td.get_template_updated_bytes()
    except Exception:
        pass

    # 2. File next to the script / in PyInstaller _MEIPASS folder
    for candidate in [
        get_resource_path("Template updated.xlsx"),
        get_resource_path("Template all.xlsx"),
    ]:
        if os.path.exists(candidate):
            with open(candidate, 'rb') as f:
                return f.read()

    # 3. Minimal fallback (last resort — very basic structure)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        sample_gstin = '27AABCU9603R1ZM'
        for sheet, inv_col in [('ITC- SR', 'Vendor Inv. No/  External Doc no'),
                                ('B2B- 2B', 'Invoice No'), ('B2BA- 2B', 'Invoice No'),
                                ('CDNR- 2B', 'Invoice No'), ('CDNRA- 2B', 'Invoice No'),
                                ('IMPG- 2B', 'BOE No'),    ('IMPGSEZ- 2B', 'BOE No')]:
            pd.DataFrame({inv_col: ['SAMPLE001'], 'GSTIN': [sample_gstin],
                           'CGST': [0.0], 'SGST': [0.0], 'IGST': [0.0]}
                         ).to_excel(writer, index=False, sheet_name=sheet)
    output.seek(0)
    return output.getvalue()


def normalize_itc_columns(itc_df):
    """Normalize ITC column names - handle Invoice No and GSTIN aliases"""
    if itc_df is None or itc_df.empty:
        return itc_df

    itc_df = itc_df.copy()

    # Convert all object/datetime columns to string to avoid type issues
    for col in itc_df.columns:
        if itc_df[col].dtype == 'object' or 'datetime' in str(itc_df[col].dtype):
            itc_df[col] = itc_df[col].astype(str).replace('nan', '').replace('NaT', '')

    has_invoice_no = False
    has_vendor_inv = False
    invoice_no_col = None
    vendor_inv_col = None

    for col in itc_df.columns:
        col_lower = col.lower().strip()
        if col_lower == 'invoice no':
            has_invoice_no = True
            invoice_no_col = col
        elif 'vendor inv' in col_lower or 'external doc' in col_lower:
            has_vendor_inv = True
            vendor_inv_col = col

    if has_invoice_no:
        if has_vendor_inv:
            itc_df = itc_df.rename(columns={invoice_no_col: 'Vendor Inv. No/ External Doc No'})
        else:
            itc_df = itc_df.rename(columns={invoice_no_col: 'Vendor Inv. No/ External Doc No'})

    has_vendor_gstn = False
    gstin_col_to_rename = None

    for col in itc_df.columns:
        col_lower = col.lower().strip()
        if 'vendor' in col_lower and 'gstn' in col_lower:
            has_vendor_gstn = True
        elif col_lower == 'gstin' or ('supplier' in col_lower and 'gstin' in col_lower):
            gstin_col_to_rename = col

    if gstin_col_to_rename and not has_vendor_gstn:
        itc_df = itc_df.rename(columns={gstin_col_to_rename: "Vendor's GSTN"})

    return itc_df


def normalize_cdnr_columns(df):
    """Normalize CDNR/CDNRA column names - handle Invoice No as BOE No alias"""
    if df is None or df.empty:
        return df

    df = df.copy()

    # Convert all object/datetime columns to string to avoid type issues
    for col in df.columns:
        if df[col].dtype == 'object' or 'datetime' in str(df[col].dtype):
            df[col] = df[col].astype(str).replace('nan', '').replace('NaT', '')

    has_invoice_no = False
    has_boe_no = False
    invoice_no_col = None

    for col in df.columns:
        col_lower = col.lower().strip()
        if col_lower == 'invoice no':
            has_invoice_no = True
            invoice_no_col = col
        elif 'boe no' in col_lower:
            has_boe_no = True

    if has_invoice_no:
        if has_boe_no:
            df = df.rename(columns={invoice_no_col: 'BOE No'})
        else:
            df = df.rename(columns={invoice_no_col: 'BOE No'})

    return df


def merge_duplicate_vendor_invoices(itc_df, log_callback=None):
    """Step 1: Merge duplicate Vendor Inv. No in ITC"""
    if itc_df is None or itc_df.empty:
        if log_callback:
            log_callback("Warning: ITC table is empty or not loaded")
        return itc_df

    # Create a copy to avoid modifying original
    itc_df = itc_df.copy()

    vendor_inv_col = None
    vendor_gstn_col = None

    for col in itc_df.columns:
        col_lower = col.lower().strip()
        if 'vendor inv' in col_lower or 'external doc' in col_lower:
            vendor_inv_col = col
        elif 'vendor' in col_lower and 'gstn' in col_lower:
            vendor_gstn_col = col

    # Use helper to find correct amount columns (not percentage columns)
    cgst_col, sgst_col, igst_col = find_tax_amount_columns(itc_df)

    if not vendor_inv_col:
        if log_callback:
            log_callback("Error: Could not find 'Vendor Inv. No/ External Doc No' column in ITC")
        return itc_df

    original_count = len(itc_df)

    # Convert all non-numeric columns to string to avoid datetime/mixed type issues
    for col in itc_df.columns:
        if col not in [cgst_col, sgst_col, igst_col]:
            # Convert to string, handling NaN values
            itc_df[col] = itc_df[col].astype(str).replace('nan', '').replace('NaT', '')

    numeric_cols = []
    if cgst_col:
        itc_df[cgst_col] = itc_df[cgst_col].apply(safe_numeric_conversion)
        numeric_cols.append(cgst_col)
    if sgst_col:
        itc_df[sgst_col] = itc_df[sgst_col].apply(safe_numeric_conversion)
        numeric_cols.append(sgst_col)
    if igst_col:
        itc_df[igst_col] = itc_df[igst_col].apply(safe_numeric_conversion)
        numeric_cols.append(igst_col)

    agg_dict = {}
    for col in itc_df.columns:
        if col in numeric_cols:
            agg_dict[col] = 'sum'
        else:
            agg_dict[col] = 'first'

    merged_itc = itc_df.groupby(vendor_inv_col, as_index=False).agg(agg_dict)
    merged_count = len(merged_itc)

    if log_callback:
        log_callback(f"Step 1: Merged {original_count - merged_count} duplicate entries. Original: {original_count} -> After merge: {merged_count}")

    return merged_itc


def match_and_update(main_df, amendment_df, main_invoice_col, amend_invoice_col, table_name, log_callback=None):
    """Generic function to match and update tables (Steps 2 & 3).
    Returns (updated_main_df, remaining_amendment_df) where remaining has matched rows removed."""
    if main_df is None or main_df.empty:
        if log_callback:
            log_callback(f"{table_name} table is empty")
        return main_df, amendment_df

    if amendment_df is None or amendment_df.empty:
        if log_callback:
            log_callback(f"Amendment table is empty, skipping update for {table_name}")
        return main_df, amendment_df

    # Use global helper to find correct amount columns (not percentage columns)
    main_cgst, main_sgst, main_igst = find_tax_amount_columns(main_df)
    amend_cgst, amend_sgst, amend_igst = find_tax_amount_columns(amendment_df)

    if not main_invoice_col or not amend_invoice_col:
        if log_callback:
            log_callback(f"Could not find invoice columns for {table_name}")
        return main_df, amendment_df

    for df, cols in [(main_df, [main_cgst, main_sgst, main_igst]),
                     (amendment_df, [amend_cgst, amend_sgst, amend_igst])]:
        for col in cols:
            if col and col in df.columns:
                df[col] = df[col].apply(safe_numeric_conversion)

    matched_count = 0
    matched_indices = []
    main_df = main_df.copy()

    for idx, amend_row in amendment_df.iterrows():
        amend_invoice = amend_row[amend_invoice_col]
        matching_mask = main_df[main_invoice_col] == amend_invoice

        if matching_mask.any():
            matched_count += 1
            matched_indices.append(idx)
            if main_cgst and amend_cgst:
                main_df.loc[matching_mask, main_cgst] = amend_row[amend_cgst]
            if main_sgst and amend_sgst:
                main_df.loc[matching_mask, main_sgst] = amend_row[amend_sgst]
            if main_igst and amend_igst:
                main_df.loc[matching_mask, main_igst] = amend_row[amend_igst]

    # Remove matched rows from amendment table
    remaining_amendment = amendment_df.drop(matched_indices).reset_index(drop=True)

    # Round tax columns in the main table to nearest rupee (if decimals exist) and guard against NaN/inf
    for col in [main_cgst, main_sgst, main_igst]:
        if col and col in main_df.columns:
            # Ensure numeric and round to integer using HALF_UP, replacing NaN/inf with 0
            main_df[col] = main_df[col].apply(safe_numeric_conversion).apply(lambda x: round_rupee(x))

    # Reset index and return
    main_df = main_df.reset_index(drop=True)
    remaining_amendment = remaining_amendment.reset_index(drop=True)

    if log_callback:
        if matched_count > 0:
            log_callback(f"Matched and updated {matched_count} records in {table_name}, removed from amendment table")
        else:
            log_callback(f"No matches found for {table_name}, proceeding to next step")

    return main_df, remaining_amendment


def create_merged_table(tables_dict, log_callback=None):
    """Step 4: Create MERGED table"""
    merged_data = []
    table_order = ['B2B', 'B2BA', 'CDNR', 'CDNRA', 'IMPG', 'IMPGSEZ']

    for table_name in table_order:
        df = tables_dict.get(table_name)

        if df is None or df.empty:
            if log_callback:
                log_callback(f"{table_name}: No data")
            continue

        doc_col = None
        gstn_col = None
        tax_col = None
        date_col = None

        for col in df.columns:
            col_lower = col.lower().strip()
            if 'invoice no' in col_lower or 'note no' in col_lower or 'boe no' in col_lower:
                if not doc_col:
                    doc_col = col
            # Supplier/vendor GSTIN — explicitly skip 'My GSTIN' (the buyer's own GSTIN)
            if ('gstn' in col_lower or 'gstin' in col_lower) and 'my' not in col_lower:
                if not gstn_col:
                    gstn_col = col
            if 'taxable' in col_lower and 'value' in col_lower and not tax_col:
                tax_col = col
            if col_lower == 'invoice date' and not date_col:
                date_col = col

        # Second pass: fallback to any gstin column if still not found (e.g. file has only 'My GSTIN')
        if not gstn_col:
            for col in df.columns:
                col_lower = col.lower().strip()
                if 'gstn' in col_lower or 'gstin' in col_lower:
                    gstn_col = col
                    break

        # Use helper to find correct amount columns (not percentage columns)
        cgst_col, sgst_col, igst_col = find_tax_amount_columns(df)
        bm_col = find_booking_month_column(df)

        if log_callback:
            log_callback(f"{table_name}: doc='{doc_col}' gstn='{gstn_col}' "
                         f"cgst='{cgst_col}' sgst='{sgst_col}' igst='{igst_col}'")

        if not doc_col:
            if log_callback:
                log_callback(f"{table_name}: Could not find document number column — "
                             f"available cols: {list(df.columns)}")
            continue

        if not gstn_col and log_callback:
            log_callback(f"{table_name}: Warning — GSTIN column not found, matching will be by invoice only")

        for _, row in df.iterrows():
            record = {
                'Document_number': row[doc_col] if doc_col else None,
                'GSTN': row[gstn_col] if gstn_col else None,
                'CGST': safe_numeric_conversion(row[cgst_col]) if cgst_col else None,
                'SGST': safe_numeric_conversion(row[sgst_col]) if sgst_col else None,
                'IGST': safe_numeric_conversion(row[igst_col]) if igst_col else None,
                'TAX': safe_numeric_conversion(row[tax_col]) if tax_col else 0.0,
                'Invoice_Date': str(row[date_col]) if date_col else '',
                'Booking_Month': str(row[bm_col]).strip() if bm_col else '',
                'TYPE': table_name
            }
            merged_data.append(record)

        if log_callback:
            log_callback(f"Added {len(df)} records from {table_name}")

    merged_df = pd.DataFrame(merged_data)
    if log_callback:
        log_callback(f"Step 4: MERGED table created with {len(merged_df)} total records")

    return merged_df


def create_itc_register(itc_df, log_callback=None):
    """Step 5: Create as_per_itc_register table"""
    if itc_df is None or itc_df.empty:
        if log_callback:
            log_callback("Error: ITC table is empty")
        return pd.DataFrame()

    vendor_gstn_col = None
    vendor_inv_col = None
    tax_col = None

    for col in itc_df.columns:
        col_lower = col.lower().strip()
        if 'vendor' in col_lower and 'gstn' in col_lower:
            vendor_gstn_col = col
        elif 'vendor inv' in col_lower or 'external doc' in col_lower:
            vendor_inv_col = col
        elif 'taxable' in col_lower and 'value' in col_lower and not tax_col:
            tax_col = col

    # Use helper to find correct amount columns (not percentage columns)
    cgst_col, sgst_col, igst_col = find_tax_amount_columns(itc_df)

    if log_callback:
        log_callback(f"ITC columns: gstn='{vendor_gstn_col}' inv='{vendor_inv_col}' "
                     f"cgst='{cgst_col}' sgst='{sgst_col}' igst='{igst_col}'")

    if not vendor_gstn_col or not vendor_inv_col:
        if log_callback:
            log_callback(f"Error: Could not find required columns in ITC — "
                         f"available cols: {list(itc_df.columns)}")
        return pd.DataFrame()

    itc_register = pd.DataFrame({
        'GSTINinvoice': itc_df[vendor_gstn_col].astype(str) + itc_df[vendor_inv_col].astype(str),
        'CGST': itc_df[cgst_col].apply(safe_numeric_conversion) if cgst_col else 0.0,
        'SGST': itc_df[sgst_col].apply(safe_numeric_conversion) if sgst_col else 0.0,
        'IGST': itc_df[igst_col].apply(safe_numeric_conversion) if igst_col else 0.0,
        'TAX': itc_df[tax_col].apply(safe_numeric_conversion) if tax_col else 0.0
    })

    if log_callback:
        log_callback(f"Step 5: Created as_per_itc_register with {len(itc_register)} records")

    return itc_register


def create_gstr_2a(merged_df, log_callback=None):
    """Step 6: Create as_per_gtsr_2a table"""
    if merged_df is None or merged_df.empty:
        if log_callback:
            log_callback("Error: MERGED table is empty")
        return pd.DataFrame()

    gstr_2a = pd.DataFrame({
        'GSTINinvoice': merged_df['GSTN'].astype(str) + merged_df['Document_number'].astype(str),
        'CGST': merged_df['CGST'].apply(safe_numeric_conversion),
        'SGST': merged_df['SGST'].apply(safe_numeric_conversion),
        'IGST': merged_df['IGST'].apply(safe_numeric_conversion),
        'TAX': merged_df['TAX'].apply(safe_numeric_conversion) if 'TAX' in merged_df.columns else 0.0
    })

    if log_callback:
        log_callback(f"Step 6: Created as_per_gtsr_2a with {len(gstr_2a)} records")

    return gstr_2a


def create_itc_register_2(itc_df, log_callback=None):
    """Step 9: Create as_per_itc_register(2) table (per-line) with normalized GSTIN and numeric TAX"""
    if itc_df is None or itc_df.empty:
        if log_callback:
            log_callback("Error: ITC table is empty for Step 9")
        return pd.DataFrame()

    vendor_gstn_col = None
    tax_col = None

    for col in itc_df.columns:
        col_lower = col.lower().strip()
        if 'vendor' in col_lower and 'gstn' in col_lower:
            vendor_gstn_col = col
        elif 'taxable' in col_lower and 'value' in col_lower and not tax_col:
            tax_col = col

    # Use helper to find correct amount columns (not percentage columns)
    cgst_col, sgst_col, igst_col = find_tax_amount_columns(itc_df)

    if not vendor_gstn_col or not tax_col:
        if log_callback:
            log_callback("Error: Could not find GSTIN or Taxable Value columns in ITC for Step 9")
        return pd.DataFrame()

    # Build per-line register with normalized GSTIN and numeric TAX (rounded to nearest rupee)
    reg = pd.DataFrame()
    reg['GSTIN'] = itc_df[vendor_gstn_col].apply(normalize_gstin)
    reg['TAX'] = itc_df[tax_col].apply(safe_numeric_conversion).apply(lambda x: round_rupee(x))
    reg['CGST'] = (itc_df[cgst_col].apply(safe_numeric_conversion).apply(lambda x: round_rupee(x))) if cgst_col else 0
    reg['SGST'] = (itc_df[sgst_col].apply(safe_numeric_conversion).apply(lambda x: round_rupee(x))) if sgst_col else 0
    reg['IGST'] = (itc_df[igst_col].apply(safe_numeric_conversion).apply(lambda x: round_rupee(x))) if igst_col else 0

    if log_callback:
        log_callback(f"Step 9: Created as_per_itc_register(2) (per-line) with {len(reg)} records")

    return reg


def create_gstr_2a_2(merged_df, log_callback=None):
    """Step 10: Create as_per_gtsr_2a(2) table with normalized GSTIN and numeric TAX"""
    if merged_df is None or merged_df.empty:
        if log_callback:
            log_callback("Error: MERGED table is empty for Step 10")
        return pd.DataFrame()

    df = merged_df.copy()
    df['GSTIN'] = df['GSTN'].astype(str).apply(normalize_gstin)
    df['TAX'] = (df['TAX'].apply(safe_numeric_conversion).apply(lambda x: round_rupee(x))) if 'TAX' in df.columns else 0
    df['CGST'] = df['CGST'].apply(safe_numeric_conversion).apply(lambda x: round_rupee(x))
    df['SGST'] = df['SGST'].apply(safe_numeric_conversion).apply(lambda x: round_rupee(x))
    df['IGST'] = df['IGST'].apply(safe_numeric_conversion).apply(lambda x: round_rupee(x))

    gstr_2a_2 = df[['GSTIN', 'TAX', 'CGST', 'SGST', 'IGST']].copy()

    if log_callback:
        log_callback(f"Step 10: Created as_per_gtsr_2a(2) with {len(gstr_2a_2)} records")

    return gstr_2a_2


def compare_tables_by_tax(itc_register_2, gstr_2a_2, log_callback=None):
    """Step 11: Compare tables by GSTIN and TAX (numeric), and map results back to per-line ITC rows."""
    if itc_register_2.empty or gstr_2a_2.empty:
        if log_callback:
            log_callback("Error: One or both comparison tables are empty for tax matching")
        return pd.DataFrame()

    # Both inputs are per-line DataFrames with 'GSTIN' and 'TAX' numeric columns
    # Aggregate to (GSTIN, TAX) level for stable comparison
    itc_agg = itc_register_2.groupby(['GSTIN', 'TAX'], as_index=False).agg({'CGST':'sum','SGST':'sum','IGST':'sum'})
    gstr_agg = gstr_2a_2.groupby(['GSTIN', 'TAX'], as_index=False).agg({'CGST':'sum','SGST':'sum','IGST':'sum'})

    # Round aggregated tax columns to nearest rupee (integers) to remove decimal points
    for c in ['CGST', 'SGST', 'IGST']:
        if c in itc_agg.columns:
            vals = itc_agg[c].fillna(0).astype(float).to_numpy()
            itc_agg[c] = np.floor(vals + 0.5).astype(int)
        if c in gstr_agg.columns:
            vals = gstr_agg[c].fillna(0).astype(float).to_numpy()
            gstr_agg[c] = np.floor(vals + 0.5).astype(int)

    # Merge on GSTIN and TAX
    comparison = pd.merge(itc_agg, gstr_agg, on=['GSTIN','TAX'], how='outer', suffixes=('_ITC','_GSTR2A'))

    for col in ['CGST_ITC', 'SGST_ITC', 'IGST_ITC', 'CGST_GSTR2A', 'SGST_GSTR2A', 'IGST_GSTR2A']:
        if col in comparison.columns:
            comparison[col] = comparison[col].fillna(0)
        else:
            comparison[col] = 0

    comparison['CGST_Difference'] = comparison['CGST_ITC'] - comparison['CGST_GSTR2A']
    comparison['SGST_Difference'] = comparison['SGST_ITC'] - comparison['SGST_GSTR2A']
    comparison['IGST_Difference'] = comparison['IGST_ITC'] - comparison['IGST_GSTR2A']

    tolerance = 10.0

    def get_status(row):
        cgst_match = abs(row['CGST_Difference']) <= tolerance
        sgst_match = abs(row['SGST_Difference']) <= tolerance
        igst_match = abs(row['IGST_Difference']) <= tolerance

        return 'MATCH' if (cgst_match and sgst_match and igst_match) else 'MISMATCH'

    comparison['Status'] = comparison.apply(get_status, axis=1)

    # Map statuses back to per-line ITC rows so user sees full original lines
    # Itc_register_2 contains per-line GSTIN and TAX
    itc_lines = itc_register_2.copy()
    itc_lines['Status'] = itc_lines.merge(
        comparison[['GSTIN','TAX','Status']],
        on=['GSTIN','TAX'], how='left')['Status']
    itc_lines['Status'] = itc_lines['Status'].fillna('MISMATCH')

    matched = len(comparison[comparison['Status'] == 'MATCH'])
    mismatched = len(comparison[comparison['Status'] == 'MISMATCH'])
    total = len(comparison)

    if log_callback:
        log_callback(f"Step 11: Tax Value Matching - Total invoice groups: {total}, Matched: {matched}, Mismatched: {mismatched}")
        log_callback(f"Step 11: Tax value mapped back to {len(itc_lines)} ITC line items")

    return itc_lines


def compare_tables(itc_register, gstr_2a, log_callback=None):
    """Step 7: Compare and show detailed matching results using normalized GSTIN+Invoice keys."""
    if itc_register is None or itc_register.empty or gstr_2a is None or gstr_2a.empty:
        if log_callback:
            log_callback("Error: One or both comparison tables are empty")
        return pd.DataFrame()

    # Work on copies
    itc = itc_register.copy()
    gstr = gstr_2a.copy()

    # Helper to split concatenated GSTINinvoice and normalize parts
    def _split_and_normalize(s):
        if pd.isna(s) or str(s).strip() == '':
            return ('', '')
        raw = str(s).upper()
        m = re.search(r'([0-9A-Z]{15})', raw)
        if m:
            gstin_raw = m.group(1)
            inv_raw = raw.replace(gstin_raw, '', 1)
        else:
            gstin_raw = raw[:15] if len(raw) >= 15 else raw
            inv_raw = raw[15:] if len(raw) > 15 else ''
        return (normalize_gstin(gstin_raw), normalize_invoice(inv_raw))

    def derive_norm_parts(df_input):
        """Extract and normalize GSTIN/Invoice from concatenated or separate columns."""
        if 'GSTINinvoice' in df_input.columns:
            parts = df_input['GSTINinvoice'].astype(str).apply(_split_and_normalize)
            df_input['_norm_gstin'] = parts.apply(lambda x: x[0])
            df_input['_norm_inv'] = parts.apply(lambda x: x[1])
        else:
            df_input['_norm_gstin'] = ''
            df_input['_norm_inv'] = ''
        return df_input

    itc = derive_norm_parts(itc)
    gstr = derive_norm_parts(gstr)

    # Build normalized composite key
    itc['_key'] = itc['_norm_gstin'].fillna('').astype(str) + '|' + itc['_norm_inv'].fillna('').astype(str)
    gstr['_key'] = gstr['_norm_gstin'].fillna('').astype(str) + '|' + gstr['_norm_inv'].fillna('').astype(str)

    # Aggregate taxes by normalized key
    itc_agg = itc.groupby('_key', as_index=False).agg({'CGST': 'sum', 'SGST': 'sum', 'IGST': 'sum'})
    gstr_agg = gstr.groupby('_key', as_index=False).agg({'CGST': 'sum', 'SGST': 'sum', 'IGST': 'sum'})

    # Ensure numeric and round to nearest rupee
    for df_agg in [itc_agg, gstr_agg]:
        for col in ['CGST', 'SGST', 'IGST']:
            df_agg[col] = df_agg[col].apply(safe_numeric_conversion).apply(lambda x: round_rupee(x))

    # Rename columns before merge
    itc_agg = itc_agg.rename(columns={'CGST': 'CGST_ITC', 'SGST': 'SGST_ITC', 'IGST': 'IGST_ITC'})
    gstr_agg = gstr_agg.rename(columns={'CGST': 'CGST_GSTR2A', 'SGST': 'SGST_GSTR2A', 'IGST': 'IGST_GSTR2A'})

    # Merge on normalized key
    comparison = pd.merge(itc_agg, gstr_agg, on='_key', how='outer')

    # Fill NaN with 0
    for col in ['CGST_ITC', 'SGST_ITC', 'IGST_ITC', 'CGST_GSTR2A', 'SGST_GSTR2A', 'IGST_GSTR2A']:
        if col not in comparison.columns:
            comparison[col] = 0
        comparison[col] = comparison[col].fillna(0)

    # Compute differences
    comparison['CGST_Difference'] = comparison['CGST_ITC'] - comparison['CGST_GSTR2A']
    comparison['SGST_Difference'] = comparison['SGST_ITC'] - comparison['SGST_GSTR2A']
    comparison['IGST_Difference'] = comparison['IGST_ITC'] - comparison['IGST_GSTR2A']

    # Build compact GSTINinvoice (no separator) for downstream compatibility
    def _compact(k):
        if pd.isna(k) or k == '':
            return ''
        gstin, inv = k.split('|', 1) if '|' in k else (k[:15], k[15:])
        return (gstin or '') + (inv or '')
    
    comparison['GSTINinvoice'] = comparison['_key'].apply(_compact)

    # Compare with tolerance
    tolerance = 10.0
    def _get_status(r):
        return 'MATCH' if (abs(r['CGST_Difference']) <= tolerance and 
                          abs(r['SGST_Difference']) <= tolerance and 
                          abs(r['IGST_Difference']) <= tolerance) else 'MISMATCH'
    
    comparison['Status'] = comparison.apply(_get_status, axis=1)

    # Clean up and reorder columns
    comparison = comparison.drop(columns=['_key'])
    
    cols_order = ['GSTINinvoice', 'CGST_ITC', 'SGST_ITC', 'IGST_ITC', 'CGST_GSTR2A', 'SGST_GSTR2A', 'IGST_GSTR2A', 'CGST_Difference', 'SGST_Difference', 'IGST_Difference', 'Status']
    existing_cols = [c for c in cols_order if c in comparison.columns]
    remaining_cols = [c for c in comparison.columns if c not in existing_cols]
    comparison = comparison[existing_cols + remaining_cols]

    total_records = len(comparison)
    matched = len(comparison[comparison['Status'] == 'MATCH'])
    mismatched = len(comparison[comparison['Status'] == 'MISMATCH'])

    if log_callback:
        log_callback(f"Step 7: Reconciliation complete - Total: {total_records}, Matched: {matched}, Mismatched: {mismatched}")

    return comparison


def create_itc_result(itc_df, itc_register, gstr_2a, comparison_df=None, merged_df=None, log_callback=None):
    """Create ITC table with improved matching logic and use `comparison_df` when available for consistency.
    If `comparison_df` is provided (aggregated matches from compare_tables), it will be used to assign invoice-level
    statuses (preferred) and then mapped back to original ITC rows so counts in ITC Results mirror reconciliation.

    Status values: Matched, Unmatched, Higher in 2A, Lower in 2A"""
    if itc_df is None or itc_df.empty:
        if log_callback:
            log_callback("Error: ITC table is empty")
        return pd.DataFrame()

    result = itc_df.copy()

    # Find GSTN and Invoice No columns in ITC
    vendor_gstn_col = None
    vendor_inv_col = None

    for col in itc_df.columns:
        col_lower = col.lower().strip()
        if 'vendor' in col_lower and 'gstn' in col_lower:
            vendor_gstn_col = col
        elif 'vendor inv' in col_lower or 'external doc' in col_lower:
            vendor_inv_col = col

    # Use helper to find correct amount columns (not percentage columns)
    itc_cgst_col, itc_sgst_col, itc_igst_col = find_tax_amount_columns(itc_df)

    if not vendor_gstn_col or not vendor_inv_col:
        if log_callback:
            log_callback("Error: Could not find GSTN or Invoice No columns in ITC")
        result['Status'] = 'Unmatched'
        return result

    # Build normalized keys on original ITC rows
    result['_norm_gstin'] = result[vendor_gstn_col].apply(normalize_gstin)
    result['_norm_inv'] = result[vendor_inv_col].apply(normalize_invoice)
    result['GSTINinvoice_norm'] = result['_norm_gstin'] + '|' + result['_norm_inv']

    # If comparison_df is available, use it to build a status lookup (prefer this for consistency)
    status_lookup = {}
    tolerance = 10.0


    # If any keys are still missing, fall back to earlier normalized matching algorithm
    # Aggregate original ITC to invoice level for comparison
    # Convert tax columns to numeric BEFORE aggregation (they are strings from dtype=str loading)
    agg_src = result[['GSTINinvoice_norm']].copy()
    if itc_cgst_col:
        agg_src['CGST'] = result[itc_cgst_col].apply(safe_numeric_conversion)
    if itc_sgst_col:
        agg_src['SGST'] = result[itc_sgst_col].apply(safe_numeric_conversion)
    if itc_igst_col:
        agg_src['IGST'] = result[itc_igst_col].apply(safe_numeric_conversion)

    agg_cols = {c: 'sum' for c in ['CGST', 'SGST', 'IGST'] if c in agg_src.columns}
    itc_agg = agg_src.groupby('GSTINinvoice_norm').agg(agg_cols).reset_index()

    # Build gstr lookup from provided gstr_2a (normalized)
    gstr_2a_local = gstr_2a.copy()
    if 'GSTINinvoice' in gstr_2a_local.columns:
        gstr_2a_local['_gstin_raw'] = gstr_2a_local['GSTINinvoice'].astype(str).apply(lambda x: x[:15] if len(x) >= 15 else x)
        gstr_2a_local['_inv_raw'] = gstr_2a_local['GSTINinvoice'].astype(str).apply(lambda x: x[15:] if len(x) > 15 else '')
    else:
        gstr_2a_local['_gstin_raw'] = ''
        gstr_2a_local['_inv_raw'] = ''
    if 'GSTN' in gstr_2a_local.columns and 'Document_number' in gstr_2a_local.columns:
        gstr_2a_local['_gstin_raw'] = gstr_2a_local['GSTN'].astype(str)
        gstr_2a_local['_inv_raw'] = gstr_2a_local['Document_number'].astype(str)
    gstr_2a_local['_norm_gstin'] = gstr_2a_local['_gstin_raw'].apply(normalize_gstin)
    gstr_2a_local['_norm_inv'] = gstr_2a_local['_inv_raw'].apply(normalize_invoice)
    gstr_2a_local['GSTINinvoice_norm'] = gstr_2a_local['_norm_gstin'] + '|' + gstr_2a_local['_norm_inv']
    gstr_agg = gstr_2a_local.groupby('GSTINinvoice_norm', as_index=False).agg({'CGST': 'sum', 'SGST': 'sum', 'IGST': 'sum'})
    for c in ['CGST', 'SGST', 'IGST']:
        if c in gstr_agg.columns:
            gstr_agg[c] = gstr_agg[c].apply(safe_numeric_conversion)

    gstr_lookup = {row['GSTINinvoice_norm']:{'CGST':row['CGST'],'SGST':row['SGST'],'IGST':row['IGST']} for _, row in gstr_agg.iterrows()}
    itc_lookup = {row['GSTINinvoice_norm']:{'CGST':row.get('CGST',0),'SGST':row.get('SGST',0),'IGST':row.get('IGST',0)} for _, row in itc_agg.iterrows()}

    # Build TYPE, Booking Month, Invoice Number, and Invoice Date lookups from merged_df for 2A columns
    type_lookup = {}
    booking_month_2a_lookup = {}  # key → booking month string from 2A files
    inv_2a_lookup = {}  # key → original Document_number (invoice no) from 2A
    inv_date_2b_lookup = {}     # key → invoice date string from 2B file
    taxable_2b_lookup  = {}     # key → taxable value from 2B file
    if merged_df is not None and not merged_df.empty:
        for _, row in merged_df.iterrows():
            gstin = normalize_gstin(str(row.get('GSTN', '')))
            inv = normalize_invoice(str(row.get('Document_number', '')))
            key = gstin + '|' + inv
            if key not in type_lookup:
                type_lookup[key] = str(row.get('TYPE', ''))
            if key not in booking_month_2a_lookup:
                bm = str(row.get('Booking_Month', '')).strip()
                if bm:
                    booking_month_2a_lookup[key] = bm
            if key not in inv_2a_lookup:
                inv_2a_lookup[key] = str(row.get('Document_number', ''))
            if key not in inv_date_2b_lookup:
                inv_date_2b_lookup[key] = str(row.get('Invoice_Date', '')).strip()
            if key not in taxable_2b_lookup:
                taxable_2b_lookup[key] = safe_numeric_conversion(row.get('TAX', 0))
    vals_2a_lookup = {}  # key → {'CGST', 'SGST', 'IGST', 'TYPE'}
    matched_2a_status = {}  # normalized 2A key → '2A Status' string ('Matched' / 'Unmatched')

    if comparison_df is not None and not comparison_df.empty:
        # Only use MATCH statuses from comparison_df
        # For MISMATCH, let the fuzzy matching below handle it (can find cross-fiscal-year matches by GSTIN + similar amounts)
        for _, row in comparison_df.iterrows():
            raw = str(row.get('GSTINinvoice', ''))
            # try to extract 15-char GSTIN if present
            m = re.search(r'[0-9A-Z]{15}', raw.upper())
            if m:
                gstin_raw = m.group(0)
                inv_raw = raw.upper().replace(gstin_raw, '', 1)
            else:
                gstin_raw = raw[:15]
                inv_raw = raw[15:]

            key = normalize_gstin(gstin_raw) + '|' + normalize_invoice(inv_raw)
            status_raw = str(row.get('Status', '')).upper()
            if status_raw == 'MATCH' or status_raw == 'MATCHED':
                # Only accept this MATCH if a real 2A entry actually exists for this key.
                # When ITC amounts aggregate to zero (e.g. invoice + reversal) and 2A has no
                # entry for that invoice, compare_tables produces a spurious 0-vs-0 MATCH.
                # In that case, skip here and let the fallback loop assign 'Not found in 2B'.
                if key in gstr_lookup:
                    status_lookup[key] = 'Matched'
                    g = gstr_lookup[key]
                    vals_2a_lookup[key] = {'CGST': g['CGST'], 'SGST': g['SGST'], 'IGST': g['IGST'], 'TAX': taxable_2b_lookup.get(key, 0), 'TYPE': type_lookup.get(key, ''), 'BM': booking_month_2a_lookup.get(key, ''), 'INV': inv_2a_lookup.get(key, 'Not Found'), 'DATE': inv_date_2b_lookup.get(key, ''), 'GSTIN': key.split('|', 1)[0], '_2A_KEY': key}
                    # Only mark 2A entry as Matched when there is a real ITC entry for this key.
                    # Pure 2A-only keys (no ITC counterpart) with small amounts can get a spurious
                    # MATCH from compare_tables (0 ITC vs small 2A ≤ tolerance). Without this
                    # guard those entries incorrectly show Status='Matched' in 2B Results.
                    if key in itc_lookup:
                        matched_2a_status[key] = 'Matched'
            # For MISMATCH cases, don't set status here - let fuzzy matching below handle them
            # This allows finding matches with different fiscal year suffixes (e.g., 21-22 vs 22-23)

    # per-GSTIN candidate sets
    gstr_by_gstin = {}
    for _, row in gstr_agg.iterrows():
        gstin = row['GSTINinvoice_norm'].split('|',1)[0]
        if gstin not in gstr_by_gstin:
            gstr_by_gstin[gstin] = []
        gstr_by_gstin[gstin].append({'key': row['GSTINinvoice_norm'], 'CGST': row['CGST'], 'SGST': row['SGST'], 'IGST': row['IGST']})

    # Employ fallback matching only for keys not set yet
    for key in set(list(itc_lookup.keys()) + list(gstr_lookup.keys())):
        if key in status_lookup:
            continue
        matched_key = None
        if key in gstr_lookup:
            itc_vals = itc_lookup.get(key, {'CGST':0,'SGST':0,'IGST':0})
            gstr_vals = gstr_lookup.get(key, {'CGST':0,'SGST':0,'IGST':0})
            matched_key = key
        else:
            gstin, inv = key.split('|',1)
            itc_vals = itc_lookup.get(key, {'CGST':0,'SGST':0,'IGST':0})
            candidates = gstr_by_gstin.get(gstin, [])
            gstr_vals = None
            best = None
            best_diff = None
            for cand in candidates:
                cand_inv = cand['key'].split('|', 1)[1]
                # Only allow amount-based matching when invoice numbers have a substring
                # relationship (one is contained in the other). This handles cross-fiscal-year
                # formats like '2021676' ↔ '20202021676' or '2' ↔ '22021', while preventing
                # false matches between genuinely different invoices (e.g. '3344' ↔ '3349').
                _s, _l = (inv, cand_inv) if len(inv) <= len(cand_inv) else (cand_inv, inv)
                if _s and _l and not _l.endswith(_s) and not _l.startswith(_s):
                    continue
                cg_diff = abs(itc_vals['CGST'] - cand['CGST'])
                sg_diff = abs(itc_vals['SGST'] - cand['SGST'])
                ig_diff = abs(itc_vals['IGST'] - cand['IGST'])
                total_diff = cg_diff + sg_diff + ig_diff
                if best is None or total_diff < best_diff:
                    best = cand
                    best_diff = total_diff
            if best is not None and best_diff <= (tolerance * 3):
                gstr_vals = {'CGST':best['CGST'],'SGST':best['SGST'],'IGST':best['IGST']}
                matched_key = best['key']
            if gstr_vals is None and candidates:
                best_sim = 0.0
                best_cand = None
                for cand in candidates:
                    cand_inv = cand['key'].split('|',1)[1]
                    _s2, _l2 = (inv, cand_inv) if len(inv) <= len(cand_inv) else (cand_inv, inv)
                    if _s2 and _l2 and not _l2.endswith(_s2) and not _l2.startswith(_s2):
                        continue
                    sim = similarity(inv, cand_inv)
                    if sim > best_sim:
                        best_sim = sim
                        best_cand = cand
                if best_sim >= 0.85:
                    gstr_vals = {'CGST':best_cand['CGST'],'SGST':best_cand['SGST'],'IGST':best_cand['IGST']}
                    matched_key = best_cand['key']
            if gstr_vals is None:
                gstr_vals = {'CGST':0,'SGST':0,'IGST':0}
        cgst_diff = itc_vals['CGST'] - gstr_vals['CGST']
        sgst_diff = itc_vals['SGST'] - gstr_vals['SGST']
        igst_diff = itc_vals['IGST'] - gstr_vals['IGST']
        exact_match_in_2a = key in gstr_lookup
        if matched_key is not None and abs(cgst_diff) <= tolerance and abs(sgst_diff) <= tolerance and abs(igst_diff) <= tolerance:
            status = 'Matched'
        elif exact_match_in_2a:
            # Exact invoice found in 2A but amounts differ
            total_diff = cgst_diff + sgst_diff + igst_diff
            status = 'Higher in 2B' if total_diff < -tolerance else 'Lower in 2B'
        else:
            # Invoice not found in 2A at all (fuzzy match didn't produce exact match)
            status = 'Not found in 2B'
        status_lookup[key] = status
        # Track matched 2A status: only when key represents an ITC entry and a real 2A match was found
        if key in itc_lookup and matched_key is not None and status in ('Matched', 'Higher in 2B', 'Lower in 2B'):
            matched_2a_status[matched_key] = status
        # Store 2A values for keys that found a real match in 2A
        if matched_key:
            vals_2a_lookup[key] = {
                'CGST': gstr_vals['CGST'], 'SGST': gstr_vals['SGST'], 'IGST': gstr_vals['IGST'],
                'TAX': taxable_2b_lookup.get(matched_key, 0),
                'TYPE': type_lookup.get(matched_key, ''),
                'BM': booking_month_2a_lookup.get(matched_key, ''),
                'INV': inv_2a_lookup.get(matched_key, 'Not Found'),
                'DATE': inv_date_2b_lookup.get(matched_key, ''),
                'GSTIN': matched_key.split('|', 1)[0],
                '_2A_KEY': matched_key
            }

    # Ensure consistency between Status and 2A column values
    for key, status in status_lookup.items():
        if status == 'Not found in 2B':
            # Remove from vals_2a_lookup so 2A columns show the not-found label
            vals_2a_lookup.pop(key, None)
        elif status in ('Matched', 'Higher in 2B', 'Lower in 2B') and key not in vals_2a_lookup:
            # Found in comparison/matching but vals_2a_lookup not yet populated
            if key in gstr_lookup:
                g = gstr_lookup[key]
                vals_2a_lookup[key] = {'CGST': g['CGST'], 'SGST': g['SGST'], 'IGST': g['IGST'], 'TAX': taxable_2b_lookup.get(key, 0), 'TYPE': type_lookup.get(key, ''), 'BM': booking_month_2a_lookup.get(key, ''), 'INV': inv_2a_lookup.get(key, 'Not Found'), 'DATE': inv_date_2b_lookup.get(key, ''), 'GSTIN': key.split('|', 1)[0], '_2A_KEY': key}
                if key in itc_lookup:
                    matched_2a_status[key] = status
            else:
                # Try same-GSTIN fuzzy match from gstr_by_gstin
                # Apply the same substring constraint as the main fallback loop to prevent
                # false matches between invoices with unrelated numbers.
                gstin_part = key.split('|', 1)[0]
                inv_part = key.split('|', 1)[1]
                cands = gstr_by_gstin.get(gstin_part, [])
                # Filter candidates using suffix/prefix check to prevent short invoice numbers
                # (e.g. '10') from matching unrelated longer strings (e.g. '20202105').
                # A match is allowed only when the shorter number appears at the END or START
                # of the longer one (e.g. '10' at end of '20202110', '676' at end of '20202021676').
                filtered_cands = []
                for _fc in cands:
                    _ci = _fc['key'].split('|', 1)[1]
                    _s3, _l3 = (inv_part, _ci) if len(inv_part) <= len(_ci) else (_ci, inv_part)
                    if not _s3 or not _l3 or _l3.endswith(_s3) or _l3.startswith(_s3):
                        filtered_cands.append(_fc)
                if filtered_cands:
                    # Pick the candidate with smallest total tax difference to ITC
                    itc_v = itc_lookup.get(key, {"CGST": 0, "SGST": 0, "IGST": 0})
                    best_c = min(filtered_cands, key=lambda c: abs(itc_v["CGST"] - c["CGST"]) + abs(itc_v["SGST"] - c["SGST"]) + abs(itc_v["IGST"] - c["IGST"]))
                    vals_2a_lookup[key] = {'CGST': best_c['CGST'], 'SGST': best_c['SGST'], 'IGST': best_c['IGST'], 'TAX': taxable_2b_lookup.get(best_c['key'], 0), 'TYPE': type_lookup.get(best_c['key'], ''), 'BM': booking_month_2a_lookup.get(best_c['key'], ''), 'INV': inv_2a_lookup.get(best_c['key'], 'Not Found'), 'DATE': inv_date_2b_lookup.get(best_c['key'], ''), 'GSTIN': best_c['key'].split('|', 1)[0], '_2A_KEY': best_c['key']}
                    if key in itc_lookup:
                        matched_2a_status[best_c['key']] = status

    # Enforce one-to-one matching: if multiple ITC keys claimed the same 2A key via fuzzy matching,
    # keep only the best match (lowest total tax diff; exact key match wins ties) and set the rest
    # to 'Not found in 2B'.  This prevents false double-counting of 2A CGST/SGST/IGST totals
    # when two ITC entries like '90620634' and '90620634 CANCELL' both match the same 2A entry.
    _a2a_claims = {}
    for _itc_k, _vals in vals_2a_lookup.items():
        _a2a_k = _vals.get('_2A_KEY')
        if not _a2a_k or _itc_k not in itc_lookup:
            continue
        _itc_v = itc_lookup.get(_itc_k, {'CGST': 0, 'SGST': 0, 'IGST': 0})
        _diff = abs(_itc_v['CGST'] - _vals['CGST']) + abs(_itc_v['SGST'] - _vals['SGST']) + abs(_itc_v['IGST'] - _vals['IGST'])
        if _itc_k == _a2a_k:
            _diff = -1  # exact key match gets top priority
        _a2a_claims.setdefault(_a2a_k, []).append((_itc_k, _diff))
    for _a2a_k, _claimants in _a2a_claims.items():
        if len(_claimants) <= 1:
            continue
        _claimants.sort(key=lambda x: x[1])
        for _loser_k, _ in _claimants[1:]:
            if status_lookup.get(_loser_k) in ('Matched', 'Higher in 2B', 'Lower in 2B'):
                status_lookup[_loser_k] = 'Not found in 2B'
                vals_2a_lookup.pop(_loser_k, None)

    # Map status back to each original ITC row
    result['Status'] = result['GSTINinvoice_norm'].map(status_lookup).fillna('Unmatched')

    # Add 2A columns (CGST/SGST/IGST as per 2A, Type)
    not_found_label = 'Not found in 2B'

    # When multiple ITC rows share the same invoice key (duplicate invoice numbers),
    # distribute 2A aggregate values proportionally based on each row's ITC share.
    # E.g. if ITC rows are 1000 and 500 (total 1500) and 2A total is 1500,
    # row 1 gets 1000 and row 2 gets 500 — not 1500 repeated on both rows.
    _r = result  # alias for brevity
    itc_tax_col = next((c for c in itc_df.columns
                        if 'taxable' in c.lower() and 'value' in c.lower()), None)
    _itc_c  = _r[itc_cgst_col].apply(safe_numeric_conversion)  if itc_cgst_col  else pd.Series(0.0, index=_r.index)
    _itc_s  = _r[itc_sgst_col].apply(safe_numeric_conversion)  if itc_sgst_col  else pd.Series(0.0, index=_r.index)
    _itc_i  = _r[itc_igst_col].apply(safe_numeric_conversion)  if itc_igst_col  else pd.Series(0.0, index=_r.index)
    _itc_tv = _r[itc_tax_col].apply(safe_numeric_conversion)    if itc_tax_col   else pd.Series(0.0, index=_r.index)
    _key_c_sum  = _itc_c.groupby(_r['GSTINinvoice_norm']).transform('sum')
    _key_s_sum  = _itc_s.groupby(_r['GSTINinvoice_norm']).transform('sum')
    _key_i_sum  = _itc_i.groupby(_r['GSTINinvoice_norm']).transform('sum')
    _key_tv_sum = _itc_tv.groupby(_r['GSTINinvoice_norm']).transform('sum')
    _key_cnt    = _r.groupby('GSTINinvoice_norm')['GSTINinvoice_norm'].transform('count')

    def _prop_val(k, tax_key, itc_row_val, itc_key_total, row_cnt):
        if k not in vals_2a_lookup:
            return not_found_label
        v = vals_2a_lookup[k].get(tax_key, not_found_label)
        if isinstance(v, str):
            return v
        if row_cnt == 1:
            return v  # single row — no distribution needed
        if itc_key_total != 0:
            return round(itc_row_val / itc_key_total * v, 2)
        return round(v / row_cnt, 2)  # equal split when ITC total is zero

    result['CGST as per 2B'] = [_prop_val(k, 'CGST', c, ct, n)
        for k, c, ct, n in zip(_r['GSTINinvoice_norm'], _itc_c, _key_c_sum, _key_cnt)]
    result['SGST as per 2B'] = [_prop_val(k, 'SGST', s, st, n)
        for k, s, st, n in zip(_r['GSTINinvoice_norm'], _itc_s, _key_s_sum, _key_cnt)]
    result['IGST as per 2B'] = [_prop_val(k, 'IGST', i, it, n)
        for k, i, it, n in zip(_r['GSTINinvoice_norm'], _itc_i, _key_i_sum, _key_cnt)]
    if itc_tax_col:
        result['Taxable Value as per 2B'] = [_prop_val(k, 'TAX', tv, tvt, n)
            for k, tv, tvt, n in zip(_r['GSTINinvoice_norm'], _itc_tv, _key_tv_sum, _key_cnt)]
    result['Type'] = result['GSTINinvoice_norm'].map(
        lambda k: vals_2a_lookup[k]['TYPE'] if k in vals_2a_lookup else not_found_label)

    # Add Booking Month columns
    # 'Booking Month as per GSTR-2B': from the matched 2A file's booking month/period
    result['Booking Month as per GSTR-2B'] = result['GSTINinvoice_norm'].map(
        lambda k: vals_2a_lookup[k]['BM'] if k in vals_2a_lookup and vals_2a_lookup[k].get('BM') else not_found_label)
    # 'Booking Month as per ITC': from the ITC file's own booking month column
    itc_bm_col = find_booking_month_column(itc_df)
    if itc_bm_col and itc_bm_col in result.columns:
        result['Booking Month as per ITC'] = result[itc_bm_col].astype(str).str.strip()
    else:
        result['Booking Month as per ITC'] = ''

    # Add 2A Invoice No column: the original invoice number from the matched 2A row
    result['2B Invoice No'] = result['GSTINinvoice_norm'].map(
        lambda k: vals_2a_lookup[k]['INV'] if k in vals_2a_lookup and 'INV' in vals_2a_lookup[k] else 'Not Found')
    # Add 2A GSTIN column: the GSTIN of the matched 2A supplier
    result['2B GSTIN'] = result['GSTINinvoice_norm'].map(
        lambda k: vals_2a_lookup[k]['GSTIN'] if k in vals_2a_lookup and 'GSTIN' in vals_2a_lookup[k] else 'Not Found')

    # Invoice Date Match: compare ITC invoice date vs 2B invoice date for found rows
    itc_date_col = next(
        (c for c in itc_df.columns if c.lower().strip() == 'invoice date'), None)

    def _norm_date(s):
        """Return a comparable date object, or None on parse failure."""
        s = str(s).strip()
        if not s or s in ('nan', 'NaT', 'None', ''):
            return None
        try:
            d = pd.to_datetime(s, dayfirst=False, errors='coerce')
            if pd.isna(d):
                d = pd.to_datetime(s, dayfirst=True, errors='coerce')
            return d.date() if not pd.isna(d) else None
        except Exception:
            return None

    def _date_match(norm_key, itc_status, itc_date_val):
        if itc_status in ('Not found in 2B', 'Unmatched'):
            return ''
        b2b_date_str = vals_2a_lookup.get(norm_key, {}).get('DATE', '')
        if not b2b_date_str or b2b_date_str in ('nan', 'NaT', 'None', ''):
            return ''
        d_itc = _norm_date(itc_date_val)
        d_2b  = _norm_date(b2b_date_str)
        if d_itc is None or d_2b is None:
            return ''
        return 'Yes' if d_itc == d_2b else 'No'

    if itc_date_col and itc_date_col in result.columns:
        result['Invoice Date Match'] = [
            _date_match(k, st, dv)
            for k, st, dv in zip(
                result['GSTINinvoice_norm'],
                result['Status'],
                result[itc_date_col])
        ]
    else:
        result['Invoice Date Match'] = ''

    # Diff columns: ITC value − 2B value (blank for Not found / Unmatched rows)
    _skip_statuses = {'Not found in 2B', 'Unmatched'}

    def _diff_col(itc_series, b2b_col_name):
        """Return a series of (ITC − 2B) differences; blank string where not applicable."""
        out = []
        b2b_series = result[b2b_col_name] if b2b_col_name in result.columns else None
        for idx, (st, itc_v) in enumerate(zip(result['Status'], itc_series)):
            if st in _skip_statuses or b2b_series is None:
                out.append('')
                continue
            b2b_v = b2b_series.iloc[idx]
            if isinstance(b2b_v, str):   # 'Not found in 2B' label
                out.append('')
                continue
            out.append(round(safe_numeric_conversion(itc_v) - safe_numeric_conversion(b2b_v), 2))
        return out

    if itc_cgst_col and itc_cgst_col in result.columns:
        result['Diff CGST'] = _diff_col(result[itc_cgst_col], 'CGST as per 2B')
    if itc_sgst_col and itc_sgst_col in result.columns:
        result['Diff SGST'] = _diff_col(result[itc_sgst_col], 'SGST as per 2B')
    if itc_igst_col and itc_igst_col in result.columns:
        result['Diff IGST'] = _diff_col(result[itc_igst_col], 'IGST as per 2B')
    if itc_tax_col and itc_tax_col in result.columns and 'Taxable Value as per 2B' in result.columns:
        result['Diff Taxable Value'] = _diff_col(result[itc_tax_col], 'Taxable Value as per 2B')

    # Auto-remarks for many-to-one / one-to-many matches
    # Case 1: single ITC invoice matched with multiple 2A invoice rows (same norm key in 2A)
    _2a_row_count = gstr_2a_local.groupby('GSTINinvoice_norm').size().to_dict()
    # Case 2: multiple distinct ITC invoice keys mapped to same 2A key (via fuzzy match)
    _2a_key_to_itc_keys = {}
    for _ik, _vv in vals_2a_lookup.items():
        _ak = _vv.get('_2A_KEY')
        if _ak and _ik in itc_lookup:
            _2a_key_to_itc_keys.setdefault(_ak, set()).add(_ik)

    _matched_statuses = {'Matched', 'Higher in 2B', 'Lower in 2B'}
    remark_list = []
    for _nk in result['GSTINinvoice_norm']:
        _status = status_lookup.get(_nk, '')
        if _status not in _matched_statuses or _nk not in vals_2a_lookup:
            remark_list.append('')
            continue
        _ak = vals_2a_lookup[_nk].get('_2A_KEY', _nk)
        _parts = []
        if _2a_row_count.get(_ak, 1) > 1:
            _parts.append('invoice matches with multiple invoice in 2b')
        if len(_2a_key_to_itc_keys.get(_ak, set())) > 1:
            _parts.append('multiple invoice matches with single invoice in 2B')
        remark_list.append('; '.join(_parts))
    result['Remarks'] = remark_list

    # Clean helper columns
    result = result.drop(columns=[c for c in ['GSTINinvoice_norm','_norm_gstin','_norm_inv'] if c in result.columns])

    # Summary
    total = len(result)
    matched_count = (result['Status'] == 'Matched').sum()
    unmatched_count = (result['Status'] == 'Unmatched').sum()
    higher_count = (result['Status'] == 'Higher in 2B').sum()
    lower_count = (result['Status'] == 'Lower in 2B').sum()
    not_found_count = (result['Status'] == 'Not found in 2B').sum()

    if log_callback:
        log_callback(f"ITC Results: Total: {total}, Matched: {matched_count}, Unmatched: {unmatched_count}, Higher in 2B: {higher_count}, Lower in 2B: {lower_count}, Not found in 2B: {not_found_count}")
        log_callback(f"ITC result table created with {total} records (mapped back to original ITC line items)")

    return result, matched_2a_status


def match_cdnr_negatives(cdnr_df, cdnra_df, itc_result_df, log_callback=None):
    """Match CDNR/CDNRA entries with negative ITC values by GSTIN + tax amount key.

    CDNR/CDNRA have positive tax values; corresponding ITC entries have negative values.
    This creates a composite key (GSTIN + IGST or CGST) to match them.
    Only upgrades unmatched ITC rows to 'Matched' — never downgrades already-matched rows.
    """
    if itc_result_df is None or itc_result_df.empty:
        return itc_result_df, set()

    # Combine CDNR + remaining CDNRA
    cdnr_frames = []
    if cdnr_df is not None and not cdnr_df.empty:
        cdnr_frames.append(cdnr_df)
    if cdnra_df is not None and not cdnra_df.empty:
        cdnr_frames.append(cdnra_df)
    if not cdnr_frames:
        if log_callback:
            log_callback("CDNR matching: No CDNR/CDNRA data available, skipping")
        return itc_result_df, set()

    combined_cdnr = pd.concat(cdnr_frames, ignore_index=True)

    # Find CDNR columns
    cdnr_gstin_col = None
    cdnr_doc_col = None  # Note No column for tracking matched 2A keys
    cdnr_cgst_col, cdnr_sgst_col, cdnr_igst_col = find_tax_amount_columns(combined_cdnr)
    for col in combined_cdnr.columns:
        cl = col.lower().strip()
        if 'gstin' in cl or 'gstn' in cl:
            cdnr_gstin_col = col
        if ('note no' in cl or 'invoice no' in cl or 'boe no' in cl) and not cdnr_doc_col:
            cdnr_doc_col = col

    if not cdnr_gstin_col:
        if log_callback:
            log_callback("CDNR matching: Could not find GSTIN column in CDNR, skipping")
        return itc_result_df, set()

    # Find SGST column in CDNR for 2A column values
    cdnr_sgst_col = None
    for col in combined_cdnr.columns:
        cl = col.lower().strip()
        if cl == 'sgst' or ('sgst' in cl and 'amount' in cl):
            cdnr_sgst_col = col
            break

    # Find booking month column in CDNR
    cdnr_bm_col = find_booking_month_column(combined_cdnr)

    # Build CDNR key dict: key → list of {original values} (for one-to-one consumption)
    cdnr_keys = {}
    for idx, row in combined_cdnr.iterrows():
        gstin = normalize_gstin(str(row.get(cdnr_gstin_col, '')))
        igst = safe_numeric_conversion(row.get(cdnr_igst_col, 0)) if cdnr_igst_col else 0
        cgst = safe_numeric_conversion(row.get(cdnr_cgst_col, 0)) if cdnr_cgst_col else 0
        sgst = safe_numeric_conversion(row.get(cdnr_sgst_col, 0)) if cdnr_sgst_col else 0

        # Store original (positive) values for 2A columns
        orig_igst = igst
        orig_cgst = cgst
        orig_sgst = sgst

        # Negate positive values for matching key
        if igst > 0:
            igst = -igst
        if cgst > 0:
            cgst = -cgst

        # Determine key tax: IGST if non-zero, else CGST
        if igst != 0:
            key_tax = igst
        elif cgst != 0:
            key_tax = cgst
        else:
            continue  # both zero, skip

        # Determine source table
        src_type = 'CDNR'
        if idx >= len(cdnr_frames[0]) if len(cdnr_frames) > 1 else False:
            src_type = 'CDNRA'

        key = gstin + '|' + str(round(key_tax, 2))
        if key not in cdnr_keys:
            cdnr_keys[key] = []
        cdnr_bm = str(row.get(cdnr_bm_col, '')).strip() if cdnr_bm_col else ''
        cdnr_note_no = str(row.get(cdnr_doc_col, '')).strip() if cdnr_doc_col else ''
        cdnr_keys[key].append({
            'CGST': orig_cgst, 'SGST': orig_sgst, 'IGST': orig_igst, 'TYPE': src_type, 'BM': cdnr_bm,
            'GSTIN': gstin, 'NOTE_NO': cdnr_note_no
        })

    if not cdnr_keys:
        if log_callback:
            log_callback("CDNR matching: No valid CDNR entries with non-zero tax, skipping")
        return itc_result_df, set()

    # Set to track which 2A (CDNR) normalized keys were matched
    matched_cdnr_2a_keys = set()

    # Find ITC columns
    itc_gstin_col = None
    itc_inv_col = None
    itc_cgst_col, itc_sgst_col, itc_igst_col = find_tax_amount_columns(itc_result_df)
    for col in itc_result_df.columns:
        cl = col.lower().strip()
        if 'vendor' in cl and 'gstn' in cl:
            itc_gstin_col = col
        elif 'vendor inv' in cl or 'external doc' in cl:
            itc_inv_col = col

    if not itc_gstin_col or not itc_inv_col:
        if log_callback:
            log_callback("CDNR matching: Could not find GSTIN/Invoice columns in ITC, skipping")
        return itc_result_df, set()

    # Statuses that should NOT be changed (already matched)
    protected_statuses = {'Matched', 'Matched but invoice number is not accurate'}

    # Track which GSTINinvoice_norm keys to upgrade
    matched_norm_keys = set()
    matched_cdnr_info = {}  # inv_key → {'CGST', 'SGST', 'IGST', 'TYPE'}
    result = itc_result_df.copy()

    # Build normalized keys for grouping sibling rows
    result['_norm_gstin'] = result[itc_gstin_col].apply(normalize_gstin)
    result['_norm_inv'] = result[itc_inv_col].apply(normalize_invoice)
    result['_gstin_inv_norm'] = result['_norm_gstin'] + '|' + result['_norm_inv']

    # Process unique invoice groups (not individual line items)
    seen_inv_keys = set()
    for idx, row in result.iterrows():
        inv_key = row['_gstin_inv_norm']
        if inv_key in seen_inv_keys:
            continue
        seen_inv_keys.add(inv_key)

        # Skip already matched rows
        if row.get('Status', '') in protected_statuses:
            continue

        gstin = row['_norm_gstin']
        igst = safe_numeric_conversion(row.get(itc_igst_col, 0)) if itc_igst_col else 0
        cgst = safe_numeric_conversion(row.get(itc_cgst_col, 0)) if itc_cgst_col else 0

        # Only consider rows with negative IGST or CGST (credit note entries)
        if igst >= 0 and cgst >= 0:
            continue

        # Determine key tax: IGST if non-zero, else CGST
        if igst != 0:
            key_tax = igst
        elif cgst != 0:
            key_tax = cgst
        else:
            continue

        key = gstin + '|' + str(round(key_tax, 2))

        # Check if CDNR has a matching entry
        if key in cdnr_keys and len(cdnr_keys[key]) > 0:
            cdnr_info = cdnr_keys[key].pop()  # consume one CDNR entry
            matched_norm_keys.add(inv_key)
            matched_cdnr_info[inv_key] = cdnr_info
            # Track the matched CDNR 2A key (GSTIN + normalized Note No)
            cdnr_2a_key = cdnr_info.get('GSTIN', gstin) + '|' + normalize_invoice(cdnr_info.get('NOTE_NO', ''))
            if cdnr_2a_key.split('|', 1)[1]:  # only add if Note No is non-empty
                matched_cdnr_2a_keys.add(cdnr_2a_key)

    # Update all sibling rows for matched invoices
    has_2a_cols = 'CGST as per 2B' in result.columns
    if matched_norm_keys:
        mask = result['_gstin_inv_norm'].isin(matched_norm_keys)
        result.loc[mask, 'Status'] = 'Matched'
        # Update 2A columns with CDNR values
        if has_2a_cols:
            for inv_key, info in matched_cdnr_info.items():
                row_mask = result['_gstin_inv_norm'] == inv_key
                # Negate CDNR amounts to match 2B_Results sign convention (CDNR is negated there)
                result.loc[row_mask, 'CGST as per 2B'] = -info['CGST']
                result.loc[row_mask, 'SGST as per 2B'] = -info['SGST']
                result.loc[row_mask, 'IGST as per 2B'] = -info['IGST']
                result.loc[row_mask, 'Type'] = info['TYPE']
                if 'Booking Month as per GSTR-2B' in result.columns and info.get('BM'):
                    result.loc[row_mask, 'Booking Month as per GSTR-2B'] = info['BM']
                if '2B Invoice No' in result.columns:
                    result.loc[row_mask, '2B Invoice No'] = info.get('NOTE_NO', 'Not Found')
                if '2B GSTIN' in result.columns:
                    result.loc[row_mask, '2B GSTIN'] = info.get('GSTIN', 'Not Found')

    # Clean helper columns
    result = result.drop(columns=['_norm_gstin', '_norm_inv', '_gstin_inv_norm'])

    matched_count = len(matched_norm_keys)
    matched_rows = mask.sum() if matched_norm_keys else 0
    if log_callback:
        log_callback(f"CDNR matching: {matched_count} invoices ({matched_rows} rows) matched via CDNR/CDNRA negative matching")

    return result, matched_cdnr_2a_keys


def find_same_month_cancellations(b2b_df, cdnr_df, log_callback=None):
    """Find B2B invoices whose GSTIN and CGST/SGST/IGST amounts match a CDNR entry.

    Matching rules:
    - GSTIN must match exactly (normalized)
    - CGST, SGST, and IGST amounts must all match within a ₹10 tolerance
    - Invoice number does NOT need to match

    Returns a DataFrame with matched B2B-CDNR pairs side by side, or an empty DataFrame.
    """
    if b2b_df is None or b2b_df.empty or cdnr_df is None or cdnr_df.empty:
        return pd.DataFrame()

    tolerance = 10.0

    # ── Locate columns in B2B ───────────────────────────────────────────
    b2b_gstin_col = None
    b2b_inv_col = None
    b2b_date_col = None
    for col in b2b_df.columns:
        cl = col.lower().strip()
        if ('gstin' in cl or 'gstn' in cl) and not b2b_gstin_col:
            b2b_gstin_col = col
        if 'invoice no' in cl and not b2b_inv_col:
            b2b_inv_col = col
        if 'invoice date' in cl and not b2b_date_col:
            b2b_date_col = col
    b2b_cgst_col, b2b_sgst_col, b2b_igst_col = find_tax_amount_columns(b2b_df)

    # ── Locate columns in CDNR ──────────────────────────────────────────
    cdnr_gstin_col = None
    cdnr_note_col = None
    cdnr_date_col = None
    for col in cdnr_df.columns:
        cl = col.lower().strip()
        if ('gstin' in cl or 'gstn' in cl) and not cdnr_gstin_col:
            cdnr_gstin_col = col
        if ('note no' in cl or 'invoice no' in cl or 'boe no' in cl) and not cdnr_note_col:
            cdnr_note_col = col
        if ('note date' in cl or 'invoice date' in cl) and not cdnr_date_col:
            cdnr_date_col = col
    cdnr_cgst_col, cdnr_sgst_col, cdnr_igst_col = find_tax_amount_columns(cdnr_df)

    if not b2b_gstin_col or not cdnr_gstin_col:
        if log_callback:
            log_callback("Same Month Cancellation: Could not find GSTIN column in B2B or CDNR")
        return pd.DataFrame()

    # ── Build CDNR lookup: norm_GSTIN → list of entries ────────────────
    cdnr_by_gstin = {}
    for _, row in cdnr_df.iterrows():
        gstin = normalize_gstin(str(row.get(cdnr_gstin_col, '')))
        if not gstin:
            continue
        cgst = abs(safe_numeric_conversion(row.get(cdnr_cgst_col, 0) if cdnr_cgst_col else 0))
        sgst = abs(safe_numeric_conversion(row.get(cdnr_sgst_col, 0) if cdnr_sgst_col else 0))
        igst = abs(safe_numeric_conversion(row.get(cdnr_igst_col, 0) if cdnr_igst_col else 0))
        cdnr_by_gstin.setdefault(gstin, []).append({
            'cgst': cgst, 'sgst': sgst, 'igst': igst,
            'note_no': str(row.get(cdnr_note_col, '')) if cdnr_note_col else '',
            'date': str(row.get(cdnr_date_col, '')) if cdnr_date_col else '',
            'raw_gstin': str(row.get(cdnr_gstin_col, '')),
        })

    # ── Match B2B rows against CDNR lookup ─────────────────────────────
    results = []
    for _, b2b_row in b2b_df.iterrows():
        b2b_gstin = normalize_gstin(str(b2b_row.get(b2b_gstin_col, '')))
        if not b2b_gstin:
            continue
        b2b_cgst = abs(safe_numeric_conversion(b2b_row.get(b2b_cgst_col, 0) if b2b_cgst_col else 0))
        b2b_sgst = abs(safe_numeric_conversion(b2b_row.get(b2b_sgst_col, 0) if b2b_sgst_col else 0))
        b2b_igst = abs(safe_numeric_conversion(b2b_row.get(b2b_igst_col, 0) if b2b_igst_col else 0))

        if b2b_cgst == 0 and b2b_sgst == 0 and b2b_igst == 0:
            continue

        for cdnr_entry in cdnr_by_gstin.get(b2b_gstin, []):
            if (abs(cdnr_entry['cgst'] - b2b_cgst) <= tolerance and
                    abs(cdnr_entry['sgst'] - b2b_sgst) <= tolerance and
                    abs(cdnr_entry['igst'] - b2b_igst) <= tolerance):
                results.append({
                    'B2B_GSTIN': str(b2b_row.get(b2b_gstin_col, '')),
                    'B2B_Invoice_No': str(b2b_row.get(b2b_inv_col, '')) if b2b_inv_col else '',
                    'B2B_Invoice_Date': str(b2b_row.get(b2b_date_col, '')) if b2b_date_col else '',
                    'B2B_CGST': b2b_cgst,
                    'B2B_SGST': b2b_sgst,
                    'B2B_IGST': b2b_igst,
                    'CDNR_GSTIN': cdnr_entry['raw_gstin'],
                    'CDNR_Note_No': cdnr_entry['note_no'],
                    'CDNR_Note_Date': cdnr_entry['date'],
                    'CDNR_CGST': cdnr_entry['cgst'],
                    'CDNR_SGST': cdnr_entry['sgst'],
                    'CDNR_IGST': cdnr_entry['igst'],
                })

    result_df = pd.DataFrame(results) if results else pd.DataFrame()
    if log_callback:
        log_callback(f"Same Month Cancellation: Found {len(result_df)} B2B-CDNR matching pairs")
    return result_df


def create_comprehensive_report(itc_df, itc_register, merged_df, comparison_df, log_callback=None):
    """Step 8: Create comprehensive merged report with all tables side by side"""
    if itc_df is None or itc_df.empty:
        if log_callback:
            log_callback("Error: ITC table is empty")
        return pd.DataFrame()

    invoice_col = None
    for col in itc_df.columns:
        col_lower = col.lower().strip()
        if 'vendor inv' in col_lower or 'external doc' in col_lower:
            invoice_col = col
            break

    if not invoice_col:
        if log_callback:
            log_callback("Error: Could not find Invoice column in ITC")
        return pd.DataFrame()

    itc_columns = {}
    for col in itc_df.columns:
        itc_columns[col] = f"ITC_{col}"
    itc_prepared = itc_df.rename(columns=itc_columns)
    itc_invoice_col = f"ITC_{invoice_col}"

    itc_reg_prepared = itc_register.copy()
    itc_reg_columns = {}
    for col in itc_reg_prepared.columns:
        itc_reg_columns[col] = f"as_per_itc_register_{col}"
    itc_reg_prepared = itc_reg_prepared.rename(columns=itc_reg_columns)

    gstr2a_prepared = merged_df.copy()
    gstr2a_columns = {}
    for col in gstr2a_prepared.columns:
        gstr2a_columns[col] = f"as_per_GSTR-2A_{col}"
    gstr2a_prepared = gstr2a_prepared.rename(columns=gstr2a_columns)

    vendor_gstn_col = None
    vendor_inv_col = None
    for col in itc_df.columns:
        col_lower = col.lower().strip()
        if 'vendor' in col_lower and 'gstn' in col_lower:
            vendor_gstn_col = col
        elif 'vendor inv' in col_lower or 'external doc' in col_lower:
            vendor_inv_col = col

    if vendor_gstn_col and vendor_inv_col:
        itc_prepared['_merge_key'] = itc_df[vendor_gstn_col].astype(str) + itc_df[vendor_inv_col].astype(str)

    itc_reg_prepared['_merge_key'] = itc_register['GSTINinvoice']
    gstr2a_prepared['_merge_key'] = merged_df['GSTN'].astype(str) + merged_df['Document_number'].astype(str)

    comprehensive = pd.merge(
        itc_prepared,
        itc_reg_prepared,
        on='_merge_key',
        how='outer'
    )

    comprehensive = pd.merge(
        comprehensive,
        gstr2a_prepared,
        on='_merge_key',
        how='outer'
    )

    if not comparison_df.empty:
        # Include the reconciliation Status as well so comprehensive report follows the same MATCH/MISMATCH decision
        comparison_subset = comparison_df[['GSTINinvoice', 'CGST_Difference', 'SGST_Difference', 'IGST_Difference', 'Status']].copy()
        comparison_subset = comparison_subset.rename(columns={'GSTINinvoice': '_merge_key', 'Status': 'Reconciliation_Status'})
        comprehensive = pd.merge(
            comprehensive,
            comparison_subset,
            on='_merge_key',
            how='left'
        )

    tolerance = 10.0

    # Derive numeric CGST/SGST/IGST from displayed ITC and GSTR columns (prefer Amount columns if present)
    def _find_col(cols, keyword):
        """Return the best column name from cols containing keyword. Prefer columns with 'amount' else the first match."""
        candidates = [c for c in cols if keyword in c.lower()]
        if not candidates:
            return None
        for c in candidates:
            if 'amount' in c.lower():
                return c
        return candidates[0]

    # Build lists of candidate columns (these are defined later originally but we need them now to pick numeric columns)
    itc_cols = [col for col in comprehensive.columns if col.startswith('ITC_')]
    gstr2a_cols = [col for col in comprehensive.columns if col.startswith('as_per_GSTR-2A_')]

    itc_cgst_col = _find_col(itc_cols, 'cgst')
    itc_sgst_col = _find_col(itc_cols, 'sgst')
    itc_igst_col = _find_col(itc_cols, 'igst')

    gstr_cgst_col = _find_col(gstr2a_cols, 'cgst')
    gstr_sgst_col = _find_col(gstr2a_cols, 'sgst')
    gstr_igst_col = _find_col(gstr2a_cols, 'igst')

    # Compute rounded numeric columns to use for determining Status (HALF_UP rounding)
    if itc_igst_col and gstr_igst_col:
        comprehensive['ITC_IGST_num'] = pd.to_numeric(comprehensive[itc_igst_col].astype(str).str.replace(',',''), errors='coerce').fillna(0).apply(lambda x: round_rupee(x))
        comprehensive['GSTR_IGST_num'] = pd.to_numeric(comprehensive[gstr_igst_col].astype(str).str.replace(',',''), errors='coerce').fillna(0).apply(lambda x: round_rupee(x))
        comprehensive['IGST_Diff_num'] = comprehensive['ITC_IGST_num'] - comprehensive['GSTR_IGST_num']
    else:
        comprehensive['IGST_Diff_num'] = comprehensive.get('IGST_Difference', 0)

    if itc_cgst_col and gstr_cgst_col:
        comprehensive['ITC_CGST_num'] = pd.to_numeric(comprehensive[itc_cgst_col].astype(str).str.replace(',',''), errors='coerce').fillna(0).apply(lambda x: round_rupee(x))
        comprehensive['GSTR_CGST_num'] = pd.to_numeric(comprehensive[gstr_cgst_col].astype(str).str.replace(',',''), errors='coerce').fillna(0).apply(lambda x: round_rupee(x))
        comprehensive['CGST_Diff_num'] = comprehensive['ITC_CGST_num'] - comprehensive['GSTR_CGST_num']
    else:
        comprehensive['CGST_Diff_num'] = comprehensive.get('CGST_Difference', 0)

    if itc_sgst_col and gstr_sgst_col:
        comprehensive['ITC_SGST_num'] = pd.to_numeric(comprehensive[itc_sgst_col].astype(str).str.replace(',',''), errors='coerce').fillna(0).apply(lambda x: round_rupee(x))
        comprehensive['GSTR_SGST_num'] = pd.to_numeric(comprehensive[gstr_sgst_col].astype(str).str.replace(',',''), errors='coerce').fillna(0).apply(lambda x: round_rupee(x))
        comprehensive['SGST_Diff_num'] = comprehensive['ITC_SGST_num'] - comprehensive['GSTR_SGST_num']
    else:
        comprehensive['SGST_Diff_num'] = comprehensive.get('SGST_Difference', 0)

    # Prefer differences from comparison_df (invoice-level) when available, otherwise fallback to per-line numeric diffs
    def _choose_diff(row, comp_col, num_col):
        # comp_col is e.g. 'CGST_Difference' inserted from comparison_df; num_col is per-line computed diff
        comp_val = row.get(comp_col, None)
        if pd.notna(comp_val):
            # Round comp_val to nearest rupee using HALF_UP
            try:
                return round_rupee(comp_val)
            except Exception:
                return 0
        # else, use num_col
        return row.get(num_col, 0) if pd.notna(row.get(num_col, 0)) else 0

    def get_detailed_status(row):
        # If reconciliation already determined a MATCH at invoice level, prefer that decision (consistent with comparison_df)
        recon_status = row.get('Reconciliation_Status', None)
        if recon_status and str(recon_status).upper() == 'MATCH':
            return 'MATCH'

        cgst_diff = _choose_diff(row, 'CGST_Difference', 'CGST_Diff_num')
        sgst_diff = _choose_diff(row, 'SGST_Difference', 'SGST_Diff_num')
        igst_diff = _choose_diff(row, 'IGST_Difference', 'IGST_Diff_num')

        cgst_match = abs(cgst_diff) <= tolerance
        sgst_match = abs(sgst_diff) <= tolerance
        igst_match = abs(igst_diff) <= tolerance

        # If both per-row displayed ITC and GSTR numeric values exist and are effectively equal, prefer them as match
        try:
            itc_c = row.get('ITC_CGST_num', None)
            gstr_c = row.get('GSTR_CGST_num', None)
            if itc_c is not None and gstr_c is not None and abs(itc_c - gstr_c) <= tolerance:
                cgst_match = True
                cgst_diff = 0
        except Exception:
            pass
        try:
            itc_s = row.get('ITC_SGST_num', None)
            gstr_s = row.get('GSTR_SGST_num', None)
            if itc_s is not None and gstr_s is not None and abs(itc_s - gstr_s) <= tolerance:
                sgst_match = True
                sgst_diff = 0
        except Exception:
            pass
        try:
            itc_i = row.get('ITC_IGST_num', None)
            gstr_i = row.get('GSTR_IGST_num', None)
            if itc_i is not None and gstr_i is not None and abs(itc_i - gstr_i) <= tolerance:
                igst_match = True
                igst_diff = 0
        except Exception:
            pass

        if cgst_match and sgst_match and igst_match:
            return 'MATCH'

        mismatches = []
        if not cgst_match:
            if cgst_diff > 0:
                mismatches.append('CGST Lower in 2B')
            else:
                mismatches.append('CGST Higher in 2B')
        if not sgst_match:
            if sgst_diff > 0:
                mismatches.append('SGST Lower in 2B')
            else:
                mismatches.append('SGST Higher in 2B')
        if not igst_match:
            if igst_diff > 0:
                mismatches.append('IGST Lower in 2B')
            else:
                mismatches.append('IGST Higher in 2B')

        if len(mismatches) > 1:
            return 'MISMATCH'
        elif len(mismatches) == 1:
            return mismatches[0]

        return 'MISMATCH'
    comprehensive['Status'] = comprehensive.apply(get_detailed_status, axis=1)

    comprehensive = comprehensive.rename(columns={
        'CGST_Difference': 'CGST',
        'SGST_Difference': 'SGST',
        'IGST_Difference': 'IGST'
    })

    if '_merge_key' in comprehensive.columns:
        comprehensive = comprehensive.drop(columns=['_merge_key'])

    itc_cols = [col for col in comprehensive.columns if col.startswith('ITC_')]
    itc_reg_cols = [col for col in comprehensive.columns if col.startswith('as_per_itc_register_')]
    gstr2a_cols = [col for col in comprehensive.columns if col.startswith('as_per_GSTR-2A_')]
    status_cols = ['Status', 'CGST', 'SGST', 'IGST']
    status_cols = [col for col in status_cols if col in comprehensive.columns]

    ordered_cols = itc_cols + itc_reg_cols + gstr2a_cols + status_cols
    remaining_cols = [col for col in comprehensive.columns if col not in ordered_cols]
    ordered_cols = ordered_cols + remaining_cols

    comprehensive = comprehensive[ordered_cols]

    if log_callback:
        log_callback(f"Step 8: Created comprehensive report with {len(comprehensive)} records")

    return comprehensive


def _autofit_ws(ws, max_width=55):
    """Set each column width to fit its widest cell value."""
    from openpyxl.utils import get_column_letter
    for col_cells in ws.columns:
        max_len = max(
            (len(str(c.value)) if c.value is not None else 0 for c in col_cells),
            default=8,
        )
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 3, max_width)


def to_excel(df):
    """Convert dataframe to Excel file for download"""
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Reconciliation')
        _autofit_ws(writer.sheets['Reconciliation'])
    output.seek(0)
    return output.getvalue()


def to_excel_with_highlight(df):
    """Convert dataframe to Excel file with mismatch highlighting"""
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Comprehensive_Report')

        workbook = writer.book
        worksheet = writer.sheets['Comprehensive_Report']

        from openpyxl.styles import PatternFill
        red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')

        status_col_idx = None
        for idx, col in enumerate(df.columns, 1):
            if col == 'Status':
                status_col_idx = idx
                break

        if status_col_idx:
            for row_idx, row in enumerate(df.itertuples(), 2):
                status_value = getattr(row, 'Status', None)
                if status_value and status_value != 'MATCH':
                    for col_idx in range(1, len(df.columns) + 1):
                        worksheet.cell(row=row_idx, column=col_idx).fill = red_fill

        _autofit_ws(worksheet)

    output.seek(0)
    return output.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# YTD Database Manager
# ─────────────────────────────────────────────────────────────────────────────

class GSTDatabaseManager:
    """Manages a SQLite database for year-to-date GST invoice tracking.
    ITC and 2A invoices are stored in separate tables (itc_invoices / twoa_invoices).
    """

    def __init__(self, db_path):
        self.db_path = db_path
        self.conn = sqlite3.connect(db_path, check_same_thread=False)
        self.conn.row_factory = sqlite3.Row
        self._create_tables()

    def _create_tables(self):
        self.conn.executescript("""
            CREATE TABLE IF NOT EXISTS itc_invoices (
                id                   INTEGER PRIMARY KEY AUTOINCREMENT,
                gstin                TEXT,
                invoice_number       TEXT,
                invoice_norm         TEXT,
                invoice_date         TEXT,
                invoice_month        INTEGER,
                invoice_year         INTEGER,
                taxable_value        REAL DEFAULT 0,
                cgst                 REAL DEFAULT 0,
                sgst                 REAL DEFAULT 0,
                igst                 REAL DEFAULT 0,
                source_type          TEXT,
                status               TEXT DEFAULT 'Unmatched',
                matched_invoice      TEXT DEFAULT '',
                matched_gstin        TEXT DEFAULT '',
                saved_date           TEXT,
                updated_at           TEXT
            );
            CREATE TABLE IF NOT EXISTS twoa_invoices (
                id                   INTEGER PRIMARY KEY AUTOINCREMENT,
                gstin                TEXT,
                invoice_number       TEXT,
                invoice_norm         TEXT,
                invoice_date         TEXT,
                invoice_month        INTEGER,
                invoice_year         INTEGER,
                taxable_value        REAL DEFAULT 0,
                cgst                 REAL DEFAULT 0,
                sgst                 REAL DEFAULT 0,
                igst                 REAL DEFAULT 0,
                source_type          TEXT,
                status               TEXT DEFAULT 'Unmatched',
                matched_invoice      TEXT DEFAULT '',
                matched_gstin        TEXT DEFAULT '',
                saved_date           TEXT,
                updated_at           TEXT
            );
        """)
        # Migrate old gst_invoices table if it exists
        try:
            self.conn.execute("SELECT 1 FROM gst_invoices LIMIT 1")
            self.conn.executescript("""
                INSERT OR IGNORE INTO itc_invoices
                    (gstin, invoice_number, invoice_norm, invoice_date,
                     invoice_month, invoice_year, taxable_value, cgst, sgst, igst,
                     source_type, status, matched_invoice, matched_gstin,
                     saved_date, updated_at)
                SELECT gstin, invoice_number, invoice_norm, invoice_date,
                       invoice_month, invoice_year, taxable_value, cgst, sgst, igst,
                       source_type, status, matched_invoice, matched_gstin,
                       reconciliation_period, updated_at
                FROM gst_invoices WHERE record_type='ITC';

                INSERT OR IGNORE INTO twoa_invoices
                    (gstin, invoice_number, invoice_norm, invoice_date,
                     invoice_month, invoice_year, taxable_value, cgst, sgst, igst,
                     source_type, status, matched_invoice, matched_gstin,
                     saved_date, updated_at)
                SELECT gstin, invoice_number, invoice_norm, invoice_date,
                       invoice_month, invoice_year, taxable_value, cgst, sgst, igst,
                       source_type, status, matched_invoice, matched_gstin,
                       reconciliation_period, updated_at
                FROM gst_invoices WHERE record_type='2A';

                DROP TABLE gst_invoices;
            """)
        except sqlite3.OperationalError:
            pass
        # Per-type raw data table (template-structured columns stored as JSON)
        self.conn.executescript("""
            CREATE TABLE IF NOT EXISTS ytd_raw_data (
                id            INTEGER PRIMARY KEY AUTOINCREMENT,
                data_type     TEXT,
                invoice_year  INTEGER DEFAULT 0,
                invoice_month INTEGER DEFAULT 0,
                is_frozen     INTEGER DEFAULT 0,
                row_json      TEXT DEFAULT '{}',
                saved_date    TEXT
            );
        """)
        self.conn.commit()

        # Add extra columns if missing (migration for existing DBs)
        for tbl in ('itc_invoices', 'twoa_invoices'):
            for col_def in [
                "is_frozen INTEGER DEFAULT 0",
                "freeze_note TEXT DEFAULT ''",
                "vendor_name TEXT DEFAULT ''",
                "invoice_value REAL DEFAULT 0",
                "tax_rate TEXT DEFAULT ''",
                "source_type2 TEXT DEFAULT ''",
                "extra_data TEXT DEFAULT '{}'"
            ]:
                col_name = col_def.split()[0]
                try:
                    self.conn.execute(f"ALTER TABLE {tbl} ADD COLUMN {col_def}")
                except sqlite3.OperationalError:
                    pass
        self.conn.commit()

    # ── Raw (template-structured) row storage ─────────────────────────────────

    def save_raw_rows(self, rows_by_type: dict, saved_date: str):
        """Save per-type original rows to ytd_raw_data (JSON-serialised)."""
        c = self.conn.cursor()
        for data_type, row_list in rows_by_type.items():
            for row_dict, yr, mo in row_list:
                c.execute(
                    "INSERT INTO ytd_raw_data "
                    "(data_type, invoice_year, invoice_month, row_json, saved_date) "
                    "VALUES (?,?,?,?,?)",
                    (data_type, yr or 0, mo or 0,
                     json.dumps(row_dict, default=str), saved_date))
        self.conn.commit()

    def has_raw_data(self) -> bool:
        try:
            return bool(
                self.conn.execute("SELECT 1 FROM ytd_raw_data LIMIT 1").fetchone())
        except Exception:
            return False

    def get_raw_rows(self, year=None, month=None):
        """Return all ytd_raw_data rows (optionally filtered by year/month)."""
        q = "SELECT data_type, invoice_year, invoice_month, is_frozen, row_json FROM ytd_raw_data WHERE 1=1"
        params = []
        if year:
            q += " AND invoice_year=?"; params.append(year)
        if month:
            q += " AND invoice_month=?"; params.append(month)
        return self.conn.execute(q, params).fetchall()

    def get_raw_rows_by_type(self, data_type, year=None, month=None):
        """Return parsed JSON rows from ytd_raw_data for a specific data_type."""
        q = ("SELECT id, invoice_year, invoice_month, is_frozen, row_json, saved_date "
             "FROM ytd_raw_data WHERE data_type=?")
        params = [data_type]
        if year:
            q += " AND invoice_year=?"; params.append(year)
        if month:
            q += " AND invoice_month=?"; params.append(month)
        result = []
        for row in self.conn.execute(q, params).fetchall():
            try:
                d = json.loads(row['row_json'] or '{}')
            except Exception:
                d = {}
            d['__id__']         = row['id']
            d['__frozen__']     = row['is_frozen']
            d['__saved_date__'] = (row['saved_date'] or '')[:10]
            result.append(d)
        return result

    # ── Freeze / unfreeze ─────────────────────────────────────────────────────

    def freeze_month(self, year, month, note=''):
        """Mark all records for the given year/month as frozen."""
        now = self._now()
        for tbl in ('itc_invoices', 'twoa_invoices'):
            self.conn.execute(
                f"UPDATE {tbl} SET is_frozen=1, freeze_note=?, updated_at=? "
                f"WHERE invoice_year=? AND invoice_month=?",
                (note, now, year, month))
        self.conn.execute(
            "UPDATE ytd_raw_data SET is_frozen=1 WHERE invoice_year=? AND invoice_month=?",
            (year, month))
        self.conn.commit()

    def unfreeze_month(self, year, month):
        """Unfreeze all records for the given year/month."""
        now = self._now()
        for tbl in ('itc_invoices', 'twoa_invoices'):
            self.conn.execute(
                f"UPDATE {tbl} SET is_frozen=0, freeze_note='', updated_at=? "
                f"WHERE invoice_year=? AND invoice_month=?",
                (now, year, month))
        self.conn.execute(
            "UPDATE ytd_raw_data SET is_frozen=0 WHERE invoice_year=? AND invoice_month=?",
            (year, month))
        self.conn.commit()

    def is_month_frozen(self, year, month):
        row = self.conn.execute(
            "SELECT COUNT(*) FROM itc_invoices WHERE invoice_year=? AND invoice_month=? AND is_frozen=1",
            (year, month)).fetchone()
        return (row[0] > 0) if row else False

    def _now(self):
        return datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    def bulk_insert(self, records):
        now = self._now()
        itc_rows  = [r for r in records if r.get('record_type') == 'ITC']
        twoa_rows = [r for r in records if r.get('record_type') == '2A']

        _sql = """
            INSERT INTO {tbl}
            (gstin, invoice_number, invoice_norm, invoice_date,
             invoice_month, invoice_year, taxable_value, cgst, sgst, igst,
             source_type, status, matched_invoice, matched_gstin,
             vendor_name, invoice_value, tax_rate, extra_data,
             saved_date, updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """
        def _params(r):
            extra = r.get('extra_data', {})
            return (
                r.get('gstin', ''),
                r.get('invoice_number', ''),
                r.get('invoice_norm', ''),
                r.get('invoice_date', ''),
                r.get('invoice_month'),
                r.get('invoice_year'),
                float(r.get('taxable_value', 0) or 0),
                float(r.get('cgst', 0) or 0),
                float(r.get('sgst', 0) or 0),
                float(r.get('igst', 0) or 0),
                r.get('source_type', ''),
                r.get('status', 'Unmatched'),
                r.get('matched_invoice', ''),
                r.get('matched_gstin', ''),
                r.get('vendor_name', ''),
                float(r.get('invoice_value', 0) or 0),
                str(r.get('tax_rate', '')),
                json.dumps(extra) if isinstance(extra, dict) else str(extra),
                now, now,
            )
        if itc_rows:
            self.conn.executemany(_sql.format(tbl='itc_invoices'),  [_params(r) for r in itc_rows])
        if twoa_rows:
            self.conn.executemany(_sql.format(tbl='twoa_invoices'), [_params(r) for r in twoa_rows])
        self.conn.commit()

    def get_invoices(self, record_type=None, year=None, month=None):
        if record_type == 'ITC':
            tables = [('ITC', 'itc_invoices')]
        elif record_type == '2A':
            tables = [('2A', 'twoa_invoices')]
        else:
            tables = [('ITC', 'itc_invoices'), ('2A', 'twoa_invoices')]
        all_rows = []
        for rt, tbl in tables:
            q = f"SELECT *, '{rt}' as record_type, '{tbl}' as _table FROM {tbl} WHERE 1=1"
            p = []
            if year:
                q += " AND invoice_year=?"; p.append(year)
            if month is not None:
                q += " AND invoice_month=?"; p.append(month)
            q += " ORDER BY invoice_date, invoice_number"
            all_rows.extend([dict(r) for r in self.conn.execute(q, p).fetchall()])
        return all_rows

    def update_status(self, row_id, table, status, matched_invoice='', matched_gstin=''):
        now = self._now()
        self.conn.execute(f"""
            UPDATE {table}
            SET status=?, matched_invoice=?, matched_gstin=?, saved_date=?, updated_at=?
            WHERE id=?
        """, (status, matched_invoice, matched_gstin, now, now, row_id))
        self.conn.commit()

    def upsert_matched_pair(self, itc_rec, twoa_rec):
        """Insert ITC/2A record if not present, then mark both as Matched."""
        for rec in (itc_rec, twoa_rec):
            tbl = 'itc_invoices' if rec.get('record_type') == 'ITC' else 'twoa_invoices'
            existing = self.conn.execute(
                f"SELECT id FROM {tbl} WHERE invoice_norm=?", (rec['invoice_norm'],)
            ).fetchone()
            now = self._now()
            if existing:
                self.update_status(existing['id'], tbl, rec['status'],
                                   rec.get('matched_invoice', ''), rec.get('matched_gstin', ''))
            else:
                self.conn.execute(f"""
                    INSERT INTO {tbl}
                    (gstin, invoice_number, invoice_norm, invoice_date, invoice_month,
                     invoice_year, taxable_value, cgst, sgst, igst, source_type, status,
                     matched_invoice, matched_gstin, saved_date, updated_at)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (rec['gstin'], rec['invoice_number'], rec['invoice_norm'],
                      rec.get('invoice_date', ''), rec.get('invoice_month'),
                      rec.get('invoice_year'), rec.get('taxable_value', 0),
                      rec.get('cgst', 0), rec.get('sgst', 0), rec.get('igst', 0),
                      rec.get('source_type', ''), rec['status'],
                      rec.get('matched_invoice', ''), rec.get('matched_gstin', ''),
                      now, now))
                self.conn.commit()

    def delete_invoice(self, row_id, table):
        self.conn.execute(
            f"DELETE FROM {table} WHERE id=? AND status != 'Matched'", (row_id,))
        self.conn.commit()

    def get_distinct_periods(self):
        rows = self.conn.execute("""
            SELECT DISTINCT invoice_year, invoice_month FROM itc_invoices
            WHERE invoice_year IS NOT NULL
            UNION
            SELECT DISTINCT invoice_year, invoice_month FROM twoa_invoices
            WHERE invoice_year IS NOT NULL
            ORDER BY invoice_year, invoice_month
        """).fetchall()
        return [(r[0], r[1]) for r in rows]

    def is_norm_key_matched(self, invoice_norm, record_type='2A'):
        tbl = 'twoa_invoices' if record_type == '2A' else 'itc_invoices'
        row = self.conn.execute(
            f"SELECT id FROM {tbl} WHERE invoice_norm=? AND status='Matched'",
            (invoice_norm,)).fetchone()
        return row is not None

    def get_unmatched_2a(self):
        """Return all unmatched 2A records for cross-period matching."""
        return [dict(r) for r in self.conn.execute(
            "SELECT * FROM twoa_invoices WHERE status='Unmatched' "
            "ORDER BY gstin, invoice_number"
        ).fetchall()]

    def get_unmatched_itc(self):
        """Return all unmatched ITC records for reverse cross-period matching."""
        return [dict(r) for r in self.conn.execute(
            "SELECT * FROM itc_invoices WHERE status='Unmatched' "
            "ORDER BY gstin, invoice_number"
        ).fetchall()]

    def close(self):
        if self.conn:
            self.conn.close()
            self.conn = None


class GSTReconciliationApp(ctk.CTk):
    def __init__(self):
        super().__init__()

        # Hide the main window until license is verified
        self.withdraw()
        # Schedule license check after the event loop starts
        self.after(10, self._check_license_on_start)

        # Window setup
        self.title("GST Reconciliation Tool")
        self.geometry("1200x800")
        self.minsize(1000, 700)

        # Try to set icon
        try:
            ico_path = get_resource_path("app_icon.ico")
            if os.path.exists(ico_path):
                self.iconbitmap(ico_path)
            else:
                icon_path = get_resource_path("logo small.png")
                if os.path.exists(icon_path):
                    icon_image = Image.open(icon_path)
                    self.iconphoto(True, ctk.CTkImage(light_image=icon_image, size=(32, 32))._light_image)
        except Exception:
            pass

        # File paths storage
        self.file_paths = {
            'B2B': None,
            'B2BA': None,
            'CDNR': None,
            'CDNRA': None,
            'IMPG': None,
            'IMPGSEZ': None,
            'ITC': None
        }

        # Results storage
        self.comparison_df = None
        self.itc_result_df = None
        self.unmatched_2a_df = None
        self.gstr_2a_results_df = None
        self.merged_df = None
        self.original_itc = None
        self.itc_register = None
        self.same_month_cancel_df = None
        self.b2b_processed_df = None
        self.cdnr_processed_df = None
        self._original_itc_raw = pd.DataFrame()
        self._original_2b_tables = {}

        # YTD Database
        self.db_manager = None
        self.db_name = None

        # Debug matching — pending matches to save on next YTD save
        self._debug_final_matches = []

        # Upload mode: 'template' (default) or 'csv'
        self._upload_mode = 'template'
        self._template_upload_frame = None
        self._csv_upload_frame = None

        # Company info
        self._company_name = ''
        self._company_gst = ''
        self._company_period = ''
        self._company_info_lbl = None
        self._all_companies = []   # list of dicts with name/gst/period
        self._load_company_info()

        # Create main layout
        self.create_widgets()

    # ── License check ─────────────────────────────────────────────────────────
    # ── Company info helpers ──────────────────────────────────────────────────
    def _get_company_config_path(self):
        config_dir = os.path.join(
            os.environ.get('APPDATA', os.path.expanduser('~')),
            'GST_Reconciliation_Tool')
        os.makedirs(config_dir, exist_ok=True)
        return os.path.join(config_dir, 'company_config.json')

    def _load_company_info(self):
        try:
            path = self._get_company_config_path()
            if os.path.exists(path):
                with open(path, 'r') as f:
                    data = json.load(f)
                # Support both legacy single-company and new multi-company formats
                if 'companies' in data:
                    self._all_companies = data['companies']
                    active = data.get('active', {})
                    self._company_name   = active.get('name', '')
                    self._company_gst    = active.get('gst', '')
                    self._company_period = active.get('period', '')
                else:
                    # Migrate legacy format
                    self._company_name = data.get('company_name', '')
                    self._company_gst  = data.get('company_gst', '')
                    self._company_period = ''
                    if self._company_name:
                        self._all_companies = [{'name': self._company_name,
                                                'gst': self._company_gst,
                                                'period': ''}]
        except Exception:
            pass

    def _save_company_info_to_disk(self):
        try:
            path = self._get_company_config_path()
            # Upsert current company into the list
            entry = {'name': self._company_name,
                     'gst':  self._company_gst,
                     'period': self._company_period}
            existing = [c for c in self._all_companies
                        if c.get('gst', '').upper() == self._company_gst.upper()
                        and c.get('name', '').lower() == self._company_name.lower()]
            if existing:
                existing[0].update(entry)
            else:
                self._all_companies.append(entry)
            with open(path, 'w') as f:
                json.dump({'companies': self._all_companies,
                           'active': entry}, f, indent=2)
        except Exception:   
            pass

    def _update_company_display(self):
        if self._company_info_lbl:
            parts = [p for p in [self._company_name, self._company_gst,
                                  self._company_period] if p]
            text = "  |  ".join(parts) if parts else "No company selected — click Edit to set"
            self._company_info_lbl.configure(text=text)

    def open_help_dialog(self):
        """Show help dialog with usage info and contact details."""
        import webbrowser
        dlg = ctk.CTkToplevel(self)
        dlg.title("Help & Support")
        dlg.geometry("520x600")
        dlg.grab_set()
        dlg.lift()
        dlg.focus_force()

        # Header
        hdr = ctk.CTkFrame(dlg, fg_color="#5C6BC0", corner_radius=0)
        hdr.pack(fill="x")
        ctk.CTkLabel(hdr, text="Help & Support",
                     font=ctk.CTkFont(size=16, weight="bold"),
                     text_color="white").pack(pady=14)

        scroll = ctk.CTkScrollableFrame(dlg, fg_color="white")
        scroll.pack(fill="both", expand=True, padx=0, pady=0)

        sections = [
            ("Getting Started",
             "1. Download the template from the dashboard.\n"
             "2. Fill in your GST data in the provided sheets (B2B-2b, ITC-SR, etc.).\n"
             "3. Upload the filled Excel file using 'Browse & Upload Excel'.\n"
             "4. Click 'Process and Reconcile' to match your ITC with GSTR-2B.\n"
             "5. Review results and download the report."),
            ("Template Mode vs CSV Mode",
             "Template Mode (default): Upload a single Excel file containing all the\n"
             "required sheets. This is the recommended approach.\n\n"
             "CSV Mode: Upload individual CSV files for each sheet separately.\n"
             "Use this if you have exported CSV files from the GST portal."),
            ("Understanding Results",
             "Matched: Invoice found in both ITC and GSTR-2B with matching amounts.\n"
             "Higher in 2B: 2B shows a higher tax amount than your ITC entry.\n"
             "Lower in 2B: 2B shows a lower tax amount than your ITC entry.\n"
             "Not found in 2B: Your ITC entry has no corresponding entry in GSTR-2B.\n"
             "Unmatched: Entry in GSTR-2B with no corresponding ITC record."),
            ("Debug Matching",
             "Use 'Debug Matching' to manually review pairs where the invoice numbers\n"
             "don't match exactly but the GSTIN and tax amounts are the same.\n"
             "Use 'GSTN Debug' to find entries where the GSTIN may be incorrect,\n"
             "causing mismatches even though the invoice and amounts match."),
            ("YTD Database",
             "Use 'Save to YTD Database' to store monthly reconciliation results for\n"
             "year-to-date tracking. You can later open the YTD Database to browse,\n"
             "filter, and download historical records."),
        ]
        for heading, text in sections:
            ctk.CTkLabel(scroll, text=heading,
                         font=ctk.CTkFont(size=13, weight="bold"),
                         text_color="#333333", anchor="w").pack(anchor="w", padx=20, pady=(14, 2))
            ctk.CTkFrame(scroll, fg_color="#5C6BC0", height=2).pack(fill="x", padx=20, pady=(0, 6))
            ctk.CTkLabel(scroll, text=text,
                         font=ctk.CTkFont(size=12), text_color="#555555",
                         anchor="w", justify="left", wraplength=460).pack(anchor="w", padx=24, pady=(0, 4))

        # Contact section
        ctk.CTkFrame(scroll, fg_color="#E8EAF6", height=2).pack(fill="x", padx=20, pady=(18, 6))
        ctk.CTkLabel(scroll, text="Contact Support",
                     font=ctk.CTkFont(size=13, weight="bold"),
                     text_color="#333333").pack(pady=(6, 2))
        ctk.CTkLabel(scroll,
                     text="Monday to Saturday  •  10:00 AM – 7:00 PM",
                     font=ctk.CTkFont(size=11), text_color="#9E9E9E").pack(pady=(0, 10))

        contact_row = ctk.CTkFrame(scroll, fg_color="transparent")
        contact_row.pack(pady=(0, 20))
        ctk.CTkButton(contact_row, text="✉  tech@gscintime.com",
                      command=lambda: webbrowser.open("mailto:tech@gscintime.com"),
                      fg_color="#1565C0", hover_color="#0D47A1",
                      font=ctk.CTkFont(size=12), height=36, width=220,
                      corner_radius=8).pack(side="left", padx=8)
        ctk.CTkButton(contact_row, text="📞  022 46725611",
                      command=lambda: webbrowser.open("tel:02246725611"),
                      fg_color="#2E7D32", hover_color="#1B5E20",
                      font=ctk.CTkFont(size=12), height=36, width=180,
                      corner_radius=8).pack(side="left", padx=8)

        ctk.CTkButton(dlg, text="Close", command=dlg.destroy,
                      fg_color="#757575", hover_color="#616161",
                      height=34, width=120).pack(pady=10)

    def _prompt_company_info(self, parent=None, prefill=None):
        """Show dialog to add / edit company info (name, GSTIN, period)."""
        pre = prefill or {}
        dlg = ctk.CTkToplevel(parent or self)
        dlg.title("Company Information")
        dlg.geometry("460x360")
        dlg.grab_set(); dlg.lift(); dlg.focus_force()

        ctk.CTkLabel(dlg, text="Company Information",
                     font=ctk.CTkFont(size=16, weight="bold")).pack(pady=(20, 2))
        ctk.CTkLabel(dlg, text="This information appears on all reports and windows.",
                     font=ctk.CTkFont(size=11), text_color="#9E9E9E").pack(pady=(0, 14))

        name_var   = tk.StringVar(value=pre.get('name', self._company_name))
        gst_var    = tk.StringVar(value=pre.get('gst',  self._company_gst))
        period_var = tk.StringVar(value=pre.get('period', self._company_period))

        form = ctk.CTkFrame(dlg, fg_color="transparent")
        form.pack(fill="x", padx=36)

        for label, var, placeholder in [
            ("Company Name:", name_var, "e.g. ABC Pvt. Ltd."),
            ("GST Number (GSTIN):", gst_var, "e.g. 27AABCU9603R1ZM"),
        ]:
            ctk.CTkLabel(form, text=label, anchor="w",
                         font=ctk.CTkFont(size=12)).pack(anchor="w", pady=(0, 2))
            ctk.CTkEntry(form, textvariable=var, placeholder_text=placeholder,
                         height=34, font=ctk.CTkFont(size=12)).pack(fill="x", pady=(0, 10))
        ctk.CTkLabel(form, text="Period:", anchor="w",
                     font=ctk.CTkFont(size=12)).pack(anchor="w", pady=(0, 2))
        _build_period_pickers(form, period_var)

        saved = [False]

        def _save():
            self._company_name   = name_var.get().strip()
            self._company_gst    = gst_var.get().strip().upper()
            self._company_period = period_var.get().strip()
            self._save_company_info_to_disk()
            self._update_company_display()
            saved[0] = True
            dlg.destroy()

        bf = ctk.CTkFrame(dlg, fg_color="transparent")
        bf.pack(pady=6)
        ctk.CTkButton(bf, text="Save", command=_save,
                      fg_color=THEME_PRIMARY, hover_color=THEME_HOVER,
                      width=120).pack(side="left", padx=8)
        ctk.CTkButton(bf, text="Cancel", command=dlg.destroy,
                      fg_color="#757575", hover_color="#616161",
                      width=100).pack(side="left", padx=8)

        dlg.wait_window()
        return saved[0]

    def _show_company_selection(self):
        """Show company selection screen on startup.  Returns True when a company is chosen."""
        dlg = ctk.CTkToplevel(self)
        dlg.title("Select Company")
        dlg.geometry("500x460")
        dlg.grab_set(); dlg.lift(); dlg.focus_force()
        dlg.protocol("WM_DELETE_WINDOW", self.quit)  # close app if user X's out

        chosen = [False]

        hdr = ctk.CTkFrame(dlg, fg_color=THEME_DARK, corner_radius=0)
        hdr.pack(fill="x")
        ctk.CTkLabel(hdr, text="GST Reconciliation Tool",
                     font=ctk.CTkFont(size=16, weight="bold"),
                     text_color="white").pack(pady=12)
        ctk.CTkLabel(dlg, text="Select a company to work with, or add a new one.",
                     font=ctk.CTkFont(size=12), text_color="#555555").pack(pady=(12, 4))

        scroll = ctk.CTkScrollableFrame(dlg, fg_color="#F5F5F5", height=220)
        scroll.pack(fill="x", padx=16, pady=4)

        def _pick(co):
            self._company_name   = co.get('name', '')
            self._company_gst    = co.get('gst',  '')
            self._company_period = co.get('period', '')
            self._save_company_info_to_disk()
            self._update_company_display()
            chosen[0] = True
            dlg.destroy()

        def _refresh():
            for w in scroll.winfo_children():
                w.destroy()
            if not self._all_companies:
                ctk.CTkLabel(scroll, text="No companies yet. Click '+ Add Company' below.",
                             font=ctk.CTkFont(size=11), text_color="#9E9E9E").pack(pady=20)
                return
            for co in self._all_companies:
                row = ctk.CTkFrame(scroll, fg_color="white", corner_radius=6,
                                   border_width=1, border_color="#E0E0E0")
                row.pack(fill="x", pady=3, padx=4)
                line1 = co.get('name', '—')
                line2 = "  ·  ".join(p for p in [co.get('gst',''), co.get('period','')] if p)
                ctk.CTkLabel(row, text=line1, anchor="w",
                             font=ctk.CTkFont(size=13, weight="bold"),
                             text_color="#1B5E20").pack(anchor="w", padx=12, pady=(6, 0))
                if line2:
                    ctk.CTkLabel(row, text=line2, anchor="w",
                                 font=ctk.CTkFont(size=11), text_color="#555555").pack(anchor="w", padx=12, pady=(0, 4))
                ctk.CTkButton(row, text="Select →",
                              command=lambda c=co: _pick(c),
                              fg_color=THEME_PRIMARY, hover_color=THEME_HOVER,
                              height=28, width=100).pack(side="right", padx=10, pady=6)

        _refresh()

        def _add_new():
            # Open add-company dialog and refresh list
            new_dlg = ctk.CTkToplevel(dlg)
            new_dlg.title("Add Company"); new_dlg.geometry("460x360")
            new_dlg.grab_set(); new_dlg.lift(); new_dlg.focus_force()
            nv = tk.StringVar(); gv = tk.StringVar(); pv = tk.StringVar()
            f = ctk.CTkFrame(new_dlg, fg_color="transparent"); f.pack(padx=36, pady=20)
            for lbl, var, ph in [("Company Name:", nv, "e.g. ABC Pvt. Ltd."),
                                  ("GSTIN:", gv, "e.g. 27AABCU9603R1ZM")]:
                ctk.CTkLabel(f, text=lbl, anchor="w",
                             font=ctk.CTkFont(size=12)).pack(anchor="w", pady=(0,2))
                ctk.CTkEntry(f, textvariable=var, placeholder_text=ph,
                             height=34).pack(fill="x", pady=(0,10))
            ctk.CTkLabel(f, text="Period:", anchor="w",
                         font=ctk.CTkFont(size=12)).pack(anchor="w", pady=(0,2))
            _build_period_pickers(f, pv)
            def _save_new():
                n = nv.get().strip(); g = gv.get().strip().upper(); p = pv.get().strip()
                if not n:
                    messagebox.showwarning("Required", "Company name is required.", parent=new_dlg)
                    return
                self._all_companies.append({'name': n, 'gst': g, 'period': p})
                self._company_name = n; self._company_gst = g; self._company_period = p
                self._save_company_info_to_disk()
                new_dlg.destroy()
                _refresh()
            bf2 = ctk.CTkFrame(new_dlg, fg_color="transparent"); bf2.pack(pady=8)
            ctk.CTkButton(bf2, text="Save", command=_save_new,
                          fg_color=THEME_PRIMARY, width=120).pack(side="left", padx=8)
            ctk.CTkButton(bf2, text="Cancel", command=new_dlg.destroy,
                          fg_color="#757575", width=100).pack(side="left", padx=8)

        btn_row = ctk.CTkFrame(dlg, fg_color="transparent")
        btn_row.pack(pady=8)
        ctk.CTkButton(btn_row, text="+ Add Company", command=_add_new,
                      fg_color="#1565C0", hover_color="#0D47A1",
                      height=34, width=160).pack(side="left", padx=8)

        dlg.wait_window()
        return chosen[0]

    # ── Upload mode toggle ────────────────────────────────────────────────────
    def _set_upload_mode(self, mode):
        self._upload_mode = mode
        if mode == 'template':
            if self._csv_upload_frame:
                self._csv_upload_frame.pack_forget()
            if self._template_upload_frame:
                self._template_upload_frame.pack(fill="x", padx=12, pady=8)
            if hasattr(self, '_mode_template_btn'):
                self._mode_template_btn.configure(fg_color=THEME_PRIMARY, text_color="white")
                self._mode_csv_btn.configure(fg_color="#E0E0E0", text_color="#555555")
        else:
            if self._template_upload_frame:
                self._template_upload_frame.pack_forget()
            if self._csv_upload_frame:
                self._csv_upload_frame.pack(fill="x", padx=12, pady=8)
            if hasattr(self, '_mode_template_btn'):
                self._mode_csv_btn.configure(fg_color=THEME_PRIMARY, text_color="white")
                self._mode_template_btn.configure(fg_color="#E0E0E0", text_color="#555555")

    # ── License check ─────────────────────────────────────────────────────────
    def _check_license_on_start(self):
        """Verify license, then show company selection before opening the main window."""
        if not license_manager.is_activated():
            dlg = license_manager.ActivationDialog(self)
            self.wait_window(dlg)
            if not dlg.activated:
                self.quit()
                return
        self.deiconify()
        # First-time setup: ask user to create YTD credentials if not done yet
        if not license_manager.has_ytd_credentials():
            self._setup_ytd_credentials()
        # Always show company selection (even if one is already active)
        self.after(200, self._show_company_selection_or_prompt)

    def _setup_ytd_credentials(self):
        """Show a dialog to create YTD username and password (first-time setup)."""
        dlg = ctk.CTkToplevel(self)
        dlg.title("Set Up YTD Security")
        dlg.geometry("420x380")
        dlg.grab_set(); dlg.lift(); dlg.focus_force()
        dlg.resizable(False, False)

        hdr = ctk.CTkFrame(dlg, fg_color=THEME_DARK, corner_radius=0)
        hdr.pack(fill="x")
        ctk.CTkLabel(hdr, text="Create YTD Security Credentials",
                     font=ctk.CTkFont(size=15, weight="bold"),
                     text_color="white").pack(pady=14)

        body = ctk.CTkScrollableFrame(dlg, fg_color="white", corner_radius=0)
        body.pack(fill="both", expand=True, padx=0, pady=0)

        ctk.CTkLabel(body,
                     text="Create a username and password to protect the YTD month-unfreeze action.\n"
                          "You will need these credentials each time you unfreeze a frozen month.",
                     font=ctk.CTkFont(size=11), text_color="#555",
                     wraplength=360, justify="center").pack(pady=(18, 14), padx=20)

        ctk.CTkLabel(body, text="Username", font=ctk.CTkFont(size=12, weight="bold"),
                     anchor="w").pack(fill="x", padx=30, pady=(0, 2))
        uv = tk.StringVar()
        uentry = ctk.CTkEntry(body, textvariable=uv, width=300, height=36,
                              placeholder_text="Enter username")
        uentry.pack(padx=30, pady=(0, 10))

        ctk.CTkLabel(body, text="Password", font=ctk.CTkFont(size=12, weight="bold"),
                     anchor="w").pack(fill="x", padx=30, pady=(0, 2))
        pv = tk.StringVar()
        pentry = ctk.CTkEntry(body, textvariable=pv, width=300, height=36,
                              show="*", placeholder_text="Enter password")
        pentry.pack(padx=30, pady=(0, 10))

        ctk.CTkLabel(body, text="Confirm Password", font=ctk.CTkFont(size=12, weight="bold"),
                     anchor="w").pack(fill="x", padx=30, pady=(0, 2))
        cv = tk.StringVar()
        ctk.CTkEntry(body, textvariable=cv, width=300, height=36,
                     show="*", placeholder_text="Confirm password").pack(padx=30, pady=(0, 16))

        err_lbl = ctk.CTkLabel(body, text="", font=ctk.CTkFont(size=11),
                               text_color="#B71C1C")
        err_lbl.pack()

        def _save():
            u = uv.get().strip()
            p = pv.get()
            c2 = cv.get()
            if not u:
                err_lbl.configure(text="Username cannot be empty.")
                return
            if len(p) < 4:
                err_lbl.configure(text="Password must be at least 4 characters.")
                return
            if p != c2:
                err_lbl.configure(text="Passwords do not match.")
                return
            license_manager.save_ytd_credentials(u, p)
            dlg.destroy()
            messagebox.showinfo("Credentials Saved",
                f"YTD credentials created for '{u}'.\n\n"
                "You will be asked for these when unfreezing a month in the YTD Database.")

        bf = ctk.CTkFrame(dlg, fg_color="transparent")
        bf.pack(pady=10)
        ctk.CTkButton(bf, text="Save Credentials", command=_save,
                      fg_color=THEME_PRIMARY, hover_color=THEME_HOVER,
                      height=36, width=160).pack(side="left", padx=8)
        ctk.CTkButton(bf, text="Skip", command=dlg.destroy,
                      fg_color="#757575", hover_color="#616161",
                      height=36, width=90).pack(side="left", padx=4)
        self.wait_window(dlg)

    def _verify_ytd_creds_dialog(self) -> bool:
        """Show credential prompt before unfreezing; return True if verified."""
        if not license_manager.has_ytd_credentials():
            return True  # no credentials configured — allow unfreeze
        result = [False]
        dlg = ctk.CTkToplevel(self)
        dlg.title("YTD Security Check")
        dlg.geometry("360x280")
        dlg.grab_set(); dlg.lift(); dlg.focus_force()
        dlg.resizable(False, False)

        ctk.CTkLabel(dlg, text="Enter YTD Credentials to Unfreeze",
                     font=ctk.CTkFont(size=14, weight="bold"),
                     text_color=THEME_DARK).pack(pady=(22, 8))
        ctk.CTkLabel(dlg,
                     text=f"Username: {license_manager.get_ytd_username()}",
                     font=ctk.CTkFont(size=11), text_color="#555").pack()

        pv = tk.StringVar()
        ctk.CTkLabel(dlg, text="Password", font=ctk.CTkFont(size=12),
                     anchor="w").pack(fill="x", padx=40, pady=(14, 2))
        pe = ctk.CTkEntry(dlg, textvariable=pv, show="*", width=280, height=36)
        pe.pack(padx=40)
        pe.focus()

        err_lbl = ctk.CTkLabel(dlg, text="", font=ctk.CTkFont(size=11),
                               text_color="#B71C1C")
        err_lbl.pack(pady=(6, 0))

        def _verify():
            u = license_manager.get_ytd_username()
            if license_manager.verify_ytd_credentials(u, pv.get()):
                result[0] = True
                dlg.destroy()
            else:
                err_lbl.configure(text="Incorrect password. Try again.")
                pv.set("")
                pe.focus()

        bf = ctk.CTkFrame(dlg, fg_color="transparent")
        bf.pack(pady=14)
        ctk.CTkButton(bf, text="Confirm", command=_verify,
                      fg_color=THEME_PRIMARY, hover_color=THEME_HOVER,
                      height=36, width=120).pack(side="left", padx=8)
        ctk.CTkButton(bf, text="Cancel", command=dlg.destroy,
                      fg_color="#757575", hover_color="#616161",
                      height=36, width=90).pack(side="left", padx=4)
        dlg.bind("<Return>", lambda e: _verify())
        self.wait_window(dlg)
        return result[0]

    def _show_company_selection_or_prompt(self):
        """Show company list if any exist; otherwise open the add-company dialog."""
        if self._all_companies:
            ok = self._show_company_selection()
            if not ok:
                self.quit()
        else:
            ok = self._prompt_company_info()
            if not ok:
                self.quit()

    def create_widgets(self):
        # ══════════════════════════════════════════════════════════════════
        # ROOT: two-pane layout  [SIDEBAR | CONTENT]
        # ══════════════════════════════════════════════════════════════════
        root = ctk.CTkFrame(self, fg_color="#F0F2F5", corner_radius=0)
        root.pack(fill="both", expand=True)

        # ── LEFT SIDEBAR (fixed 400px) ────────────────────────────────────
        sidebar = ctk.CTkFrame(root, fg_color="#880E4F", width=400, corner_radius=0)
        sidebar.pack(side="left", fill="y")
        sidebar.pack_propagate(False)

        # — Branding —
        brand = ctk.CTkFrame(sidebar, fg_color="transparent")
        brand.pack(fill="x", padx=18, pady=(22, 10))
        try:
            logo_path = get_resource_path("logo small.png")
            if os.path.exists(logo_path):
                logo_image = Image.open(logo_path)
                self.logo_ctk = ctk.CTkImage(light_image=logo_image,
                                             dark_image=logo_image, size=(42, 42))
                ctk.CTkLabel(brand, image=self.logo_ctk, text="").pack(side="left", padx=(0, 10))
        except Exception:
            pass
        title_col = ctk.CTkFrame(brand, fg_color="transparent")
        title_col.pack(side="left")
        ctk.CTkLabel(title_col, text="GST Recon",
                     font=ctk.CTkFont(size=17, weight="bold"),
                     text_color="white").pack(anchor="w")
        ctk.CTkLabel(title_col, text="GSTR-2B vs Books",
                     font=ctk.CTkFont(size=10), text_color="#F48FB1").pack(anchor="w")

        ctk.CTkFrame(sidebar, fg_color="#AD1457", height=1).pack(fill="x", padx=18, pady=(4, 10))

        # Keep these dicts so that _sync_sidebar_file / _refresh_upload_counter
        # don't crash — they already guard with 'if file_type in self._sidebar_dots'
        self.display_names = {
            'B2B':    '2B-B2B',    'B2BA':    '2B-B2BA',
            'CDNR':   '2B-CDNR',   'CDNRA':   '2B-CDNRA',
            'IMPG':   '2B-IMPG',   'IMPGSEZ': '2B-IMPGSEZ',
            'ITC':    'Register-ITC',
        }
        self.file_labels      = {}
        self.file_status_dots = {}
        self._sidebar_dots    = {}
        self._sidebar_file_lbls = {}
        # Silent dummy — _refresh_upload_counter calls .configure() on this
        self.upload_summary_lbl = ctk.CTkLabel(sidebar, text="")

        # — Instructions —
        ctk.CTkLabel(sidebar, text="HOW TO USE",
                     font=ctk.CTkFont(size=11, weight="bold"),
                     text_color="#F48FB1").pack(anchor="w", padx=18, pady=(0, 6))

        instructions = [
            ("1.  Download the Template",
             "Click 'Download Template' to get a ready-made Excel file."),
            ("2.  Fill in your GST data",
             "Open the file and enter your data in these sheets: B2B-2b, B2BA-2b, CDNR-2b, CDNRA-2b, IMPG-2b, IMPGSEZ-2b and ITC-SR."),
            ("3.  Upload the filled file",
             "Click 'Browse & Upload Excel' to load your completed file. (Switch to CSV Mode if you have individual CSV files instead.)"),
            ("4.  Run the reconciliation",
             "Click 'Process and Reconcile'. The tool will automatically match your ITC entries with GSTR-2B."),
            ("5.  Check & download results",
             "View the results on screen, then download the Excel report with your matched and unmatched items."),
            ("6.  Save for year tracking",
             "Use 'Save to YTD Database' to a running record of all\nyour monthly reconciliations."),
        ]

        for title, body in instructions:
            ctk.CTkLabel(sidebar, text=title,
                         font=ctk.CTkFont(size=15, weight="bold"),
                         text_color="white", anchor="w",
                         justify="left", wraplength=310).pack(anchor="w", padx=14, pady=(10, 0))
            ctk.CTkLabel(sidebar, text=body,
                         font=ctk.CTkFont(size=13),
                         text_color="#F8BBD0", anchor="w",
                         justify="left", wraplength=304).pack(anchor="w", padx=18, pady=(2, 0))

        ctk.CTkFrame(sidebar, fg_color="#AD1457", height=1).pack(fill="x", padx=18, pady=12)

        # — Quick actions —
        ctk.CTkLabel(sidebar, text="QUICK ACTIONS",
                     font=ctk.CTkFont(size=9, weight="bold"),
                     text_color="#F48FB1").pack(anchor="w", padx=18, pady=(0, 6))

        def _sb_btn(text, cmd, fg="#C2185B", hover="#AD1457", tc="white"):
            ctk.CTkButton(sidebar, text=text, command=cmd,
                          fg_color=fg, hover_color=hover, text_color=tc,
                          height=32, corner_radius=6, anchor="w",
                          font=ctk.CTkFont(size=11)
                          ).pack(fill="x", padx=14, pady=2)

        _sb_btn("Clear All Files",  self.clear_all_files,
                fg="#6D1B3A", hover="#5C1630")
        _sb_btn("YTD Database",    self.open_ytd_database_window,
                fg="#1B5E20", hover="#2E7D32")

        # Spacer pushes footer to bottom
        ctk.CTkLabel(sidebar, text="").pack(expand=True, fill="y")

        # — Sidebar footer —
        ctk.CTkFrame(sidebar, fg_color="#AD1457", height=1).pack(fill="x", padx=18, pady=(0, 6))
        for line in ["© GSC in time 2026", "info@gscintime.com", "+91-22-4612 5600"]:
            ctk.CTkLabel(sidebar, text=line, font=ctk.CTkFont(size=9),
                         text_color="#F48FB1").pack(pady=1)
        ctk.CTkLabel(sidebar, text="").pack(pady=4)

        # ── RIGHT CONTENT AREA ────────────────────────────────────────────
        content_outer = ctk.CTkFrame(root, fg_color="#F0F2F5", corner_radius=0)
        content_outer.pack(side="left", fill="both", expand=True)

        # Top bar
        top_bar = ctk.CTkFrame(content_outer, fg_color="white",
                               corner_radius=0, height=54)
        top_bar.pack(fill="x")
        top_bar.pack_propagate(False)

        ctk.CTkLabel(top_bar, text="GST Reconciliation Dashboard",
                     font=ctk.CTkFont(size=15, weight="bold"),
                     text_color="#222222").pack(side="left", padx=20, pady=14)

        # Company info section in top bar
        company_bar = ctk.CTkFrame(top_bar, fg_color="transparent")
        company_bar.pack(side="left", padx=(4, 0), pady=10)
        ctk.CTkLabel(company_bar, text="Working on:",
                     font=ctk.CTkFont(size=10), text_color="#9E9E9E").pack(side="left", padx=(0, 4))
        self._company_info_lbl = ctk.CTkLabel(
            company_bar, text="",
            font=ctk.CTkFont(size=11, weight="bold"), text_color=THEME_PRIMARY)
        self._company_info_lbl.pack(side="left")
        ctk.CTkButton(company_bar, text="Edit", width=46, height=24,
                      fg_color="#F3E5F5", hover_color="#E1BEE7",
                      text_color=THEME_DARK, font=ctk.CTkFont(size=10),
                      command=self._prompt_company_info).pack(side="left", padx=(6, 0))
        self._update_company_display()

        # Help icon
        ctk.CTkButton(top_bar, text="?", width=34, height=34,
                      fg_color="#5C6BC0", hover_color="#3949AB",
                      text_color="white", font=ctk.CTkFont(size=15, weight="bold"),
                      corner_radius=17,
                      command=self.open_help_dialog).pack(side="right", padx=(6, 16), pady=10)


        # Scrollable content
        self.main_container = ctk.CTkScrollableFrame(
            content_outer, fg_color="#F0F2F5", corner_radius=0)
        self.main_container.pack(fill="both", expand=True)

        # ── UPLOAD CARD ───────────────────────────────────────────────────
        upload_card = ctk.CTkFrame(self.main_container, fg_color="white",
                                   corner_radius=10, border_width=1,
                                   border_color="#E8E8E8")
        upload_card.pack(fill="x", padx=16, pady=(14, 6))

        uc_hdr = ctk.CTkFrame(upload_card, fg_color="transparent")
        uc_hdr.pack(fill="x", padx=16, pady=(12, 6))
        ctk.CTkLabel(uc_hdr, text="Upload Files",
                     font=ctk.CTkFont(size=14, weight="bold"),
                     text_color="#222222").pack(side="left")

        # Mode toggle buttons
        mode_frame = ctk.CTkFrame(uc_hdr, fg_color="transparent")
        mode_frame.pack(side="right")
        self._mode_template_btn = ctk.CTkButton(
            mode_frame, text="Template Mode", width=130, height=28,
            fg_color=THEME_PRIMARY, text_color="white",
            hover_color=THEME_HOVER, font=ctk.CTkFont(size=11),
            corner_radius=6,
            command=lambda: self._set_upload_mode('template'))
        self._mode_template_btn.pack(side="left", padx=(0, 4))
        self._mode_csv_btn = ctk.CTkButton(
            mode_frame, text="CSV Mode", width=100, height=28,
            fg_color="#E0E0E0", text_color="#555555",
            hover_color="#BDBDBD", font=ctk.CTkFont(size=11),
            corner_radius=6,
            command=lambda: self._set_upload_mode('csv'))
        self._mode_csv_btn.pack(side="left")

        ctk.CTkFrame(upload_card, fg_color="#F0F0F0", height=1).pack(fill="x", padx=16)

        # ── Template mode frame ───────────────────────────────────────────
        self._template_upload_frame = ctk.CTkFrame(upload_card, fg_color="#F8F9FA",
                                                    corner_radius=8, border_width=2,
                                                    border_color="#E8E8E8")
        tuf = self._template_upload_frame
        ctk.CTkLabel(tuf, text="Upload Excel Template (All Sheets)",
                     font=ctk.CTkFont(size=13, weight="bold"),
                     text_color="#333333").pack(pady=(18, 4))
        ctk.CTkLabel(tuf,
                     text="Select the filled Excel file containing sheets:\n"
                          "B2B-2b, B2BA-2b, CDNR-2b, CDNRA-2b, IMPG-2b, IMPGSEZ-2b, ITC-SR",
                     font=ctk.CTkFont(size=11), text_color="#757575",
                     justify="center").pack(pady=(0, 12))
        tuf_btns = ctk.CTkFrame(tuf, fg_color="transparent")
        tuf_btns.pack(pady=(0, 18))
        ctk.CTkButton(tuf_btns, text="Download Template",
                      command=self.download_template,
                      fg_color="#2E7D32", hover_color="#1B5E20",
                      font=ctk.CTkFont(size=13, weight="bold"),
                      height=42, width=200, corner_radius=8).pack(side="left", padx=(0, 10))
        ctk.CTkButton(tuf_btns, text="Browse & Upload Excel",
                      command=self.upload_excel_template,
                      fg_color=THEME_PRIMARY, hover_color=THEME_HOVER,
                      font=ctk.CTkFont(size=13, weight="bold"),
                      height=42, width=220, corner_radius=8).pack(side="left")

        # ── CSV mode frame ────────────────────────────────────────────────
        self._csv_upload_frame = ctk.CTkFrame(upload_card, fg_color="transparent")
        grid_frame = ctk.CTkFrame(self._csv_upload_frame, fg_color="transparent")
        grid_frame.pack(fill="x", padx=0, pady=0)

        left_col = ctk.CTkFrame(grid_frame, fg_color="transparent")
        left_col.pack(side="left", fill="both", expand=True, padx=4)
        right_col = ctk.CTkFrame(grid_frame, fg_color="transparent")
        right_col.pack(side="right", fill="both", expand=True, padx=4)

        ctk.CTkLabel(left_col, text="GSTR-2B — From GST Portal",
                     font=ctk.CTkFont(size=10, weight="bold"),
                     text_color="#9E9E9E").pack(anchor="w", padx=4, pady=(4, 2))
        for ft in ['B2B', 'B2BA', 'CDNR', 'CDNRA']:
            self.create_file_upload(left_col, ft)

        ctk.CTkLabel(right_col, text="GSTR-2B  /  Books Register",
                     font=ctk.CTkFont(size=10, weight="bold"),
                     text_color="#9E9E9E").pack(anchor="w", padx=4, pady=(4, 2))
        for ft in ['IMPG', 'IMPGSEZ', 'ITC']:
            self.create_file_upload(right_col, ft)

        # Start in template mode (default)
        self._set_upload_mode('template')

        # ── PROCESS BUTTON ────────────────────────────────────────────────
        self.process_btn = ctk.CTkButton(
            self.main_container,
            text="Process and Reconcile",
            command=self.start_processing,
            fg_color=THEME_PRIMARY, hover_color=THEME_HOVER,
            font=ctk.CTkFont(size=15, weight="bold"),
            height=50, corner_radius=8)
        self.process_btn.pack(fill="x", padx=16, pady=8)

        # ── RESULTS (hidden until done) ───────────────────────────────────
        self.results_frame = ctk.CTkFrame(
            self.main_container, fg_color="white",
            corner_radius=10, border_width=1, border_color="#A5D6A7")

        res_hdr = ctk.CTkFrame(self.results_frame, fg_color="#2E7D32", corner_radius=8)
        res_hdr.pack(fill="x", padx=8, pady=8)
        ctk.CTkLabel(res_hdr,
                     text="Reconciliation Complete  —  Download Your Results",
                     font=ctk.CTkFont(size=13, weight="bold"),
                     text_color="white").pack(side="left", padx=14, pady=8)

        row1 = ctk.CTkFrame(self.results_frame, fg_color="transparent")
        row1.pack(fill="x", padx=12, pady=(0, 4))
        row2 = ctk.CTkFrame(self.results_frame, fg_color="transparent")
        row2.pack(fill="x", padx=12, pady=(0, 10))

        self.download_reconciliation_btn = ctk.CTkButton(
            row1, text="Download Books Results",
            command=self.download_reconciliation,
            fg_color=THEME_PRIMARY, hover_color=THEME_HOVER, height=36, width=210)
        self.download_reconciliation_btn.pack(side="left", padx=(0, 6), pady=4)

        self.download_2a_results_btn = ctk.CTkButton(
            row1, text="Download 2B Results",
            command=self.download_2a_results,
            fg_color="#1565C0", hover_color="#0D47A1", height=36, width=190)
        self.download_2a_results_btn.pack(side="left", padx=6, pady=4)

        self.download_party_wise_btn = ctk.CTkButton(
            row1, text="Party-wise Report",
            command=self.download_party_wise_report,
            fg_color="#4527A0", hover_color="#311B92", height=36, width=180)
        self.download_party_wise_btn.pack(side="left", padx=6, pady=4)

        self.view_results_btn = ctk.CTkButton(
            row1, text="View Results",
            command=self.view_results,
            fg_color="#6A1B9A", hover_color="#4A148C", height=36, width=140)
        self.view_results_btn.pack(side="left", padx=6, pady=4)

        self.debug_matching_btn = ctk.CTkButton(
            row2, text="Debug Tools",
            command=self.open_combined_debug_window,
            fg_color="#E65100", hover_color="#BF360C", height=36, width=160)
        self.debug_matching_btn.pack(side="left", padx=(0, 6), pady=4)

        self.save_to_db_btn = ctk.CTkButton(
            row2, text="Save to YTD Database",
            command=self.save_to_database,
            fg_color="#2E7D32", hover_color="#1B5E20", height=36, width=200)
        self.save_to_db_btn.pack(side="left", padx=6, pady=4)

        self.match_past_2a_btn = ctk.CTkButton(
            row2, text="Match with Past 2B DB",
            command=self.match_with_past_2a,
            fg_color="#1565C0", hover_color="#0D47A1", height=36, width=200)
        self.match_past_2a_btn.pack(side="left", padx=6, pady=4)

        self.gst_status_btn = ctk.CTkButton(
            row2, text="Check GST Status",
            command=self.check_gst_status,
            fg_color="#00695C", hover_color="#004D40", height=36, width=175)
        self.gst_status_btn.pack(side="left", padx=6, pady=4)

        self.match_past_itc_btn = ctk.CTkButton(
            row2, text="Match with Past Books DB",
            command=self.match_with_past_itc,
            fg_color="#00695C", hover_color="#004D40", height=36, width=200)
        self.match_past_itc_btn.pack(side="left", padx=6, pady=4)

        self.same_month_cancel_btn = ctk.CTkButton(
            row2, text="Same Month Cancellation",
            command=self.show_same_month_cancellations,
            fg_color="#5C6BC0", hover_color="#3949AB", height=36, width=220)
        self.same_month_cancel_btn.pack(side="left", padx=6, pady=4)

        row3 = ctk.CTkFrame(self.results_frame, fg_color="transparent")
        row3.pack(fill="x", padx=12, pady=(0, 10))
        self.summary_report_btn = ctk.CTkButton(
            row3, text="Summary Report",
            command=self.show_summary_report,
            fg_color="#AD1457", hover_color="#880E4F", height=36, width=180)
        self.summary_report_btn.pack(side="left", padx=(0, 6), pady=4)

        # GSTN Debug is now inside the combined Debug Tools window

        # ── PROGRESS ──────────────────────────────────────────────────────
        self.progress_frame = ctk.CTkFrame(
            self.main_container, fg_color="white",
            corner_radius=10, border_width=1, border_color="#E8E8E8")
        self.progress_frame.pack(fill="x", padx=16, pady=4)

        prog_row = ctk.CTkFrame(self.progress_frame, fg_color="transparent")
        prog_row.pack(fill="x", padx=16, pady=(10, 4))
        self.progress_label = ctk.CTkLabel(
            prog_row, text="Ready to process",
            font=ctk.CTkFont(size=12), text_color="#555555")
        self.progress_label.pack(side="left")

        self.progress_bar = ctk.CTkProgressBar(
            self.progress_frame, progress_color=THEME_PRIMARY, height=8)
        self.progress_bar.pack(fill="x", padx=16, pady=(0, 12))
        self.progress_bar.set(0)

        # ── HIDDEN LOG (keeps internal logging + error-popup logic working) ──
        _hidden = ctk.CTkFrame(self.main_container, fg_color="transparent", height=0)
        _hidden.pack()
        self.log_text = ctk.CTkTextbox(_hidden, height=1,
                                       font=ctk.CTkFont(size=10, family="Courier"))

        # ── SPINNER / STATUS AREA ─────────────────────────────────────────
        spin_card = ctk.CTkFrame(self.main_container, fg_color="white",
                                 corner_radius=10, border_width=1, border_color="#E8E8E8")
        spin_card.pack(fill="x", padx=16, pady=(4, 16))

        spin_inner = ctk.CTkFrame(spin_card, fg_color="transparent")
        spin_inner.pack(pady=14)

        self._spinner_lbl = ctk.CTkLabel(spin_inner, text="",
                                         font=ctk.CTkFont(size=22), text_color=THEME_PRIMARY,
                                         width=30)
        self._spinner_lbl.pack(side="left", padx=(0, 8))

        self._status_lbl = ctk.CTkLabel(spin_inner, text="Ready to process",
                                        font=ctk.CTkFont(size=13), text_color="#555555")
        self._status_lbl.pack(side="left")

        self._spinner_running = False
        self._spinner_frame   = spin_card

        # ── REPORT LOG (appears after reconciliation) ─────────────────────
        self._report_log_frame = ctk.CTkFrame(
            self.main_container, fg_color="white",
            corner_radius=10, border_width=1, border_color="#E8E8E8")
        # Not packed yet — shown only after reconciliation completes
        self._report_log_lbl = ctk.CTkLabel(
            self._report_log_frame,
            text="", font=ctk.CTkFont(size=12, family="Courier"),
            text_color="#333333", justify="left", anchor="w", wraplength=700)
        self._report_log_lbl.pack(anchor="w", padx=16, pady=10)

    def _update_upload_summary(self):
        """Refresh the 'X / 7 files uploaded' counter in the sidebar."""
        count = sum(1 for v in self.file_paths.values() if v)
        total = len(self.file_paths)
        color = "#A5D6A7" if count == total else ("#F48FB1" if count == 0 else "#FFE082")
        self.upload_summary_lbl.configure(
            text=f"{count} / {total} files uploaded", text_color=color)

    def create_file_upload(self, parent, file_type, number=0):  # noqa: ARG002
        """Create a styled file upload row with status indicator dot."""
        is_itc = (file_type == 'ITC')
        row_bg  = "#FFF8E1" if is_itc else "#FAFAFA"
        border  = "#FFB74D" if is_itc else "#EEEEEE"

        frame = ctk.CTkFrame(parent, fg_color=row_bg, corner_radius=6,
                             border_width=1, border_color=border)
        frame.pack(fill="x", pady=2, padx=2)

        # Status dot: grey = not uploaded, green = uploaded
        dot = ctk.CTkLabel(frame, text="●", font=ctk.CTkFont(size=16),
                            text_color="#BDBDBD", width=22)
        dot.pack(side="left", padx=(8, 2), pady=7)
        self.file_status_dots[file_type] = dot

        display_name = self.display_names.get(file_type, file_type)
        ctk.CTkLabel(frame, text=display_name,
                     font=ctk.CTkFont(size=12, weight="bold" if is_itc else "normal"),
                     text_color=THEME_DARK if is_itc else "#333333",
                     width=120, anchor="w").pack(side="left", padx=(4, 0))

        self.file_labels[file_type] = ctk.CTkLabel(
            frame, text="No file selected",
            font=ctk.CTkFont(size=10), text_color="#9E9E9E", anchor="w")
        self.file_labels[file_type].pack(side="left", padx=6, fill="x", expand=True)

        ctk.CTkButton(frame, text="✕",
                      command=lambda ft=file_type: self.clear_file(ft),
                      width=28, height=26, fg_color="#9E9E9E", hover_color="#616161",
                      font=ctk.CTkFont(size=11)).pack(side="right", padx=(2, 6), pady=6)

        ctk.CTkButton(frame, text="Browse",
                      command=lambda ft=file_type: self.browse_file(ft),
                      width=80, height=26,
                      fg_color=THEME_HOVER if is_itc else THEME_PRIMARY,
                      hover_color=THEME_DARK if is_itc else THEME_HOVER,
                      font=ctk.CTkFont(size=11)).pack(side="right", padx=2, pady=6)

    def _sync_sidebar_file(self, file_type, filename=None):
        """Update sidebar dot + short filename label for a file type."""
        if file_type in self._sidebar_dots:
            if filename:
                self._sidebar_dots[file_type].configure(text_color="#A5D6A7")
                short = filename if len(filename) <= 14 else filename[:13] + "…"
                self._sidebar_file_lbls[file_type].configure(
                    text=short, text_color="white")
            else:
                self._sidebar_dots[file_type].configure(text_color="#F06292")
                self._sidebar_file_lbls[file_type].configure(
                    text="—", text_color="#F48FB1")

    def browse_file(self, file_type):
        """Open file dialog to select a CSV file"""
        display_name = self.display_names.get(file_type, file_type)
        filepath = filedialog.askopenfilename(
            title=f"Select {display_name} CSV File",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if filepath:
            self.file_paths[file_type] = filepath
            filename = os.path.basename(filepath)
            self.file_labels[file_type].configure(text=filename, text_color=THEME_PRIMARY)
            if file_type in self.file_status_dots:
                self.file_status_dots[file_type].configure(text_color="#43A047")
            self._sync_sidebar_file(file_type, filename)
            self._update_upload_summary()
            self.log(f"Selected {display_name}: {filename}")

    def clear_file(self, file_type):
        """Clear selected file"""
        self.file_paths[file_type] = None
        self.file_labels[file_type].configure(text="No file selected", text_color="#9E9E9E")
        if file_type in self.file_status_dots:
            self.file_status_dots[file_type].configure(text_color="#BDBDBD")
        self._sync_sidebar_file(file_type, None)
        self._update_upload_summary()

    def clear_all_files(self):
        """Clear all selected files"""
        for file_type in self.file_paths.keys():
            self.file_paths[file_type] = None
            self.file_labels[file_type].configure(text="No file selected", text_color="#9E9E9E")
            if file_type in self.file_status_dots:
                self.file_status_dots[file_type].configure(text_color="#BDBDBD")
            self._sync_sidebar_file(file_type, None)
        self._update_upload_summary()
        self.log("All files cleared.")

    # ── Spinner helpers ───────────────────────────────────────────────────────
    _SPIN_CHARS = ['⠋', '⠙', '⠹', '⠸', '⠼', '⠴', '⠦', '⠧', '⠇', '⠏']

    def _start_spinner(self, status="Processing…"):
        self._spinner_running = True
        self._spin_idx = 0
        self.after(0, lambda: self._status_lbl.configure(text=status))
        self._tick_spinner()

    def _tick_spinner(self):
        if not self._spinner_running:
            return
        ch = self._SPIN_CHARS[self._spin_idx % len(self._SPIN_CHARS)]
        self._spinner_lbl.configure(text=ch)
        self._spin_idx += 1
        self.after(100, self._tick_spinner)

    def _stop_spinner(self, done_text="Done"):
        self._spinner_running = False
        self._spinner_lbl.configure(text="✓")
        self._status_lbl.configure(text=done_text)

    def log(self, message):
        """Add message to log (hidden); errors also appear as a popup dialog."""
        self.log_text.insert("end", message + "\n")
        if message.lower().startswith("error") or ": error" in message.lower():
            self.after(0, lambda m=message: messagebox.showerror("Error", m))

    def download_template(self):
        """Download template Excel file - copies the existing Template updated.xlsx"""
        try:
            # Try to find the existing template file
            template_source = get_resource_path("Template updated.xlsx")

            if not os.path.exists(template_source):
                # Fallback to Template all.xlsx
                template_source = get_resource_path("Template all.xlsx")

            filepath = filedialog.asksaveasfilename(
                title="Save Template",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile="Template updated.xlsx"
            )
            if filepath:
                if os.path.exists(template_source):
                    # Copy existing template file
                    import shutil
                    shutil.copy2(template_source, filepath)
                else:
                    # Generate template if source doesn't exist
                    template_data = create_template_excel()
                    with open(filepath, 'wb') as f:
                        f.write(template_data)
                self.log(f"Template saved to: {filepath}")
                messagebox.showinfo("Success", f"Template saved successfully!\n\nLocation: {filepath}")
        except Exception as e:
            self.log(f"Error saving template: {str(e)}")
            messagebox.showerror("Error", f"Failed to save template: {str(e)}")

    def upload_excel_template(self):
        """Upload a single Excel file with all sheets (B2B, B2BA, CDNR, CDNRA, IMPG, IMPGSEZ, ITC)"""
        filepath = filedialog.askopenfilename(
            title="Select Excel File with All Sheets",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filepath:
            try:
                # Read all sheets from the Excel file
                excel_file = pd.ExcelFile(filepath)
                sheet_names = excel_file.sheet_names
                self.log(f"Loading Excel file: {os.path.basename(filepath)}")
                self.log(f"Found sheets: {', '.join(sheet_names)}")

                # Map sheet names to file types — robust matching that handles any
                # spacing, dashes, underscores, or special chars between tokens.
                import re as _re
                def _norm_sheet(s):
                    """Strip all non-alphanumeric chars and uppercase — e.g. 'B2B- 2B' → 'B2B2B'"""
                    return _re.sub(r'[^A-Z0-9]', '', s.upper())

                # Exact normalised-key → logical name (checked first)
                _norm_map = {
                    'ITC': 'ITC', 'ITCSR': 'ITC',
                    'B2B': 'B2B', 'B2B2B': 'B2B',
                    'B2BA': 'B2BA', 'B2BA2B': 'B2BA',
                    'CDNR': 'CDNR', 'CDNR2B': 'CDNR',
                    'CDNRA': 'CDNRA', 'CDNRA2B': 'CDNRA',
                    'IMPG': 'IMPG', 'IMPG2B': 'IMPG',
                    'IMPGSEZ': 'IMPGSEZ', 'IMPGSEZ2B': 'IMPGSEZ',
                }
                # Priority-ordered prefix fallback (longer prefixes first to avoid B2BA → B2B)
                _prefix_order = ['IMPGSEZ', 'CDNRA', 'B2BA', 'IMPG', 'CDNR', 'B2B', 'ITC']

                sheet_mapping = {}
                for sheet in sheet_names:
                    norm = _norm_sheet(sheet)
                    ft = _norm_map.get(norm)
                    if not ft:
                        # Fallback: check if normalised name starts with a known keyword
                        for kw in _prefix_order:
                            if norm.startswith(kw):
                                ft = kw
                                break
                    if ft and ft not in sheet_mapping:
                        sheet_mapping[ft] = sheet
                        self.log(f"  Matched sheet '{sheet}' → {ft}")

                # Load each recognized sheet
                loaded_count = 0
                for file_type in ['B2B', 'B2BA', 'CDNR', 'CDNRA', 'IMPG', 'IMPGSEZ', 'ITC']:
                    if file_type in sheet_mapping:
                        # Store the Excel file path and sheet name
                        self.file_paths[file_type] = (filepath, sheet_mapping[file_type])
                        self.file_labels[file_type].configure(
                            text=f"{os.path.basename(filepath)} [{sheet_mapping[file_type]}]",
                            text_color=THEME_PRIMARY
                        )
                        if file_type in self.file_status_dots:
                            self.file_status_dots[file_type].configure(text_color="#43A047")
                        self._sync_sidebar_file(file_type, os.path.basename(filepath))
                        loaded_count += 1
                        dn = self.display_names.get(file_type, file_type)
                        self.log(f"  - {dn}: Loaded from sheet '{sheet_mapping[file_type]}'")

                self._update_upload_summary()
                if loaded_count > 0:
                    messagebox.showinfo("Success", f"Loaded {loaded_count} sheets from Excel file!\n\nSheets found: {', '.join(sheet_mapping.keys())}")
                else:
                    found_list = ', '.join(f"'{s}'" for s in sheet_names) if sheet_names else '(none)'
                    messagebox.showwarning(
                        "Warning",
                        f"No recognized sheets found!\n\n"
                        f"Sheets in file: {found_list}\n\n"
                        f"Expected names contain: B2B, B2BA, CDNR, CDNRA, IMPG, IMPGSEZ, or ITC\n"
                        f"(spaces, dashes and case are ignored)"
                    )

            except Exception as e:
                self.log(f"Error loading Excel file: {str(e)}")
                messagebox.showerror("Error", f"Failed to load Excel file: {str(e)}")

    def start_processing(self):
        """Start the reconciliation process in a separate thread"""
        # Check if ITC is provided (can be string path or tuple for Excel)
        if not self.file_paths['ITC']:
            messagebox.showerror("Error", "Register-ITC file/sheet is mandatory!")
            return

        self.process_btn.configure(state="disabled")
        self.progress_bar.set(0)
        self.log_text.delete("1.0", "end")
        self.results_frame.pack_forget()
        self._start_spinner("Loading files and running reconciliation…")

        # Run processing in a thread
        thread = threading.Thread(target=self.process_files)
        thread.start()

    def process_files(self):
        """Process all files and perform reconciliation"""
        try:
            self._debug_final_matches = []   # clear any pending matches from prior session
            tables = {}

            # Load files
            self.update_progress(0.1, "Loading files...")
            for name, filepath in self.file_paths.items():
                dn = self.display_names.get(name, name)
                if filepath:
                    try:
                        # Check if it's a tuple (Excel file with sheet name) or string (CSV file)
                        if isinstance(filepath, tuple):
                            # Excel file: (filepath, sheet_name)
                            excel_path, sheet_name = filepath
                            df = pd.read_excel(excel_path, sheet_name=sheet_name, dtype=str)
                            df = df.fillna('').replace('nan', '').replace('None', '')
                            tables[name] = fix_sci_notation_in_df(df)
                            self.log(f"Loaded {dn} from Excel sheet '{sheet_name}': {len(tables[name])} rows")
                        else:
                            # CSV file
                            try:
                                df = pd.read_csv(filepath, encoding='utf-8', dtype=str)
                            except Exception:
                                df = pd.read_csv(filepath, encoding='latin-1', dtype=str)
                            tables[name] = fix_sci_notation_in_df(df.fillna(''))
                            self.log(f"Loaded {dn}: {len(tables[name])} rows")
                    except Exception as e:
                        self.log(f"Error loading {dn}: {str(e)}")
                        tables[name] = pd.DataFrame()
                else:
                    tables[name] = pd.DataFrame()

            # Snapshot original per-type DataFrames BEFORE normalization (for template-structured YTD)
            self._original_itc_raw = tables.get('ITC', pd.DataFrame()).copy()
            self._original_2b_tables = {
                ft: tables[ft].copy()
                for ft in ('B2B', 'B2BA', 'CDNR', 'CDNRA', 'IMPG', 'IMPGSEZ')
                if ft in tables and not tables[ft].empty
            }

            # Normalize columns
            self.update_progress(0.2, "Normalizing columns...")
            if 'ITC' in tables and not tables['ITC'].empty:
                tables['ITC'] = normalize_itc_columns(tables['ITC'])
            if 'CDNR' in tables and not tables['CDNR'].empty:
                tables['CDNR'] = normalize_cdnr_columns(tables['CDNR'])
            if 'CDNRA' in tables and not tables['CDNRA'].empty:
                tables['CDNRA'] = normalize_cdnr_columns(tables['CDNRA'])

            # Step 1: Merge duplicates (keep original ITC to preserve line items)
            self.update_progress(0.3, "Step 1: Merging duplicates...")
            # Keep a copy of the original ITC (line items) before collapsing duplicates for matching
            original_itc = tables['ITC'].copy()
            self.original_itc = original_itc
            tables['ITC'] = merge_duplicate_vendor_invoices(tables['ITC'], self.log)  # merged per invoice

            # Step 2: Match B2B with B2BA
            self.update_progress(0.4, "Step 2: Matching B2B with B2BA...")
            if 'B2B' in tables and 'B2BA' in tables:
                b2b_inv_col = None
                b2ba_inv_col = None

                for col in tables['B2B'].columns:
                    if col.lower().strip() == 'invoice no':
                        b2b_inv_col = col
                        break

                if not tables['B2BA'].empty:
                    for col in tables['B2BA'].columns:
                        if col.lower().strip() == 'invoice no':
                            b2ba_inv_col = col
                            break

                if b2b_inv_col and b2ba_inv_col:
                    tables['B2B'], tables['B2BA'] = match_and_update(
                        tables['B2B'], tables['B2BA'],
                        b2b_inv_col, b2ba_inv_col, 'B2B', self.log
                    )

            # Step 3: Match CDNR with CDNRA
            self.update_progress(0.5, "Step 3: Matching CDNR with CDNRA...")
            if 'CDNR' in tables and 'CDNRA' in tables:
                cdnr_inv_col = None
                cdnra_inv_col = None

                if not tables['CDNR'].empty:
                    for col in tables['CDNR'].columns:
                        cl = col.lower().strip()
                        if cl == 'note no' or cl == 'invoice no' or cl == 'boe no':
                            cdnr_inv_col = col
                            break

                if not tables['CDNRA'].empty:
                    for col in tables['CDNRA'].columns:
                        cl = col.lower().strip()
                        if cl == 'note no' or cl == 'invoice no' or cl == 'boe no':
                            cdnra_inv_col = col
                            break

                if cdnr_inv_col and cdnra_inv_col:
                    tables['CDNR'], tables['CDNRA'] = match_and_update(
                        tables['CDNR'], tables['CDNRA'],
                        cdnr_inv_col, cdnra_inv_col, 'CDNR', self.log
                    )

            # Store processed B2B and CDNR for same-month cancellation detection
            self.b2b_processed_df = tables.get('B2B', pd.DataFrame()).copy()
            self.cdnr_processed_df = tables.get('CDNR', pd.DataFrame()).copy()

            # Step 4: Create MERGED table
            self.update_progress(0.6, "Step 4: Creating MERGED table...")
            merged_df = create_merged_table(tables, self.log)
            self.merged_df = merged_df

            # Step 5: Create ITC register
            self.update_progress(0.7, "Step 5: Creating ITC register...")
            # Use merged ITC for register (invoice-level sums)
            itc_register = create_itc_register(tables['ITC'], self.log)
            self.itc_register = itc_register

            # Step 6: Create GSTR 2A table
            self.update_progress(0.8, "Step 6: Creating GSTR-2B table...")
            gstr_2a = create_gstr_2a(merged_df, self.log)

            # Step 7: Compare tables
            self.update_progress(0.85, "Step 7: Comparing tables...")
            self.comparison_df = compare_tables(itc_register, gstr_2a, self.log)

            # Create ITC result table with Status column
            self.update_progress(0.9, "Creating ITC results with Status...")
            # Pass original (pre-merged) ITC so all original line items are present in results
            # Use the aggregated comparison dataframe to ensure ITC Results reflect reconciliation (match counts)
            self.itc_result_df, matched_2a_status = create_itc_result(original_itc, itc_register, gstr_2a, self.comparison_df, merged_df, self.log)

            # Step 7.5: CDNR/CDNRA negative value matching
            self.update_progress(0.92, "Step 7.5: Matching CDNR/CDNRA negatives with ITC...")
            self.itc_result_df, matched_cdnr_2a_keys = match_cdnr_negatives(
                tables.get('CDNR', pd.DataFrame()),
                tables.get('CDNRA', pd.DataFrame()),
                self.itc_result_df, self.log
            )
            # Update matched_2a_status with CDNR rows matched by the CDNR negative matcher
            matched_2a_status.update({k: 'Matched' for k in matched_cdnr_2a_keys})

            # Build 2A Results: all merged rows with Status column
            self.update_progress(0.95, "Building 2B Results...")
            gstr_2a_results = merged_df.copy()
            gstr_2a_results['Status'] = gstr_2a_results.apply(
                lambda r: matched_2a_status.get(
                    normalize_gstin(str(r.get('GSTN', ''))) + '|' + normalize_invoice(str(r.get('Document_number', ''))),
                    'Not Found in ITC'
                ),
                axis=1
            )
            # Negate CGST/SGST/IGST for CDNR/CDNRA rows so 2B totals align with ITC sign convention.
            # ITC records credit notes as negative amounts (reducing ITC), while raw GSTR-2B files
            # store them as positive.  Negating here ensures that filtering by Matched status and
            # summing tax columns gives consistent totals in both sheets.
            _cdnr_mask = gstr_2a_results.get('TYPE', pd.Series(dtype=str)).astype(str).str.upper().isin(['CDNR', 'CDNRA'])
            if _cdnr_mask.any():
                for _col in ['CGST', 'SGST', 'IGST']:
                    if _col in gstr_2a_results.columns:
                        gstr_2a_results.loc[_cdnr_mask, _col] = (
                            gstr_2a_results.loc[_cdnr_mask, _col].apply(safe_numeric_conversion) * -1
                        )
            self.gstr_2a_results_df = gstr_2a_results

            # Keep unmatched_2a_df (rows with no ITC entry) for debug matching functionality
            self.unmatched_2a_df = self._compute_unmatched_2a(merged_df, original_itc, matched_2a_status)

            # Same Month Cancellation: match B2B amounts against CDNR amounts (same GSTIN)
            self.update_progress(0.98, "Finding same month cancellations...")
            self.same_month_cancel_df = find_same_month_cancellations(
                self.b2b_processed_df, self.cdnr_processed_df, self.log
            )

            # Complete
            self.update_progress(1.0, "Processing completed!")
            self.log("\n" + "=" * 50)
            self.log("PROCESSING COMPLETED SUCCESSFULLY!")
            self.log("=" * 50)

            if self.itc_result_df is not None and not self.itc_result_df.empty:
                total = len(self.itc_result_df)
                matched = len(self.itc_result_df[self.itc_result_df['Status'] == 'Matched'])
                unmatched = len(self.itc_result_df[self.itc_result_df['Status'] == 'Unmatched'])
                higher = len(self.itc_result_df[self.itc_result_df['Status'] == 'Higher in 2B'])
                lower = len(self.itc_result_df[self.itc_result_df['Status'] == 'Lower in 2B'])
                not_found = len(self.itc_result_df[self.itc_result_df['Status'] == 'Not found in 2B'])
                self.log(f"\nITC RESULTS SUMMARY:")
                self.log(f"  Total Records: {total}")
                self.log(f"  Matched: {matched} ({matched/total*100:.1f}%)")
                self.log(f"  Unmatched: {unmatched} ({unmatched/total*100:.1f}%)")
                self.log(f"  Higher in 2B: {higher} ({higher/total*100:.1f}%)")
                self.log(f"  Lower in 2B: {lower} ({lower/total*100:.1f}%)")
                self.log(f"  Not found in 2B: {not_found} ({not_found/total*100:.1f}%)")

            # Show results frame
            self.after(0, self.show_results_frame)

            # Build result summary for popup
            if self.itc_result_df is not None and not self.itc_result_df.empty:
                total    = len(self.itc_result_df)
                matched  = len(self.itc_result_df[self.itc_result_df['Status'] == 'Matched'])
                higher   = len(self.itc_result_df[self.itc_result_df['Status'] == 'Higher in 2B'])
                lower    = len(self.itc_result_df[self.itc_result_df['Status'] == 'Lower in 2B'])
                not_found= len(self.itc_result_df[self.itc_result_df['Status'] == 'Not found in 2B'])
                unmatched= total - matched - higher - lower - not_found
                pct = lambda n: f"{n/total*100:.1f}%" if total else "0%"
                summary = (
                    f"Reconciliation Complete!\n\n"
                    f"Total ITC Records : {total}\n"
                    f"✔  Matched         : {matched}  ({pct(matched)})\n"
                    f"↑  Higher in 2B    : {higher}  ({pct(higher)})\n"
                    f"↓  Lower in 2B     : {lower}  ({pct(lower)})\n"
                    f"✗  Not found in 2B : {not_found}  ({pct(not_found)})\n"
                    f"?  Unmatched       : {unmatched}  ({pct(unmatched)})"
                )
            else:
                summary = "Reconciliation complete. No ITC records found."

            def _finish(s=summary):
                self._stop_spinner("Reconciliation complete")
                messagebox.showinfo("Result", s)
                # Show report log panel at bottom of dashboard
                self._report_log_lbl.configure(text=s)
                self._report_log_frame.pack(fill="x", padx=16, pady=(0, 12))
            self.after(0, _finish)

        except Exception as e:
            self.log(f"\nError during processing: {str(e)}")
            import traceback
            self.log(traceback.format_exc())
            self.after(0, lambda: self._stop_spinner("Error — see popup for details"))

        finally:
            self.after(0, lambda: self.process_btn.configure(state="normal"))

    def update_progress(self, value, message):
        """Update progress bar, label and spinner status."""
        self.after(0, lambda: self.progress_bar.set(value))
        self.after(0, lambda: self.progress_label.configure(text=message))
        self.after(0, lambda m=message: self._status_lbl.configure(text=m))
        self.log(message)

    # ──────────────────────────────────────────────────────────────────────
    # YTD Database: Save & Open
    # ──────────────────────────────────────────────────────────────────────

    def _parse_invoice_date(self, date_val):
        """Try to parse a date value to (year, month, 'YYYY-MM-DD') tuple."""
        if date_val is None or (isinstance(date_val, float) and math.isnan(date_val)):
            return None, None, ''
        s = str(date_val).strip()
        for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%d/%m/%y', '%d-%b-%Y', '%b-%Y'):
            try:
                dt = datetime.strptime(s, fmt)
                return dt.year, dt.month, dt.strftime('%Y-%m-%d')
            except ValueError:
                continue
        # pandas Timestamp
        try:
            import pandas as pd
            ts = pd.to_datetime(date_val, dayfirst=True, errors='coerce')
            if ts is not pd.NaT and not pd.isna(ts):
                return ts.year, ts.month, ts.strftime('%Y-%m-%d')
        except Exception:
            pass
        return None, None, ''

    def _find_col(self, df, *keywords):
        """Return first column whose lower-stripped name contains all keywords."""
        for col in df.columns:
            cl = col.lower().strip()
            if all(k in cl for k in keywords):
                return col
        return None

    def save_to_database(self):
        """Save current reconciliation results to the YTD SQLite database."""
        if self.itc_result_df is None and self.gstr_2a_results_df is None:
            messagebox.showwarning("No Data", "Run a reconciliation first before saving to the database.")
            return

        # Always ask which DB to save to (so different reconciliations go to the right DB)
        os.makedirs(GST_DB_DIR, exist_ok=True)
        pick_win = ctk.CTkToplevel(self)
        pick_win.title("Choose Database")
        pick_win.geometry("380x320")
        pick_win.grab_set()
        pick_win.lift()
        pick_win.focus_force()

        ctk.CTkLabel(pick_win, text="Select or create a database:",
                     font=ctk.CTkFont(size=13, weight="bold")).pack(pady=(18, 6))

        db_files = sorted(f[:-3] for f in os.listdir(GST_DB_DIR) if f.endswith('.db'))
        chosen = [None]

        lf = ctk.CTkScrollableFrame(pick_win, height=140)
        lf.pack(fill="x", padx=16)
        for name in db_files:
            ctk.CTkButton(lf, text=name, anchor="w",
                          fg_color="transparent", hover_color="#FCE4EC",
                          text_color="#333333", height=30,
                          command=lambda n=name: [chosen.__setitem__(0, n),
                                                  pick_win.destroy()]
                          ).pack(fill="x", pady=1)

        ctk.CTkLabel(pick_win, text="— or create new —",
                     font=ctk.CTkFont(size=10), text_color="#9E9E9E").pack(pady=(8, 2))
        new_var = tk.StringVar()
        new_entry = ctk.CTkEntry(pick_win, textvariable=new_var,
                                 placeholder_text="New database name", width=230)
        new_entry.pack()

        def _use_new():
            n = new_var.get().strip()
            if n:
                chosen[0] = n
                pick_win.destroy()

        bf = ctk.CTkFrame(pick_win, fg_color="transparent")
        bf.pack(pady=10)
        ctk.CTkButton(bf, text="Create & Use", command=_use_new,
                      fg_color="#2E7D32", hover_color="#1B5E20", width=130).pack(side="left", padx=6)
        ctk.CTkButton(bf, text="Cancel", command=pick_win.destroy,
                      fg_color="#757575", hover_color="#616161", width=100).pack(side="left", padx=6)

        self.wait_window(pick_win)
        if not chosen[0]:
            return
        save_db_name_raw = chosen[0] if chosen[0].endswith('.db') else chosen[0] + '.db'
        save_db_path = os.path.join(GST_DB_DIR, save_db_name_raw)
        save_db = GSTDatabaseManager(save_db_path)
        save_db_label = save_db_name_raw[:-3]
        self.log(f"Saving to YTD Database: {save_db_label}")

        records = []

        # Columns to exclude from extra_data (already stored in named columns)
        _CORE_COLS = {'status', 'cgst', 'sgst', 'igst', 'taxable', 'gstin', 'gstn',
                      'invoice', 'doc no', 'date', '2b invoice', '2b gstin', 'remarks'}

        # ── ITC rows ──────────────────────────────────────────────────────
        if self.itc_result_df is not None and not self.itc_result_df.empty:
            df = self.itc_result_df
            gstin_col  = self._find_col(df, 'vendor', 'gstn') or self._find_col(df, 'gstin')
            inv_col    = self._find_col(df, 'vendor inv') or self._find_col(df, 'external doc') or self._find_col(df, 'invoice no')
            date_col   = self._find_col(df, 'invoice', 'date') or self._find_col(df, 'date')
            tv_col     = self._find_col(df, 'taxable') or self._find_col(df, 'base amount')
            cgst_col   = self._find_col(df, 'cgst')
            sgst_col   = self._find_col(df, 'sgst')
            igst_col   = self._find_col(df, 'igst')
            name_col   = self._find_col(df, 'vendor name') or self._find_col(df, 'trade', 'name')
            iv_col     = self._find_col(df, 'invoice value')
            tr_col     = self._find_col(df, 'tax rate')

            for _, row in df.iterrows():
                gstin = str(row[gstin_col]).strip() if gstin_col else ''
                inv   = str(row[inv_col]).strip()   if inv_col   else ''
                date_raw = row[date_col] if date_col else None
                yr, mo, date_str = self._parse_invoice_date(date_raw)
                status = str(row.get('Status', 'Unmatched'))
                matched_inv  = str(row.get('2B Invoice No', ''))
                matched_gstn = str(row.get('2B GSTIN', ''))
                db_status = 'Matched' if status in ('Matched', 'Higher in 2B', 'Lower in 2B',
                                                      'Matched but invoice number is not accurate') else 'Unmatched'
                # Capture extra columns not in the core schema
                extra = {}
                for col in df.columns:
                    cl = col.lower().strip()
                    if not any(k in cl for k in _CORE_COLS):
                        v = row[col]
                        if pd.notna(v) and str(v) not in ('', 'nan'):
                            extra[col] = str(v)
                records.append({
                    'record_type': 'ITC',
                    'gstin': gstin,
                    'invoice_number': inv,
                    'invoice_norm': normalize_gstin(gstin) + '|' + normalize_invoice(inv),
                    'invoice_date': date_str,
                    'invoice_month': mo,
                    'invoice_year': yr,
                    'taxable_value': safe_numeric_conversion(row[tv_col]) if tv_col else 0,
                    'cgst': safe_numeric_conversion(row[cgst_col]) if cgst_col else 0,
                    'sgst': safe_numeric_conversion(row[sgst_col]) if sgst_col else 0,
                    'igst': safe_numeric_conversion(row[igst_col]) if igst_col else 0,
                    'source_type': 'ITC',
                    'status': db_status,
                    'matched_invoice': matched_inv,
                    'matched_gstin': matched_gstn,
                    'vendor_name': str(row[name_col]).strip() if name_col else '',
                    'invoice_value': safe_numeric_conversion(row[iv_col]) if iv_col else 0,
                    'tax_rate': str(row[tr_col]).strip() if tr_col else '',
                    'extra_data': extra,
                })

        # ── 2A rows ───────────────────────────────────────────────────────
        if self.gstr_2a_results_df is not None and not self.gstr_2a_results_df.empty:
            df2 = self.gstr_2a_results_df
            for _, row in df2.iterrows():
                gstin = str(row.get('GSTN', '')).strip()
                inv   = str(row.get('Document_number', '')).strip()
                date_raw = (row.get('Invoice_Date') or row.get('Invoice_date')
                            or row.get('Invoice Date') or row.get('Date')
                            or row.get('Booking_Month'))
                yr, mo, date_str = self._parse_invoice_date(date_raw)
                status = str(row.get('Status', ''))
                db_status = 'Matched' if status == 'Matched' else 'Unmatched'
                records.append({
                    'record_type': '2A',
                    'gstin': gstin,
                    'invoice_number': inv,
                    'invoice_norm': normalize_gstin(gstin) + '|' + normalize_invoice(inv),
                    'invoice_date': date_str,
                    'invoice_month': mo,
                    'invoice_year': yr,
                    'taxable_value': safe_numeric_conversion(row.get('Taxable_value', 0)),
                    'cgst': safe_numeric_conversion(row.get('CGST', 0)),
                    'sgst': safe_numeric_conversion(row.get('SGST', 0)),
                    'igst': safe_numeric_conversion(row.get('IGST', 0)),
                    'source_type': str(row.get('TYPE', '')),
                    'status': db_status,
                    'matched_invoice': '',
                    'matched_gstin': '',
                })

        if records:
            save_db.bulk_insert(records)

        # ── Save template-structured raw rows ─────────────────────────────
        _raw_rows_by_type = {}
        # ITC original rows (template column names)
        if not self._original_itc_raw.empty:
            _itc_date_cols = [c for c in self._original_itc_raw.columns
                              if any(k in c.lower() for k in ('booking date', 'invoice date'))]
            _itc_date_col  = _itc_date_cols[0] if _itc_date_cols else None
            _itc_rows = []
            for _, row in self._original_itc_raw.iterrows():
                _dv = row[_itc_date_col] if _itc_date_col else None
                _yr, _mo, _ = self._parse_invoice_date(_dv)
                _itc_rows.append((row.to_dict(), _yr, _mo))
            _raw_rows_by_type['ITC'] = _itc_rows
        # Per-type 2B original rows
        for _ft, _fdf in self._original_2b_tables.items():
            if _fdf.empty:
                continue
            _period_cols = [c for c in _fdf.columns
                            if c.lower().strip() in ('period', 'invoice date')]
            _period_col  = _period_cols[0] if _period_cols else None
            _type_rows   = []
            for _, row in _fdf.iterrows():
                _dv = row[_period_col] if _period_col else None
                _yr, _mo, _ = self._parse_invoice_date(_dv)
                _type_rows.append((row.to_dict(), _yr, _mo))
            _raw_rows_by_type[_ft] = _type_rows
        if _raw_rows_by_type:
            save_db.save_raw_rows(_raw_rows_by_type,
                                  datetime.now().strftime('%Y-%m-%d'))

        # ── Also persist any debug-matched pairs ──────────────────────────
        debug_saved = 0
        for mp in self._debug_final_matches:
            itc_norm  = normalize_gstin(mp['itc_gstin'])  + '|' + normalize_invoice(mp['itc_invoice'])
            twoa_norm = normalize_gstin(mp['twoa_gstin']) + '|' + normalize_invoice(mp['twoa_invoice'])
            yr, mo, date_str = self._parse_invoice_date(mp.get('itc_date') or mp.get('twoa_date'))
            itc_rec = {
                'record_type': 'ITC',
                'gstin': mp['itc_gstin'],
                'invoice_number': mp['itc_invoice'],
                'invoice_norm': itc_norm,
                'invoice_date': mp.get('itc_date', ''),
                'invoice_month': mo,
                'invoice_year': yr,
                'taxable_value': mp.get('itc_taxable', 0),
                'cgst': mp.get('itc_cgst', 0),
                'sgst': mp.get('itc_sgst', 0),
                'igst': mp.get('itc_igst', 0),
                'source_type': 'ITC',
                'status': 'Matched',
                'matched_invoice': mp['twoa_invoice'],
                'matched_gstin': mp['twoa_gstin'],
            }
            tyr, tmo, tdate_str = self._parse_invoice_date(
                mp.get('twoa_date') or mp.get('twoa_booking_month'))
            twoa_rec = {
                'record_type': '2A',
                'gstin': mp['twoa_gstin'],
                'invoice_number': mp['twoa_invoice'],
                'invoice_norm': twoa_norm,
                'invoice_date': tdate_str,
                'invoice_month': tmo,
                'invoice_year': tyr,
                'taxable_value': mp.get('twoa_taxable', 0),
                'cgst': mp.get('twoa_cgst', 0),
                'sgst': mp.get('twoa_sgst', 0),
                'igst': mp.get('twoa_igst', 0),
                'source_type': mp.get('twoa_type', '2A'),
                'status': 'Matched',
                'matched_invoice': mp['itc_invoice'],
                'matched_gstin': mp['itc_gstin'],
            }
            save_db.upsert_matched_pair(itc_rec, twoa_rec)
            debug_saved += 1
        self._debug_final_matches.clear()

        save_db.close()

        if records or debug_saved:
            parts = []
            if records:
                parts.append(f"{len(records)} records")
            if debug_saved:
                parts.append(f"{debug_saved} debug-matched pair(s)")
            summary = " + ".join(parts)
            messagebox.showinfo("Saved",
                f"{summary} saved to database '{save_db_label}'.")
            self.log(f"Saved {summary} to YTD database '{save_db_label}'.")
        else:
            messagebox.showinfo("Nothing to Save", "No records to save.")

    def match_with_past_2a(self):
        """Match current ITC against unmatched 2A records stored in a YTD database."""
        if self.original_itc is None or self.itc_register is None:
            messagebox.showwarning("No Data", "Run a reconciliation first before matching with past 2B data.")
            return

        # Pick a database
        os.makedirs(GST_DB_DIR, exist_ok=True)
        db_files = sorted(f[:-3] for f in os.listdir(GST_DB_DIR) if f.endswith('.db'))
        if not db_files:
            messagebox.showwarning("No Databases", "No YTD databases found. Save a reconciliation first.")
            return

        pick_win = ctk.CTkToplevel(self)
        pick_win.title("Select Database")
        pick_win.geometry("360x280")
        pick_win.grab_set(); pick_win.lift(); pick_win.focus_force()

        ctk.CTkLabel(pick_win, text="Select a database to match against:",
                     font=ctk.CTkFont(size=13, weight="bold")).pack(pady=(18, 6))

        chosen = [None]
        lf = ctk.CTkScrollableFrame(pick_win, height=160)
        lf.pack(fill="x", padx=16)
        for name in db_files:
            ctk.CTkButton(lf, text=name, anchor="w",
                          fg_color="transparent", hover_color="#E3F2FD",
                          text_color="#333333", height=30,
                          command=lambda n=name: [chosen.__setitem__(0, n), pick_win.destroy()]
                          ).pack(fill="x", pady=1)

        ctk.CTkButton(pick_win, text="Cancel", command=pick_win.destroy,
                      fg_color="#757575", hover_color="#616161", width=100).pack(pady=10)
        self.wait_window(pick_win)

        if not chosen[0]:
            return

        db_path = os.path.join(GST_DB_DIR, chosen[0] + '.db')
        db = GSTDatabaseManager(db_path)
        unmatched = db.get_unmatched_2a()

        if not unmatched:
            messagebox.showinfo("Nothing to Match", f"No unmatched 2B records in '{chosen[0]}'.")
            db.close()
            return

        self.log(f"Matching {len(unmatched)} unmatched 2B records from '{chosen[0]}'...")

        # Build a merged_df from the unmatched DB records
        past_merged = pd.DataFrame([{
            'GSTN':            r['gstin'],
            'Document_number': r['invoice_number'],
            'CGST':            r['cgst'],
            'SGST':            r['sgst'],
            'IGST':            r['igst'],
            'Taxable_value':   r['taxable_value'],
            'TYPE':            r['source_type'],
            'Invoice_date':    r['invoice_date'],
            'Booking_Month':   '',
        } for r in unmatched])

        try:
            past_gstr_2a = create_gstr_2a(past_merged, self.log)
            past_itc_result_df, _ = create_itc_result(
                self.original_itc, self.itc_register,
                gstr_2a=past_gstr_2a, comparison_df=None,
                merged_df=past_merged, log_callback=self.log
            )
        except Exception as e:
            messagebox.showerror("Error", f"Matching failed: {e}")
            db.close()
            return

        # Build matched pairs from past_itc_result_df
        gstin_col = self._find_col(past_itc_result_df, 'vendor', 'gstn') or self._find_col(past_itc_result_df, 'gstin')
        inv_col   = self._find_col(past_itc_result_df, 'vendor inv') or self._find_col(past_itc_result_df, 'external doc')

        _MNAMES = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'May',6:'Jun',
                   7:'Jul',8:'Aug',9:'Sep',10:'Oct',11:'Nov',12:'Dec'}
        db_by_norm = {r['invoice_norm']: r for r in unmatched}
        db_by_gi   = {(normalize_gstin(r['gstin']), normalize_invoice(r['invoice_number'])): r
                      for r in unmatched}

        matched_statuses = {'Matched', 'Higher in 2B', 'Lower in 2B'}
        match_pairs = []
        for _, row in past_itc_result_df.iterrows():
            st = str(row.get('Status', ''))
            if st not in matched_statuses:
                continue
            itc_inv   = str(row[inv_col]).strip()   if inv_col   else ''
            itc_gstin = str(row[gstin_col]).strip() if gstin_col else ''
            itc_norm  = normalize_gstin(itc_gstin) + '|' + normalize_invoice(itc_inv)
            twoa_inv   = str(row.get('2B Invoice No', ''))
            twoa_gstin = str(row.get('2B GSTIN', ''))
            db_rec = db_by_norm.get(
                normalize_gstin(twoa_gstin) + '|' + normalize_invoice(twoa_inv)
            ) or db_by_gi.get((normalize_gstin(twoa_gstin), normalize_invoice(twoa_inv)))
            db_yr = db_rec['invoice_year']  if db_rec else None
            db_mo = db_rec['invoice_month'] if db_rec else None
            month_year = (f"{_MNAMES.get(db_mo, '')}-{db_yr}"
                          if db_yr and db_mo else '')
            match_pairs.append({
                'db_rec':    db_rec,
                'itc_inv':   itc_inv,
                'itc_gstin': itc_gstin,
                'itc_norm':  itc_norm,
                'twoa_inv':  twoa_inv,
                'twoa_gstin':twoa_gstin,
                'status':    st,
                'month_year':month_year,
                '2a_cgst':   row.get('CGST as per 2B', 0),
                '2a_sgst':   row.get('SGST as per 2B', 0),
                '2a_igst':   row.get('IGST as per 2B', 0),
                '2a_type':   row.get('Type', ''),
                '2a_bm':     row.get('Booking Month as per GSTR-2B', ''),
            })

        if not match_pairs:
            db.close()
            self.log("Past 2B match: no matches found.")
            messagebox.showinfo("No Matches", f"No ITC invoices matched against '{chosen[0]}'.")
            return

        # ── Interactive results dialog ──────────────────────────────────────
        res_win = ctk.CTkToplevel(self)
        res_win.title(f"Match Results — {chosen[0]}")
        res_win.geometry("960x540")
        res_win.grab_set(); res_win.lift(); res_win.focus_force()

        ctk.CTkLabel(res_win,
            text=f"{len(match_pairs)} match(es) found against '{chosen[0]}'. Select rows to save:",
            font=ctk.CTkFont(size=13, weight="bold")).pack(pady=(14, 4), padx=14, anchor="w")

        hdr_f = ctk.CTkFrame(res_win, fg_color="#E8F5E9", corner_radius=4)
        hdr_f.pack(fill="x", padx=14, pady=(0, 2))
        for txt, w in [('', 36), ('Books Invoice', 210), ('Books GSTIN', 200),
                       ('2B Invoice (DB)', 200), ('Month-Year', 110), ('Status', 150)]:
            ctk.CTkLabel(hdr_f, text=txt, font=ctk.CTkFont(size=10, weight="bold"),
                         width=w, anchor="w").pack(side="left", padx=4, pady=4)

        scroll = ctk.CTkScrollableFrame(res_win, fg_color="white")
        scroll.pack(fill="both", expand=True, padx=14, pady=4)

        check_vars = []
        for mp in match_pairs:
            row_f = ctk.CTkFrame(scroll, fg_color="transparent")
            row_f.pack(fill="x", pady=1)
            var = tk.BooleanVar(value=True)
            check_vars.append(var)
            ctk.CTkCheckBox(row_f, text='', variable=var, width=36).pack(side="left", padx=4)
            ctk.CTkLabel(row_f, text=mp['itc_inv'],    width=210, anchor="w").pack(side="left", padx=4)
            ctk.CTkLabel(row_f, text=mp['itc_gstin'],  width=200, anchor="w").pack(side="left", padx=4)
            ctk.CTkLabel(row_f, text=mp['twoa_inv'],   width=200, anchor="w").pack(side="left", padx=4)
            ctk.CTkLabel(row_f, text=mp['month_year'], width=110, anchor="w").pack(side="left", padx=4)
            ctk.CTkLabel(row_f, text=mp['status'],     width=150, anchor="w").pack(side="left", padx=4)

        def _update_itc_result_for_ytd(mp):
            if self.itc_result_df is None:
                return
            df = self.itc_result_df
            gc = self._find_col(df, 'vendor', 'gstn') or self._find_col(df, 'gstin')
            ic = self._find_col(df, 'vendor inv') or self._find_col(df, 'external doc')
            if not gc or not ic:
                return
            mask = df.apply(
                lambda r: (normalize_gstin(str(r[gc])) + '|' + normalize_invoice(str(r[ic]))) == mp['itc_norm'],
                axis=1)
            if not mask.any():
                return
            remark = (f"matched with {mp['twoa_inv']} of {mp['month_year']}"
                      if mp['month_year'] else f"matched with {mp['twoa_inv']} from YTD database")
            df.loc[mask, 'Status']                       = 'Matched'
            df.loc[mask, '2B Invoice No']                = mp['twoa_inv']
            df.loc[mask, '2B GSTIN']                     = mp['twoa_gstin']
            df.loc[mask, 'CGST as per 2B']               = mp['2a_cgst']
            df.loc[mask, 'SGST as per 2B']               = mp['2a_sgst']
            df.loc[mask, 'IGST as per 2B']               = mp['2a_igst']
            df.loc[mask, 'Type']                         = 'ytd database'
            df.loc[mask, 'Booking Month as per GSTR-2B'] = mp['2a_bm']
            if 'Remarks' not in df.columns:
                df['Remarks'] = ''
            df.loc[mask, 'Remarks'] = remark

        def _save_selected():
            saved = 0
            for mp, var in zip(match_pairs, check_vars):
                if not var.get():
                    continue
                if mp['db_rec']:
                    db.update_status(mp['db_rec']['id'], 'twoa_invoices', 'Matched',
                                     matched_invoice=mp['itc_inv'],
                                     matched_gstin=mp['itc_gstin'])
                _update_itc_result_for_ytd(mp)
                saved += 1
            db.close()
            res_win.destroy()
            self.log(f"Past 2B match: saved {saved} of {len(match_pairs)} match(es).")
            messagebox.showinfo("Saved", f"Saved {saved} match(es) to '{chosen[0]}'.")

        def _cancel_match():
            db.close()
            res_win.destroy()

        btn_f = ctk.CTkFrame(res_win, fg_color="transparent")
        btn_f.pack(pady=8)
        ctk.CTkButton(btn_f, text="Save Selected", command=_save_selected,
                      fg_color="#2E7D32", hover_color="#1B5E20", width=140).pack(side="left", padx=8)
        ctk.CTkButton(btn_f, text="Cancel", command=_cancel_match,
                      fg_color="#757575", hover_color="#616161", width=100).pack(side="left", padx=8)

    def match_with_past_itc(self):
        """Match current 2B data against unmatched ITC records stored in a YTD database.
        Reverse of match_with_past_2a: the DB supplies the ITC side, the current run supplies the 2B side."""
        if self.merged_df is None or self.merged_df.empty:
            messagebox.showwarning(
                "No Data",
                "Run a reconciliation first before matching with past ITC data.")
            return

        # ── Pick a database ───────────────────────────────────────────────
        os.makedirs(GST_DB_DIR, exist_ok=True)
        db_files = sorted(f[:-3] for f in os.listdir(GST_DB_DIR) if f.endswith('.db'))
        if not db_files:
            messagebox.showwarning("No Databases",
                                   "No YTD databases found. Save a reconciliation first.")
            return

        pick_win = ctk.CTkToplevel(self)
        pick_win.title("Select Database")
        pick_win.geometry("360x280")
        pick_win.grab_set(); pick_win.lift(); pick_win.focus_force()

        ctk.CTkLabel(pick_win, text="Select a database to match against:",
                     font=ctk.CTkFont(size=13, weight="bold")).pack(pady=(18, 6))

        chosen = [None]
        lf = ctk.CTkScrollableFrame(pick_win, height=160)
        lf.pack(fill="x", padx=16)
        for name in db_files:
            ctk.CTkButton(lf, text=name, anchor="w",
                          fg_color="transparent", hover_color="#E0F2F1",
                          text_color="#333333", height=30,
                          command=lambda n=name: [chosen.__setitem__(0, n), pick_win.destroy()]
                          ).pack(fill="x", pady=1)

        ctk.CTkButton(pick_win, text="Cancel", command=pick_win.destroy,
                      fg_color="#757575", hover_color="#616161", width=100).pack(pady=10)
        self.wait_window(pick_win)

        if not chosen[0]:
            return

        db_path = os.path.join(GST_DB_DIR, chosen[0] + '.db')
        db = GSTDatabaseManager(db_path)
        unmatched_itc = db.get_unmatched_itc()

        if not unmatched_itc:
            messagebox.showinfo("Nothing to Match",
                                f"No unmatched ITC records in '{chosen[0]}'.")
            db.close()
            return

        self.log(f"Matching {len(unmatched_itc)} unmatched ITC records from '{chosen[0]}'...")

        # ── Build a fake ITC DataFrame from the DB records ────────────────
        # Column names must match what create_itc_register expects
        past_itc_df = pd.DataFrame([{
            'Vendor GSTN':    r['gstin'],
            'Vendor Inv. No.': r['invoice_number'],
            'Taxable Value':  r['taxable_value'],
            'CGST':           r['cgst'],
            'SGST':           r['sgst'],
            'IGST':           r['igst'],
        } for r in unmatched_itc])

        try:
            past_itc_register  = create_itc_register(past_itc_df, self.log)
            current_gstr_2a    = create_gstr_2a(self.merged_df, self.log)
            past_itc_result_df, _ = create_itc_result(
                past_itc_df, past_itc_register,
                gstr_2a=current_gstr_2a, comparison_df=None,
                merged_df=self.merged_df, log_callback=self.log,
            )
        except Exception as e:
            messagebox.showerror("Error", f"Matching failed: {e}")
            db.close()
            return

        # ── Find matched rows in the result ───────────────────────────────
        gstin_col = (self._find_col(past_itc_result_df, 'vendor', 'gstn')
                     or self._find_col(past_itc_result_df, 'gstin'))
        inv_col   = (self._find_col(past_itc_result_df, 'vendor inv')
                     or self._find_col(past_itc_result_df, 'external doc'))

        _MNAMES = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'May',6:'Jun',
                   7:'Jul',8:'Aug',9:'Sep',10:'Oct',11:'Nov',12:'Dec'}
        db_by_norm = {r['invoice_norm']: r for r in unmatched_itc}
        db_by_gi   = {(normalize_gstin(r['gstin']), normalize_invoice(r['invoice_number'])): r
                      for r in unmatched_itc}

        matched_statuses = {'Matched', 'Higher in 2B', 'Lower in 2B'}
        match_pairs = []
        for _, row in past_itc_result_df.iterrows():
            st = str(row.get('Status', ''))
            if st not in matched_statuses:
                continue

            # Past ITC side (from DB)
            itc_inv   = str(row[inv_col]).strip()   if inv_col   else ''
            itc_gstin = str(row[gstin_col]).strip() if gstin_col else ''
            itc_norm  = normalize_gstin(itc_gstin) + '|' + normalize_invoice(itc_inv)

            db_rec = (db_by_norm.get(itc_norm)
                      or db_by_gi.get((normalize_gstin(itc_gstin), normalize_invoice(itc_inv))))
            db_yr = db_rec['invoice_year']  if db_rec else None
            db_mo = db_rec['invoice_month'] if db_rec else None
            month_year = (f"{_MNAMES.get(db_mo, '')}-{db_yr}"
                          if db_yr and db_mo else '')

            # Current 2B side (matched column from result — stored as "2A Invoice No")
            twob_inv   = str(row.get('2B Invoice No', ''))
            twob_gstin = str(row.get('2B GSTIN', ''))

            match_pairs.append({
                'db_rec':     db_rec,
                'itc_inv':    itc_inv,
                'itc_gstin':  itc_gstin,
                'itc_norm':   itc_norm,
                'twob_inv':   twob_inv,
                'twob_gstin': twob_gstin,
                'status':     st,
                'month_year': month_year,
                'itc_cgst':   row.get('CGST', 0),
                'itc_sgst':   row.get('SGST', 0),
                'itc_igst':   row.get('IGST', 0),
            })

        if not match_pairs:
            db.close()
            self.log("Past ITC match: no matches found.")
            messagebox.showinfo("No Matches",
                                f"No 2B invoices matched against ITC records in '{chosen[0]}'.")
            return

        # ── Interactive results dialog ────────────────────────────────────
        res_win = ctk.CTkToplevel(self)
        res_win.title(f"Match Results — {chosen[0]}")
        res_win.geometry("960x540")
        res_win.grab_set(); res_win.lift(); res_win.focus_force()

        ctk.CTkLabel(res_win,
            text=f"{len(match_pairs)} match(es) found against Books records in '{chosen[0]}'. Select rows to save:",
            font=ctk.CTkFont(size=13, weight="bold")).pack(pady=(14, 4), padx=14, anchor="w")

        hdr_f = ctk.CTkFrame(res_win, fg_color="#E0F2F1", corner_radius=4)
        hdr_f.pack(fill="x", padx=14, pady=(0, 2))
        for txt, w in [('', 36), ('Past Books Invoice', 200), ('Past Books GSTIN', 190),
                       ('Current 2B Invoice', 200), ('Books Period', 110), ('Status', 150)]:
            ctk.CTkLabel(hdr_f, text=txt, font=ctk.CTkFont(size=10, weight="bold"),
                         width=w, anchor="w").pack(side="left", padx=4, pady=4)

        scroll = ctk.CTkScrollableFrame(res_win, fg_color="white")
        scroll.pack(fill="both", expand=True, padx=14, pady=4)

        check_vars = []
        for mp in match_pairs:
            row_f = ctk.CTkFrame(scroll, fg_color="transparent")
            row_f.pack(fill="x", pady=1)
            var = tk.BooleanVar(value=True)
            check_vars.append(var)
            ctk.CTkCheckBox(row_f, text='', variable=var, width=36).pack(side="left", padx=4)
            ctk.CTkLabel(row_f, text=mp['itc_inv'],    width=200, anchor="w").pack(side="left", padx=4)
            ctk.CTkLabel(row_f, text=mp['itc_gstin'],  width=190, anchor="w").pack(side="left", padx=4)
            ctk.CTkLabel(row_f, text=mp['twob_inv'],   width=200, anchor="w").pack(side="left", padx=4)
            ctk.CTkLabel(row_f, text=mp['month_year'], width=110, anchor="w").pack(side="left", padx=4)
            ctk.CTkLabel(row_f, text=mp['status'],     width=150, anchor="w").pack(side="left", padx=4)

        def _update_2a_result_for_itc_match(mp):
            """Mark the matched 2B row in gstr_2a_results_df as Matched."""
            if self.gstr_2a_results_df is None:
                return
            df = self.gstr_2a_results_df
            twob_norm = normalize_gstin(mp['twob_gstin']) + '|' + normalize_invoice(mp['twob_inv'])
            mask = df.apply(
                lambda r: (normalize_gstin(str(r.get('GSTN', ''))) + '|'
                           + normalize_invoice(str(r.get('Document_number', '')))) == twob_norm,
                axis=1,
            )
            if not mask.any():
                return
            remark = (f"matched with past ITC {mp['itc_inv']} of {mp['month_year']}"
                      if mp['month_year'] else f"matched with past ITC {mp['itc_inv']}")
            df.loc[mask, 'Status'] = 'Matched'
            if 'Remarks' not in df.columns:
                df['Remarks'] = ''
            df.loc[mask, 'Remarks'] = remark

        def _save_selected():
            saved = 0
            for mp, var in zip(match_pairs, check_vars):
                if not var.get():
                    continue
                if mp['db_rec']:
                    db.update_status(
                        mp['db_rec']['id'], 'itc_invoices', 'Matched',
                        matched_invoice=mp['twob_inv'],
                        matched_gstin=mp['twob_gstin'],
                    )
                _update_2a_result_for_itc_match(mp)
                saved += 1
            db.close()
            res_win.destroy()
            self.log(f"Past ITC match: saved {saved} of {len(match_pairs)} match(es).")
            messagebox.showinfo("Saved", f"Saved {saved} match(es) to '{chosen[0]}'.")

        def _cancel_match():
            db.close()
            res_win.destroy()

        btn_f = ctk.CTkFrame(res_win, fg_color="transparent")
        btn_f.pack(pady=8)
        ctk.CTkButton(btn_f, text="Save Selected", command=_save_selected,
                      fg_color="#00695C", hover_color="#004D40", width=140).pack(side="left", padx=8)
        ctk.CTkButton(btn_f, text="Cancel", command=_cancel_match,
                      fg_color="#757575", hover_color="#616161", width=100).pack(side="left", padx=8)

    def open_ytd_database_window(self):
        """Open the YTD Database manager — 3-panel browser (DBs | Year/Month | Invoices)."""
        os.makedirs(GST_DB_DIR, exist_ok=True)

        MONTH_NAMES = {1:'January',2:'February',3:'March',4:'April',5:'May',6:'June',
                       7:'July',8:'August',9:'September',10:'October',11:'November',12:'December'}

        win = ctk.CTkToplevel(self)
        win.title("YTD Database Manager")
        win.geometry("1200x700")
        win.lift()
        win.focus_force()
        win.after(100, lambda: win.attributes('-topmost', False))
        win.attributes('-topmost', True)

        # ── Header ────────────────────────────────────────────────────────
        hdr = ctk.CTkFrame(win, fg_color="#1B5E20", corner_radius=0)
        hdr.pack(fill="x")
        ctk.CTkLabel(hdr, text="YTD Database Manager",
                     font=ctk.CTkFont(size=15, weight="bold"),
                     text_color="white").pack(side="left", padx=16, pady=10)
        co_parts = [p for p in [self._company_name, self._company_gst,
                                 self._company_period] if p]
        if co_parts:
            ctk.CTkLabel(hdr, text="  |  " + "  ·  ".join(co_parts),
                         font=ctk.CTkFont(size=11), text_color="#A5D6A7").pack(side="left")
        ctk.CTkLabel(hdr, text=f"  |  {GST_DB_DIR}",
                     font=ctk.CTkFont(size=9), text_color="#81C784").pack(side="left", padx=8)

        # ── 3-pane body ───────────────────────────────────────────────────
        body = ctk.CTkFrame(win, fg_color="#F0F2F5", corner_radius=0)
        body.pack(fill="both", expand=True)

        # ── PANE 1: Database list (220px) ─────────────────────────────────
        p1 = ctk.CTkFrame(body, fg_color="white", width=220, corner_radius=0,
                           border_width=1, border_color="#E0E0E0")
        p1.pack(side="left", fill="y")
        p1.pack_propagate(False)

        ctk.CTkLabel(p1, text="DATABASES",
                     font=ctk.CTkFont(size=9, weight="bold"),
                     text_color="#9E9E9E").pack(anchor="w", padx=14, pady=(12, 4))

        db_list_scroll = ctk.CTkScrollableFrame(p1, fg_color="transparent")
        db_list_scroll.pack(fill="both", expand=True, padx=4)

        _state = {'db': None, 'db_name': None, 'year': None, 'month': None}
        _sel_db_btn   = [None]
        _sel_period_btn = [None]
        _row_id_map   = {}

        def _new_db_dialog():
            nd = ctk.CTkToplevel(win)
            nd.title("New Database")
            nd.geometry("340x220")
            nd.grab_set(); nd.lift(); nd.focus_force()
            ctk.CTkLabel(nd, text="Database name:",
                         font=ctk.CTkFont(size=13)).pack(pady=(20, 5))
            nv = tk.StringVar()
            ctk.CTkEntry(nd, textvariable=nv, width=200).pack(pady=5)
            def _create():
                name = nv.get().strip()
                if not name:
                    return
                fn = name if name.endswith('.db') else name + '.db'
                path = os.path.join(GST_DB_DIR, fn)
                if os.path.exists(path):
                    messagebox.showwarning("Exists", f"'{fn}' already exists.", parent=nd)
                    return
                GSTDatabaseManager(path).close()
                nd.destroy()
                _load_db_list()
            bf = ctk.CTkFrame(nd, fg_color="transparent"); bf.pack(pady=10)
            ctk.CTkButton(bf, text="Create", command=_create,
                          fg_color="#2E7D32", hover_color="#1B5E20", width=100).pack(side="left", padx=6)
            ctk.CTkButton(bf, text="Cancel", command=nd.destroy,
                          fg_color="#757575", hover_color="#616161", width=100).pack(side="left", padx=6)

        ctk.CTkButton(p1, text="+ New Database", command=_new_db_dialog,
                      fg_color="#2E7D32", hover_color="#1B5E20",
                      height=30, font=ctk.CTkFont(size=11)
                      ).pack(fill="x", padx=8, pady=(0, 2))

        def _delete_db():
            import gc
            if not _state['db']:
                messagebox.showwarning("No Selection", "Select a database first.", parent=win)
                return
            name = _state['db_name']
            path = _state['db'].db_path
            if not messagebox.askyesno("Confirm Delete",
                    f"Permanently delete '{name}'?\nThis cannot be undone.", parent=win):
                return
            # Close viewer connection
            _state['db'].close()
            _state['db'] = None
            _state['db_name'] = None
            _state['year'] = None
            _state['month'] = None
            _sel_db_btn[0] = None
            # Also close main app's connection if it points to the same file
            if self.db_manager and os.path.abspath(self.db_manager.db_path) == os.path.abspath(path):
                self.db_manager.close()
                self.db_manager = None
                self.db_name = None
            gc.collect()  # ensure SQLite file handle is released on Windows
            try:
                os.remove(path)
            except PermissionError as e:
                messagebox.showerror("Cannot Delete",
                    f"File is still in use by another process.\n{e}", parent=win)
                return
            for _t in _tab_trees.values():
                for _ti in _t.get_children(): _t.delete(_ti)
            _row_id_map.clear()
            for w in period_scroll.winfo_children(): w.destroy()
            breadcrumb_lbl.configure(text="Select a database  ›  then a period")
            count_lbl.configure(text="")
            _load_db_list()

        ctk.CTkButton(p1, text="Delete Database", command=_delete_db,
                      fg_color="#C62828", hover_color="#B71C1C",
                      height=30, font=ctk.CTkFont(size=11)
                      ).pack(fill="x", padx=8, pady=(0, 6))

        # ── PANE 2: Year / Month nav (190px) ──────────────────────────────
        p2 = ctk.CTkFrame(body, fg_color="white", width=190, corner_radius=0,
                           border_width=1, border_color="#E0E0E0")
        p2.pack(side="left", fill="y")
        p2.pack_propagate(False)

        p2_title = ctk.CTkLabel(p2, text="PERIOD",
                                font=ctk.CTkFont(size=9, weight="bold"),
                                text_color="#9E9E9E")
        p2_title.pack(anchor="w", padx=14, pady=(12, 4))

        period_scroll = ctk.CTkScrollableFrame(p2, fg_color="transparent")
        period_scroll.pack(fill="both", expand=True, padx=4)

        # ── PANE 3: Invoice table (flex) ───────────────────────────────────
        p3 = ctk.CTkFrame(body, fg_color="white", corner_radius=0)
        p3.pack(side="left", fill="both", expand=True)

        p3_hdr = ctk.CTkFrame(p3, fg_color="transparent")
        p3_hdr.pack(fill="x", padx=14, pady=(10, 2))
        breadcrumb_lbl = ctk.CTkLabel(p3_hdr,
                                      text="Select a database  ›  then a period",
                                      font=ctk.CTkFont(size=13, weight="bold"),
                                      text_color="#333333")
        breadcrumb_lbl.pack(side="left")
        count_lbl = ctk.CTkLabel(p3_hdr, text="",
                                 font=ctk.CTkFont(size=11), text_color="#9E9E9E")
        count_lbl.pack(side="right")

        # Status filter
        frow = ctk.CTkFrame(p3, fg_color="transparent")
        frow.pack(fill="x", padx=14, pady=(0, 4))
        ctk.CTkLabel(frow, text="Status:", font=ctk.CTkFont(size=11)).pack(side="left")
        status_var = tk.StringVar(value="All")
        ctk.CTkOptionMenu(frow, variable=status_var, values=["All", "Matched", "Unmatched"],
                          width=110, command=lambda _: _load_invoices()).pack(side="left", padx=6)

        # Sheet-type tab configs  (rt=record_type, st=source_type, raw=use ytd_raw_data JSON)
        # Column names match Template updated.xlsx exactly; 'Saved Date' appended by loader.
        _TAB_CONFIGS = {
            'All': {
                'rt': None, 'st': None, 'raw': False,
                'cols':   ('Type','GSTIN','Invoice No','Date','Taxable','CGST','SGST','IGST','Source','Status','Matched With','Saved Date'),
                'widths': (45,    145,    165,         90,    85,       70,    70,    70,    60,      90,      165,          75),
            },
            'ITC- SR': {
                'rt': 'ITC', 'st': None, 'raw': True,
                'cols': (
                    'Sr. No.','My GSTIN','Type of Credit','Vendor GSTIN Status',
                    'Vendor  GSTN ','Vendor Name','Booking Date',
                    'Vendor Inv. No/  External Doc no','Invoice Date','Type of voucher',
                    'Taxable  Value','IGST Amount','CGST Amount','SGST Amount','cess','Total tax',
                    'Description of goods/service',
                    'Whether Eligible for ITC\nYes/No/Partially',
                    'Type of inward supply: Import, RCM, All other ITC',
                    'Month of Availment in GSTR-3B','3B TABLE','3B MONTH','Date of Payment',
                    '2B MONTH','Diff Taxable value','Diff IGST','Diff CGST','Diff SGST',
                    'GSTR 2B Invoice No\n(After consolidation of all 2Bs)',
                    'Common ITC?',
                    'Whether Credit is reversed in Subsequent Month (Yes/No)',
                    'Month Of reversal','Amount of Reversal','Reason For Reversal',
                    'Whether reavailed?','Reavailment GSTR-3B month','Reason For Reavailment',
                    'Whether reavailed in FY 2024-25 or 2025-26?',
                    'Whether Inward supply of 2023-24 but availed in 2024-25 within prescribed time limit\n(Yes/No)',
                    'Whether Inward supply of 2023-24 and  Availed  in 2023-24\n(Yes/No)',
                    'Whether Inward supply of 2024-25 but Carried Fowarded and Availed in 2025-26\n(Yes/No)',
                    'Saved Date',
                ),
                'widths': (
                    55,120,110,130,145,130,95,175,95,105,
                    105,85,85,85,60,75,
                    140,
                    100,
                    120,
                    110,75,85,95,85,90,75,75,75,
                    160,
                    80,90,85,85,110,85,100,110,100,
                    100,100,100,
                    75,
                ),
            },
            'B2B- 2B': {
                'rt': '2A', 'st': 'B2B', 'raw': True,
                'cols': (
                    'Period','My GSTIN','Trade/Legal name','GSTIN','POS','Invoice Type',
                    'Reverse Charge','Invoice No','Invoice Date','Invoice Value',
                    'Taxable Value','Tax Rate','IGST','CGST','SGST','Cess',
                    'GSTR-1/5 Filling Period','GSTR-1/5 Filling Date',
                    'ITC Availibility','Reason','Source Type',
                    'E-Invoice Applicable','IRN','IRN Generated Date','Saved Date',
                ),
                'widths': (80,130,145,145,55,95,85,155,90,100,100,70,75,75,75,60,110,110,100,95,80,95,170,110,75),
            },
            'B2BA- 2B': {
                'rt': '2A', 'st': 'B2BA', 'raw': True,
                'cols': (
                    'Period','My GSTIN','Trade/Legal name','GSTIN','POS','Invoice Type',
                    'Reverse Charge','Invoice No','Invoice Date',
                    'Revised Invoice No','Revised Invoice Date',
                    'Invoice Value','Taxable Value','Tax Rate','IGST','CGST','SGST','Cess',
                    'GSTR-1/5 Filling Period','GSTR-1/5 Filling Date',
                    'ITC Availibility','Reason','E-Invoice Applicable','Saved Date',
                ),
                'widths': (80,130,145,145,55,95,85,155,90,155,100,100,100,70,75,75,75,60,110,110,100,95,95,75),
            },
            'CDNR- 2B': {
                'rt': '2A', 'st': 'CDNR', 'raw': True,
                'cols': (
                    'Period','My GSTIN','Trade/Legal name','GSTIN','POS','Invoice Type',
                    'Reverse Charge','Note Type','Invoice No','Invoice Date',
                    'Invoice Value','Taxable Value','Tax Rate','IGST','CGST','SGST','Cess',
                    'GSTR-1/5 Filling Period','GSTR-1/5 Filling Date',
                    'ITC Availibility','Reason','Source Type',
                    'E-Invoice Applicable','IRN','IRN Generated Date','Saved Date',
                ),
                'widths': (80,130,145,145,55,95,85,85,155,90,100,100,70,75,75,75,60,110,110,100,95,80,95,170,110,75),
            },
            'CDNRA- 2B': {
                'rt': '2A', 'st': 'CDNRA', 'raw': True,
                'cols': (
                    'Period','My GSTIN','Trade/Legal name','GSTIN','POS','Invoice Type',
                    'Reverse Charge','Note Type','Note No','Invoice Date',
                    'Revised Note No','Revised Note Date',
                    'Invoice Value','Taxable Value','Tax Rate','IGST','CGST','SGST','Cess',
                    'GSTR-1/5 Filling Period','GSTR-1/5 Filling Date',
                    'ITC Availibility','Reason','E-Invoice Applicable','Saved Date',
                ),
                'widths': (80,130,145,145,55,95,85,85,155,90,155,100,100,100,70,75,75,75,60,110,110,100,95,95,75),
            },
            'IMPG- 2B': {
                'rt': '2A', 'st': 'IMPG', 'raw': True,
                'cols': (
                    'My GSTIN','Period','Icegate Reference Date','Port Code',
                    'BOE No','Invoice Date','Taxable Value','IGST','Cess',
                    'Amended (Yes)','Saved Date',
                ),
                'widths': (130,80,120,80,155,90,105,80,60,80,75),
            },
            'IMPGSEZ- 2B': {
                'rt': '2A', 'st': 'IMPGSEZ', 'raw': True,
                'cols': (
                    'My GSTIN','Period','Trade/Legal name','GSTIN',
                    'Icegate Reference Date','Port Code','BOE No','Invoice Date',
                    'Taxable Value',' IGST ','Cess','Amended (Yes)','Saved Date',
                ),
                'widths': (130,80,145,145,120,80,155,90,105,80,60,80,75),
            },
        }
        _TAB_ORDER = ['All', 'ITC- SR', 'B2B- 2B', 'B2BA- 2B', 'CDNR- 2B', 'CDNRA- 2B', 'IMPG- 2B', 'IMPGSEZ- 2B']

        # CTkTabview for sheet types
        _sheet_tabs = ctk.CTkTabview(p3, fg_color="#F8F8F8", height=500)
        _sheet_tabs.pack(fill="both", expand=True, padx=4, pady=(0, 2))
        for _tn in _TAB_ORDER:
            _sheet_tabs.add(_tn)

        # Build one Treeview per tab
        style = ttk.Style()
        style.configure("YTD.Treeview", rowheight=22, font=('Helvetica', 10))
        style.configure("YTD.Treeview.Heading", font=('Helvetica', 10, 'bold'))
        style.map("YTD.Treeview", background=[('selected', '#A5D6A7')])

        _tab_trees = {}   # tab_name → ttk.Treeview
        for _tn in _TAB_ORDER:
            _tab = _sheet_tabs.tab(_tn)
            _cfg = _TAB_CONFIGS[_tn]
            _tf = ctk.CTkFrame(_tab, fg_color="transparent")
            _tf.pack(fill="both", expand=True, padx=4, pady=4)
            _vsb = ttk.Scrollbar(_tf, orient="vertical")
            _hsb = ttk.Scrollbar(_tf, orient="horizontal")
            _tv = ttk.Treeview(_tf, columns=_cfg['cols'], show='headings',
                               style="YTD.Treeview",
                               yscrollcommand=_vsb.set, xscrollcommand=_hsb.set,
                               selectmode='extended')
            _vsb.configure(command=_tv.yview)
            _hsb.configure(command=_tv.xview)
            _vsb.pack(side="right", fill="y")
            _hsb.pack(side="bottom", fill="x")
            _tv.pack(fill="both", expand=True)
            for _col, _w in zip(_cfg['cols'], _cfg['widths']):
                _tv.heading(_col, text=_col, anchor='w')
                _tv.column(_col, width=_w, anchor='w', stretch=False)
            _tv.tag_configure('matched',   background='#C8E6C9', foreground='#1B5E20')
            _tv.tag_configure('unmatched', background='#FFFFFF')
            _tv.tag_configure('frozen',    background='#E3F2FD', foreground='#0D47A1')
            _tab_trees[_tn] = _tv

        def _active_tree():
            return _tab_trees.get(_sheet_tabs.get(), _tab_trees['All'])

        # Keep legacy 'tree' alias pointing to the active tree for buttons
        tree = _tab_trees['All']
        _sheet_tabs.configure(command=lambda: _load_invoices())

        # Action buttons
        bbar = ctk.CTkFrame(p3, fg_color="transparent")
        bbar.pack(fill="x", padx=8, pady=(0, 8))

        def _raw_val(row_dict, col_name):
            """Look up a JSON dict by exact key, then by stripped/lowered key."""
            v = row_dict.get(col_name)
            if v is not None:
                return str(v) if str(v) != 'nan' else ''
            # Fuzzy fallback: strip whitespace + lowercase
            col_clean = col_name.strip().lower().replace('\n', ' ')
            for k, val in row_dict.items():
                if isinstance(k, str) and k.strip().lower().replace('\n', ' ') == col_clean:
                    return str(val) if str(val) != 'nan' else ''
            return ''

        def _load_invoices():
            _active_tab = _sheet_tabs.get()
            _atree = _active_tree()
            for item in _atree.get_children(): _atree.delete(item)
            _row_id_map.clear()
            if not _state['db'] or _state['year'] is None:
                return
            _cfg = _TAB_CONFIGS[_active_tab]
            _cols = _cfg['cols']
            db_lbl = _state['db_name'] or ''
            yr_lbl = str(_state['year'])
            mo_lbl = MONTH_NAMES.get(_state['month'], 'All months') if _state['month'] else 'All months'
            breadcrumb_lbl.configure(text=f"{db_lbl}  ›  {yr_lbl}  ›  {mo_lbl}")

            if _cfg.get('raw'):
                # Type-specific tab: read full-column JSON from ytd_raw_data
                raw_rows = _state['db'].get_raw_rows_by_type(
                    _cfg['st'] or _cfg['rt'],
                    year=_state['year'], month=_state['month'])
                for rd in raw_rows:
                    vals = tuple(
                        rd['__saved_date__'] if col == 'Saved Date' else _raw_val(rd, col)
                        for col in _cols
                    )
                    tag = 'frozen' if rd.get('__frozen__') else 'unmatched'
                    iid = _atree.insert('', 'end', tags=(tag,), values=vals)
                    _row_id_map[iid] = ('raw', rd['__id__'])
                count_lbl.configure(text=f"{len(raw_rows)} records")
            else:
                # All tab: read normalized rows with status
                st = status_var.get() if status_var.get() != "All" else None
                rows = _state['db'].get_invoices(record_type=None,
                                                 year=_state['year'],
                                                 month=_state['month'])
                if st:
                    rows = [r for r in rows if r['status'] == st]
                _ytd_matched = {'Matched', 'Matched but invoice number is not accurate'}
                for r in rows:
                    tag = 'matched' if r['status'] in _ytd_matched else 'unmatched'
                    iid = _atree.insert('', 'end', tags=(tag,), values=(
                        r['record_type'], r['gstin'], r['invoice_number'], r['invoice_date'],
                        f"{r['taxable_value']:,.2f}", f"{r['cgst']:,.2f}",
                        f"{r['sgst']:,.2f}", f"{r['igst']:,.2f}",
                        r['source_type'], r['status'],
                        r.get('matched_invoice') or '', (r.get('saved_date') or '')[:10],
                    ))
                    _row_id_map[iid] = (r['_table'], r['id'])
                count_lbl.configure(text=f"{len(rows)} records")

        def _load_period_nav():
            for w in period_scroll.winfo_children(): w.destroy()
            if not _state['db']:
                return
            periods = _state['db'].get_distinct_periods()
            if not periods:
                ctk.CTkLabel(period_scroll, text="No data yet",
                             font=ctk.CTkFont(size=11), text_color="#9E9E9E").pack(pady=20)
                return
            from collections import defaultdict as _dd
            by_year = _dd(list)
            for yr, mo in periods:
                if yr and mo:
                    by_year[yr].append(mo)
            for yr in sorted(by_year.keys(), reverse=True):
                yr_card = ctk.CTkFrame(period_scroll, fg_color="#F5F5F5", corner_radius=6)
                yr_card.pack(fill="x", pady=(4, 1))
                ctk.CTkLabel(yr_card, text=str(yr),
                             font=ctk.CTkFont(size=12, weight="bold"),
                             text_color="#333333").pack(anchor="w", padx=10, pady=4)
                for mo in sorted(by_year[yr]):
                    def _pick(y=yr, m=mo):
                        if _sel_period_btn[0]:
                            _sel_period_btn[0].configure(fg_color="transparent",
                                                          text_color="#333333")
                        _state['year'] = y
                        _state['month'] = m
                        _load_invoices()
                    btn = ctk.CTkButton(period_scroll,
                                        text=f"  {MONTH_NAMES[mo]}",
                                        fg_color="transparent", hover_color="#E3F2FD",
                                        text_color="#333333", anchor="w",
                                        height=30, corner_radius=6,
                                        font=ctk.CTkFont(size=11))
                    btn.configure(command=lambda b=btn, fn=_pick: [
                        _sel_period_btn.__setitem__(0, b),
                        b.configure(fg_color="#E3F2FD", text_color="#1565C0"),
                        fn()])
                    btn.pack(fill="x", pady=1, padx=6)

        def _select_db(db_path, db_name, btn):
            if _sel_db_btn[0]:
                _sel_db_btn[0].configure(fg_color="transparent", text_color="#333333")
            _sel_db_btn[0] = btn
            btn.configure(fg_color="#E8F5E9", text_color="#1B5E20")
            if _state['db']:
                _state['db'].close()
            _state['db']      = GSTDatabaseManager(db_path)
            _state['db_name'] = db_name
            _state['year']    = None
            _state['month']   = None
            _load_period_nav()
            for _t in _tab_trees.values():
                for _ti in _t.get_children(): _t.delete(_ti)
            _row_id_map.clear()
            breadcrumb_lbl.configure(text=f"{db_name}  ›  select a period")
            count_lbl.configure(text="")

        def _load_db_list():
            for w in db_list_scroll.winfo_children(): w.destroy()
            dbs = sorted(f for f in os.listdir(GST_DB_DIR) if f.endswith('.db'))
            if not dbs:
                ctk.CTkLabel(db_list_scroll, text="No databases yet",
                             font=ctk.CTkFont(size=11), text_color="#9E9E9E").pack(pady=20)
                return
            for fn in dbs:
                name = fn[:-3]
                path = os.path.join(GST_DB_DIR, fn)
                btn  = ctk.CTkButton(db_list_scroll, text=name,
                                     fg_color="transparent", hover_color="#E8F5E9",
                                     text_color="#333333", anchor="w",
                                     height=32, corner_radius=6,
                                     font=ctk.CTkFont(size=12))
                btn.configure(command=lambda p=path, n=name, b=btn: _select_db(p, n, b))
                btn.pack(fill="x", pady=2)

        # ── Action button callbacks ────────────────────────────────────────
        def _is_frozen_sel(sel):
            yr, mo = _state.get('year'), _state.get('month')
            if yr and mo and _state['db'] and _state['db'].is_month_frozen(yr, mo):
                return True
            return False

        def _mark_matched():
            _atree = _active_tree()
            if _is_frozen_sel(_atree.selection()):
                messagebox.showwarning("Frozen",
                    "This month is frozen. Unfreeze it first.", parent=win)
                return
            for iid in _atree.selection():
                tbl, rid = _row_id_map[iid]
                _state['db'].update_status(rid, tbl, 'Matched')
            _load_invoices()

        def _mark_unmatched():
            _atree = _active_tree()
            if _is_frozen_sel(_atree.selection()):
                messagebox.showwarning("Frozen",
                    "This month is frozen. Unfreeze it first.", parent=win)
                return
            for iid in _atree.selection():
                tbl, rid = _row_id_map[iid]
                _state['db'].update_status(rid, tbl, 'Unmatched', '', '')
            _load_invoices()

        def _delete_sel():
            _atree = _active_tree()
            sel = _atree.selection()
            if not sel:
                return
            if _is_frozen_sel(sel):
                messagebox.showwarning("Frozen",
                    "This month is frozen. Unfreeze it first to delete records.", parent=win)
                return
            locked = [i for i in sel if _atree.item(i, 'tags')[0] == 'matched']
            free   = [i for i in sel if i not in locked]
            if locked:
                messagebox.showwarning("Locked",
                    f"{len(locked)} Matched record(s) cannot be deleted.", parent=win)
            if not free:
                return
            if not messagebox.askyesno("Confirm", f"Delete {len(free)} record(s)?", parent=win):
                return
            for iid in free:
                tbl, rid = _row_id_map[iid]
                _state['db'].delete_invoice(rid, tbl)
            _load_invoices()
            _load_period_nav()

        # ── Columns that the user may edit when they re-upload ────────────────
        EDITABLE_COLS  = ['Status', 'Matched With', 'Matched GSTIN']
        SYSTEM_COLS    = ['DB_ID', 'DB_Table']
        VALID_STATUSES = ['Matched', 'Unmatched']

        def _rows_to_df(rows):
            """Convert DB rows to a DataFrame.  DB_ID / DB_Table are the re-import keys."""
            return pd.DataFrame([{
                'DB_ID':          r.get('id', ''),
                'DB_Table':       r.get('_table', ''),
                'Type':           r['record_type'],
                'GSTIN':          r['gstin'],
                'Invoice No':     r['invoice_number'],
                'Invoice Date':   r['invoice_date'],
                'Year':           r.get('invoice_year', ''),
                'Month':          MONTH_NAMES.get(r.get('invoice_month'), ''),
                'Taxable Value':  r['taxable_value'],
                'CGST':           r['cgst'],
                'SGST':           r['sgst'],
                'IGST':           r['igst'],
                'Source':         r['source_type'],
                'Status':         r['status'],
                'Matched With':   r['matched_invoice'],
                'Matched GSTIN':  r['matched_gstin'],
                'Saved Date':     (r.get('saved_date') or '')[:10],
            } for r in rows])

        def _write_formatted_excel(writer, df, sheet_name):
            """Write df to a named sheet with colour-coded columns and Status validation."""
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            from openpyxl.worksheet.datavalidation import DataValidation

            df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]

            HDR_FILL    = PatternFill("solid", fgColor="1A237E")   # dark navy
            EDIT_FILL   = PatternFill("solid", fgColor="FFF9C4")   # pale yellow
            SYS_FILL    = PatternFill("solid", fgColor="ECEFF1")   # light grey
            HDR_FONT    = Font(bold=True, color="FFFFFF", size=11)
            EDIT_HDR    = Font(bold=True, color="F57F17", size=11)  # amber for editable headers
            NORMAL_FONT = Font(size=10)
            BOLD_FONT   = Font(bold=True, size=10)
            center      = Alignment(horizontal="center", vertical="center", wrap_text=True)
            left        = Alignment(horizontal="left", vertical="center")
            thin        = Border(
                left=Side(style='thin', color='BDBDBD'),
                right=Side(style='thin', color='BDBDBD'),
                top=Side(style='thin', color='BDBDBD'),
                bottom=Side(style='thin', color='BDBDBD'))

            col_names = list(df.columns)
            for ci, cname in enumerate(col_names, start=1):
                cell = ws.cell(row=1, column=ci)
                if cname in EDITABLE_COLS:
                    cell.fill = PatternFill("solid", fgColor="F9A825")
                    cell.font = EDIT_HDR
                elif cname in SYSTEM_COLS:
                    cell.fill = SYS_FILL
                    cell.font = Font(bold=True, color="757575", size=11)
                else:
                    cell.fill = HDR_FILL
                    cell.font = HDR_FONT
                cell.alignment = center
                cell.border = thin

            max_row = ws.max_row
            for ri in range(2, max_row + 1):
                for ci, cname in enumerate(col_names, start=1):
                    cell = ws.cell(row=ri, column=ci)
                    cell.border = thin
                    cell.alignment = left
                    cell.font = NORMAL_FONT
                    if cname in EDITABLE_COLS:
                        cell.fill = EDIT_FILL
                        cell.font = BOLD_FONT
                    elif cname in SYSTEM_COLS:
                        cell.fill = SYS_FILL

            # Status dropdown validation
            status_ci = col_names.index('Status') + 1 if 'Status' in col_names else None
            if status_ci:
                col_letter = get_column_letter(status_ci)
                dv = DataValidation(
                    type="list",
                    formula1=f'"{",".join(VALID_STATUSES)}"',
                    allow_blank=False,
                    showDropDown=False,
                    showErrorMessage=True,
                    errorTitle="Invalid Status",
                    error=f"Allowed values: {', '.join(VALID_STATUSES)}")
                dv.sqref = f"{col_letter}2:{col_letter}{max_row}"
                ws.add_data_validation(dv)

            # Auto-fit column widths
            for ci, cname in enumerate(col_names, start=1):
                max_len = max(len(str(cname)),
                              max((len(str(ws.cell(r, ci).value or '')) for r in range(2, max_row+1)),
                                  default=0))
                ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 40)

            ws.row_dimensions[1].height = 28
            ws.freeze_panes = "A2"

        def _add_instructions_sheet(writer):
            """Add a plain-English Instructions sheet to the workbook."""
            from openpyxl.styles import PatternFill, Font, Alignment
            ws = writer.book.create_sheet("Instructions", 0)  # insert first
            ws.sheet_properties.tabColor = "F9A825"
            lines = [
                ("GST Reconciliation Tool — YTD Data Edit Guide", True,  "1A237E", 14),
                ("",                                              False, None,    11),
                ("HOW TO EDIT THIS FILE",                        True,  "E65100", 12),
                ("1. Only edit the YELLOW columns: Status, Matched With, Matched GSTIN.", False, None, 11),
                ("2. Valid Status values (use the dropdown): Matched  |  Unmatched",      False, None, 11),
                ("3. Do NOT change DB_ID or DB_Table — these are used to identify each", False, None, 11),
                ("   record in the database and are required for re-upload.",             False, None, 11),
                ("4. Do NOT add or delete rows.",                                          False, None, 11),
                ("5. Save the file after editing.",                                        False, None, 11),
                ("",                                                                       False, None, 11),
                ("HOW TO RE-UPLOAD",                                                       True,  "2E7D32", 12),
                ("1. Open the YTD Database Manager in the tool.",                         False, None, 11),
                ("2. Select the relevant database from the left panel.",                   False, None, 11),
                ("3. Click  ⬆ Re-upload / Amend  and select this edited file.",           False, None, 11),
                ("4. The tool will compare every row against the database and update only",False, None, 11),
                ("   the records where Status, Matched With, or Matched GSTIN changed.",  False, None, 11),
                ("5. A summary will show how many records were updated.",                  False, None, 11),
                ("",                                                                       False, None, 11),
                ("COLOUR GUIDE",                                                            True,  "555555", 12),
                ("  Yellow columns  →  editable by you",                                  False, None, 11),
                ("  Grey columns    →  DB_ID / DB_Table (do not edit)",                   False, None, 11),
                ("  Dark blue cols  →  reference data (do not edit)",                     False, None, 11),
            ]
            for i, (text, bold, color, size) in enumerate(lines, start=1):
                cell = ws.cell(row=i, column=1, value=text)
                cell.font = Font(bold=bold, color=(color or "222222"), size=size)
                cell.alignment = Alignment(wrap_text=True)
            ws.column_dimensions['A'].width = 72

        # ── Template sheet name mapping ───────────────────────────────────
        _SHEET_NAME_MAP = {
            'ITC':     'ITC- SR',
            'B2B':     'B2B- 2B',
            'B2BA':    'B2BA- 2B',
            'CDNR':    'CDNR- 2B',
            'CDNRA':   'CDNRA- 2B',
            'IMPG':    'IMPG- 2B',
            'IMPGSEZ': 'IMPGSEZ- 2B',
        }
        _SHEET_ORDER = ['ITC', 'B2B', 'B2BA', 'CDNR', 'CDNRA', 'IMPG', 'IMPGSEZ']

        def _write_template_sheets(writer, raw_rows):
            """Write per-type raw rows as separate template-named sheets."""
            from collections import defaultdict
            by_type = defaultdict(list)
            for r in raw_rows:
                try:
                    row_dict = json.loads(r['row_json'])
                except Exception:
                    row_dict = {}
                by_type[r['data_type']].append(row_dict)
            written = 0
            for _ft in _SHEET_ORDER:
                if _ft not in by_type:
                    continue
                _df = pd.DataFrame(by_type[_ft])
                if _df.empty:
                    continue
                _sname = _SHEET_NAME_MAP.get(_ft, _ft)
                _df.to_excel(writer, sheet_name=_sname[:31], index=False)
                ws = writer.sheets[_sname[:31]]
                from openpyxl.styles import PatternFill, Font, Alignment
                _hdr_fill = PatternFill("solid", fgColor="1A237E")
                _hdr_font = Font(bold=True, color="FFFFFF", size=10)
                _center   = Alignment(horizontal='center', vertical='center', wrap_text=True)
                for cell in ws[1]:
                    cell.fill = _hdr_fill
                    cell.font = _hdr_font
                    cell.alignment = _center
                ws.freeze_panes = "A2"
                for col_cells in ws.columns:
                    max_len = max(
                        (len(str(c.value or '')) for c in col_cells), default=8)
                    from openpyxl.utils import get_column_letter
                    ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 3, 38)
                written += len(by_type[_ft])
            return written

        def _write_legacy_type_sheets(writer, all_rows):
            """Split DB rows by type and write separate template-named sheets."""
            from openpyxl.styles import PatternFill, Font, Alignment
            from openpyxl.utils import get_column_letter
            _HDR_FILL = PatternFill("solid", fgColor="1A237E")
            _HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
            _CTR = Alignment(horizontal='center', vertical='center', wrap_text=True)
            # ITC sheet
            itc_rows = [r for r in all_rows if r['record_type'] == 'ITC']
            twoa_rows = [r for r in all_rows if r['record_type'] != 'ITC']
            written = 0
            def _write_sheet(row_list, sheet_name):
                nonlocal written
                if not row_list:
                    return
                _df = pd.DataFrame([{
                    'GSTIN / Vendor GSTN': r['gstin'],
                    'Vendor Name':         r.get('vendor_name', ''),
                    'Invoice No':          r['invoice_number'],
                    'Invoice Date':        r['invoice_date'],
                    'Year':                r.get('invoice_year', ''),
                    'Month':               MONTH_NAMES.get(r.get('invoice_month'), ''),
                    'Taxable Value':       r['taxable_value'],
                    'CGST':                r['cgst'],
                    'SGST':                r['sgst'],
                    'IGST':                r['igst'],
                    'Tax Rate':            r.get('tax_rate', ''),
                    'Invoice Value':       r.get('invoice_value', ''),
                    'Type':                r['source_type'],
                    'Status':              r['status'],
                    'Matched With':        r['matched_invoice'],
                    'Matched GSTIN':       r['matched_gstin'],
                    'Saved Date':          (r.get('saved_date') or '')[:10],
                } for r in row_list])
                _sn = sheet_name[:31]
                _df.to_excel(writer, sheet_name=_sn, index=False)
                ws = writer.sheets[_sn]
                for cell in ws[1]:
                    cell.fill = _HDR_FILL
                    cell.font = _HDR_FONT
                    cell.alignment = _CTR
                ws.freeze_panes = "A2"
                for col_cells in ws.columns:
                    mx = max((len(str(c.value or '')) for c in col_cells), default=8)
                    ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(mx + 3, 38)
                written += len(row_list)
            _write_sheet(itc_rows, 'ITC- SR')
            for _src, _sname in [('B2B','B2B- 2B'),('B2BA','B2BA- 2B'),
                                   ('CDNR','CDNR- 2B'),('CDNRA','CDNRA- 2B'),
                                   ('IMPG','IMPG- 2B'),('IMPGSEZ','IMPGSEZ- 2B')]:
                _write_sheet([r for r in twoa_rows if r['source_type'] == _src], _sname)
            # Any remaining 2B rows not matched to a specific type
            _others = [r for r in twoa_rows if r['source_type'] not in
                       ('B2B','B2BA','CDNR','CDNRA','IMPG','IMPGSEZ')]
            _write_sheet(_others, '2B- Other')
            return written

        def _download_to_excel():
            """Export currently viewed period — separate sheets per type."""
            if not _state['db'] or _state['year'] is None:
                messagebox.showwarning("No Data", "Select a database and period first.", parent=win)
                return
            db_lbl = _state['db_name'] or 'YTD'
            yr_lbl = str(_state['year'])
            mo_lbl = MONTH_NAMES.get(_state['month'], 'All') if _state['month'] else 'All'
            fpath = filedialog.asksaveasfilename(
                title="Save Period Data as Excel",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"{db_lbl}_{yr_lbl}_{mo_lbl}.xlsx",
                parent=win)
            if not fpath:
                return
            try:
                # Use template raw rows when available (original column names)
                if _state['db'].has_raw_data():
                    raw = _state['db'].get_raw_rows(
                        year=_state['year'], month=_state['month'])
                    if raw:
                        with pd.ExcelWriter(fpath, engine='openpyxl') as writer:
                            n = _write_template_sheets(writer, raw)
                        messagebox.showinfo("Exported",
                            f"{n} records exported — separate sheets with original template columns.\n\nFile: {fpath}",
                            parent=win)
                        return
                # Fallback: separate sheets from existing DB rows (no original columns)
                rows = _state['db'].get_invoices(year=_state['year'], month=_state['month'])
                if not rows:
                    messagebox.showinfo("No Records", "No records to export.", parent=win)
                    return
                with pd.ExcelWriter(fpath, engine='openpyxl') as writer:
                    n = _write_legacy_type_sheets(writer, rows)
                messagebox.showinfo("Exported",
                    f"{n} records exported — separate sheets (ITC- SR, B2B- 2B, etc.).\n\nFile: {fpath}",
                    parent=win)
            except Exception as ex:
                messagebox.showerror("Export Failed", str(ex), parent=win)

        def _download_full_db():
            """Export entire database — separate sheets per type."""
            if not _state['db']:
                messagebox.showwarning("No Database", "Select a database first.", parent=win)
                return
            db_lbl = _state['db_name'] or 'YTD'
            fpath = filedialog.asksaveasfilename(
                title="Save Full Database as Excel",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"{db_lbl}_Full.xlsx",
                parent=win)
            if not fpath:
                return
            try:
                # Use template raw rows when available (original column names)
                if _state['db'].has_raw_data():
                    raw = _state['db'].get_raw_rows()
                    if raw:
                        with pd.ExcelWriter(fpath, engine='openpyxl') as writer:
                            n = _write_template_sheets(writer, raw)
                        messagebox.showinfo("Exported",
                            f"{n} records exported — separate sheets with original template columns.\n\n"
                            f"Sheets created: ITC- SR, B2B- 2B, B2BA- 2B, CDNR- 2B, CDNRA- 2B, "
                            f"IMPG- 2B, IMPGSEZ- 2B (only non-empty sheets are written).\n\nFile: {fpath}",
                            parent=win)
                        return
                # Fallback: separate sheets from existing DB rows
                all_rows = _state['db'].get_invoices()
                if not all_rows:
                    messagebox.showinfo("Empty", "The selected database has no records.", parent=win)
                    return
                with pd.ExcelWriter(fpath, engine='openpyxl') as writer:
                    n = _write_legacy_type_sheets(writer, all_rows)
                messagebox.showinfo("Exported",
                    f"{n} records exported — separate sheets per type.\n\nFile: {fpath}",
                    parent=win)
            except Exception as ex:
                messagebox.showerror("Export Failed", str(ex), parent=win)

        def _reupload_amend():
            """Re-import an edited YTD Excel file and apply only the changed rows."""
            if not _state['db']:
                messagebox.showwarning("No Database",
                    "Select a database first — changes will be applied to it.", parent=win)
                return
            fpath = filedialog.askopenfilename(
                title="Select Edited YTD Excel File",
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")],
                parent=win)
            if not fpath:
                return

            try:
                xl = pd.ExcelFile(fpath)
                # Prefer "All Records" sheet; fall back to first non-Instructions sheet
                target_sheet = None
                for sn in xl.sheet_names:
                    if sn.lower() == 'all records':
                        target_sheet = sn
                        break
                if not target_sheet:
                    for sn in xl.sheet_names:
                        if sn.lower() != 'instructions' and sn.lower() != 'summary':
                            target_sheet = sn
                            break
                if not target_sheet:
                    messagebox.showerror("No Data Sheet",
                        "Could not find a data sheet in this file.\n"
                        "Expected a sheet named 'All Records'.", parent=win)
                    return

                df_edit = pd.read_excel(fpath, sheet_name=target_sheet, dtype=str).fillna('')

                # Validate required columns
                missing = [c for c in ('DB_ID', 'DB_Table', 'Status') if c not in df_edit.columns]
                if missing:
                    messagebox.showerror("Missing Columns",
                        f"The file is missing required columns: {', '.join(missing)}\n\n"
                        "Only files downloaded from this tool can be re-uploaded.\n"
                        "Make sure you are using the correct file.", parent=win)
                    return

                # Load current state from DB for comparison
                all_current = _state['db'].get_invoices()
                current_lookup = {(str(r['id']), r['_table']): r for r in all_current}

                updated = skipped = errors = 0
                change_log = []

                for _, row in df_edit.iterrows():
                    db_id    = str(row.get('DB_ID', '')).strip()
                    db_table = str(row.get('DB_Table', '')).strip()
                    if not db_id or not db_table or db_id in ('', 'nan'):
                        errors += 1
                        continue

                    current = current_lookup.get((db_id, db_table))
                    if current is None:
                        errors += 1
                        change_log.append(f"  Row DB_ID={db_id}: not found in database")
                        continue

                    new_status   = str(row.get('Status', '')).strip()
                    new_matched  = str(row.get('Matched With', '')).strip()
                    new_gstin    = str(row.get('Matched GSTIN', '')).strip()

                    # Validate status
                    if new_status not in VALID_STATUSES:
                        errors += 1
                        change_log.append(f"  Row DB_ID={db_id}: invalid Status '{new_status}' — skipped")
                        continue

                    old_status  = str(current.get('status', '')).strip()
                    old_matched = str(current.get('matched_invoice', '')).strip()
                    old_gstin   = str(current.get('matched_gstin', '')).strip()

                    if (new_status == old_status and
                            new_matched == old_matched and
                            new_gstin   == old_gstin):
                        skipped += 1
                        continue

                    # Apply the change
                    _state['db'].update_status(
                        int(db_id), db_table, new_status, new_matched, new_gstin)
                    updated += 1
                    diffs = []
                    if new_status  != old_status:  diffs.append(f"Status: {old_status!r}→{new_status!r}")
                    if new_matched != old_matched:  diffs.append(f"Matched With: {old_matched!r}→{new_matched!r}")
                    if new_gstin   != old_gstin:    diffs.append(f"Matched GSTIN: {old_gstin!r}→{new_gstin!r}")
                    inv = current.get('invoice_number', '')
                    change_log.append(f"  [{db_table}] ID {db_id} ({inv}): {', '.join(diffs)}")

                summary_msg = (
                    f"Re-upload Complete\n\n"
                    f"Records updated  : {updated}\n"
                    f"Unchanged rows   : {skipped}\n"
                    f"Errors / skipped : {errors}\n"
                )
                if change_log:
                    summary_msg += "\nChanges applied:\n" + "\n".join(change_log[:20])
                    if len(change_log) > 20:
                        summary_msg += f"\n  … and {len(change_log)-20} more"

                messagebox.showinfo("Re-upload Complete", summary_msg, parent=win)
                _load_invoices()      # refresh the grid
                _load_period_nav()

            except Exception as ex:
                messagebox.showerror("Re-upload Failed", str(ex), parent=win)

        ctk.CTkButton(bbar, text="Mark Matched",   command=_mark_matched,
                      fg_color="#2E7D32", hover_color="#1B5E20", height=30, width=130).pack(side="left", padx=4)
        ctk.CTkButton(bbar, text="Mark Unmatched", command=_mark_unmatched,
                      fg_color="#E65100", hover_color="#BF360C", height=30, width=140).pack(side="left", padx=4)
        ctk.CTkButton(bbar, text="Delete Selected", command=_delete_sel,
                      fg_color="#C62828", hover_color="#B71C1C", height=30, width=130).pack(side="left", padx=4)
        ctk.CTkButton(bbar, text="Refresh", command=_load_invoices,
                      fg_color="#607D8B", hover_color="#455A64", height=30, width=90).pack(side="left", padx=4)

        def _freeze_month():
            if not _state['db'] or _state['year'] is None or _state['month'] is None:
                messagebox.showwarning("No Selection",
                    "Select a database and a specific month first.", parent=win)
                return
            yr, mo = _state['year'], _state['month']
            mn = MONTH_NAMES.get(mo, str(mo))
            already = _state['db'].is_month_frozen(yr, mo)
            if already:
                if messagebox.askyesno("Unfreeze?",
                        f"'{mn} {yr}' is already frozen.\nUnfreeze it?", parent=win):
                    if not self._verify_ytd_creds_dialog():
                        messagebox.showwarning("Access Denied",
                            "Incorrect credentials. Month not unfrozen.", parent=win)
                        return
                    _state['db'].unfreeze_month(yr, mo)
                    _load_invoices()
                    messagebox.showinfo("Unfrozen", f"'{mn} {yr}' has been unfrozen.", parent=win)
                return
            ans = messagebox.askyesnocancel("Freeze Month",
                f"Freeze '{mn} {yr}'?\n\nHas the GSTR-3B been filed for this month?",
                parent=win)
            if ans is None:
                return
            note = f"3B Filed: {'Yes' if ans else 'No'}"
            _state['db'].freeze_month(yr, mo, note)
            _load_invoices()
            messagebox.showinfo("Frozen",
                f"'{mn} {yr}' is now frozen ({note}).\nRecords cannot be deleted or status-changed.",
                parent=win)

        ctk.CTkButton(bbar, text="❄ Freeze Month", command=_freeze_month,
                      fg_color="#0277BD", hover_color="#01579B", height=30, width=140).pack(side="left", padx=4)
        ctk.CTkButton(bbar, text="⬆ Re-upload / Amend", command=_reupload_amend,
                      fg_color="#F57F17", hover_color="#E65100", height=30, width=175).pack(side="right", padx=4)
        ctk.CTkButton(bbar, text="⬇ Download Period", command=_download_to_excel,
                      fg_color="#1565C0", hover_color="#0D47A1", height=30, width=160).pack(side="right", padx=4)
        ctk.CTkButton(bbar, text="⬇ Download Full DB", command=_download_full_db,
                      fg_color="#4A148C", hover_color="#311B92", height=30, width=165).pack(side="right", padx=4)

        # Initial population
        _load_db_list()

    def show_results_frame(self):
        """Show the results download frame"""
        self.results_frame.pack(fill="x", padx=16, pady=6, before=self.progress_frame)

    def _compute_unmatched_2a(self, merged_df, original_itc, matched_2a_status=None):
        """Find rows in merged/GSTR 2A that have no matching invoice in ITC."""
        if merged_df is None or merged_df.empty:
            return pd.DataFrame()

        # Compute normalized key for each 2A row
        merged_keys = merged_df.apply(
            lambda r: normalize_gstin(str(r.get('GSTN', ''))) + '|' + normalize_invoice(str(r.get('Document_number', ''))),
            axis=1
        )

        # Prefer the pre-computed matched_2a_status from create_itc_result, which uses the same
        # sophisticated fuzzy matching as the main reconciliation (handles year-prefix and
        # leading-zero differences like "2020-2021/676" <-> "20-21/676" or "2" <-> "02/20-21").
        if matched_2a_status is not None:
            mask_unmatched = ~merged_keys.isin(matched_2a_status.keys())
            return merged_df[mask_unmatched].reset_index(drop=True)

        # Fallback: simple normalized key comparison against ITC keys
        if original_itc is None or original_itc.empty:
            return merged_df.copy()

        vendor_gstn_col = None
        vendor_inv_col = None
        for col in original_itc.columns:
            cl = col.lower().strip()
            if 'vendor' in cl and 'gstn' in cl:
                vendor_gstn_col = col
            elif 'vendor inv' in cl or 'external doc' in cl:
                vendor_inv_col = col

        if not vendor_gstn_col or not vendor_inv_col:
            return pd.DataFrame()

        itc_keys = set()
        for _, row in original_itc.iterrows():
            key = normalize_gstin(row[vendor_gstn_col]) + '|' + normalize_invoice(row[vendor_inv_col])
            itc_keys.add(key)

        mask_unmatched = ~merged_keys.isin(itc_keys)
        return merged_df[mask_unmatched].reset_index(drop=True)

    # ── Report helpers ────────────────────────────────────────────────────────
    def _write_report_info_sheet(self, writer, report_type=''):
        """Insert a 'Report Info' sheet (index 0) with company details."""
        from openpyxl.styles import PatternFill, Font, Alignment
        ws = writer.book.create_sheet("Report Info", 0)
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 44
        hdr_fill = PatternFill(start_color="880E4F", end_color="880E4F", fill_type="solid")
        hdr_font = Font(color="FFFFFF", bold=True, size=11)
        val_font = Font(size=11)
        rows = [
            ("Company Name", self._company_name or '—'),
            ("GSTIN",        self._company_gst    or '—'),
            ("Period",       self._company_period  or '—'),
            ("Report Type",  report_type),
            ("Generated On", datetime.now().strftime('%d %b %Y  %H:%M')),
        ]
        for r, (lbl, val) in enumerate(rows, start=1):
            ca = ws.cell(row=r, column=1, value=lbl)
            cb = ws.cell(row=r, column=2, value=val)
            ca.fill = hdr_fill
            ca.font = hdr_font
            ca.alignment = Alignment(horizontal='left', vertical='center')
            cb.font = val_font
            cb.alignment = Alignment(horizontal='left', vertical='center')
            ws.row_dimensions[r].height = 22

    @staticmethod
    def _apply_status_colors(ws, df, status_col='Status'):
        """Color every data row by its Status value and bold the header row."""
        from openpyxl.styles import PatternFill, Font, Alignment
        _FILLS = {
            'Matched':          PatternFill("solid", fgColor="C8E6C9"),
            'Matched but invoice number is not accurate':
                                PatternFill("solid", fgColor="FFE0B2"),
            'Higher in 2B':     PatternFill("solid", fgColor="FFF9C4"),
            'Lower in 2B':      PatternFill("solid", fgColor="BBDEFB"),
            'Not found in 2B':  PatternFill("solid", fgColor="E1BEE7"),
            'Unmatched':        PatternFill("solid", fgColor="FFCDD2"),
            'Not Found in ITC': PatternFill("solid", fgColor="FFCDD2"),
        }
        hdr_fill = PatternFill("solid", fgColor="37474F")
        hdr_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 20
        ws.auto_filter.ref = ws.dimensions
        if status_col not in df.columns:
            _autofit_ws(ws)
            return
        max_col = len(df.columns)
        for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
            fill = _FILLS.get(str(row.get(status_col, '')))
            if fill:
                for col_idx in range(1, max_col + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill
        _autofit_ws(ws)

    # ── Numeric column helper ─────────────────────────────────────────────────
    @staticmethod
    def _coerce_numeric_cols(df):
        """Convert tax/taxable-value columns from string to float in-place."""
        NUMBER_KEYWORDS = ('cgst', 'sgst', 'igst', 'taxable', 'cess', 'tax rate',
                           'invoice value', 'total tax', 'amount', 'diff ')
        for col in df.columns:
            cl = col.lower().strip()
            if any(kw in cl for kw in NUMBER_KEYWORDS):
                df[col] = pd.to_numeric(df[col].apply(safe_numeric_conversion), errors='coerce').fillna(0)
        return df

    @staticmethod
    def _format_number_cols(ws, df):
        """Apply Excel number format '##,##,##0.00' to numeric columns."""
        from openpyxl.utils import get_column_letter
        NUMBER_KEYWORDS = ('cgst', 'sgst', 'igst', 'taxable', 'cess',
                           'invoice value', 'total tax', 'amount', 'diff ')
        for col_idx, col_name in enumerate(df.columns, start=1):
            if any(kw in col_name.lower() for kw in NUMBER_KEYWORDS):
                col_letter = get_column_letter(col_idx)
                for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value is not None:
                            cell.number_format = '##,##,##0.00'

    def download_reconciliation(self):
        """Download ITC results report with unmatched sheets"""
        download_df = self.itc_result_df if self.itc_result_df is not None and not self.itc_result_df.empty else self.comparison_df
        if download_df is None or download_df.empty:
            messagebox.showwarning("Warning", "No data available!")
            return

        try:
            filepath = filedialog.asksaveasfilename(
                title="Save ITC Results",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile="GST_ITC_Results.xlsx"
            )
            if filepath:
                output = BytesIO()
                dl = self._coerce_numeric_cols(download_df.copy())
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    dl.to_excel(writer, sheet_name='ITC_Results', index=False)
                    self._format_number_cols(writer.sheets['ITC_Results'], dl)
                    self._apply_status_colors(writer.sheets['ITC_Results'], dl)

                    # 2B Results: all rows from GSTR 2B with Status
                    if self.gstr_2a_results_df is not None and not self.gstr_2a_results_df.empty:
                        self.gstr_2a_results_df.to_excel(writer, sheet_name='2B_Results', index=False)
                        self._apply_status_colors(writer.sheets['2B_Results'], self.gstr_2a_results_df)

                    # Same Month Cancellation sheet
                    if self.same_month_cancel_df is not None and not self.same_month_cancel_df.empty:
                        self.same_month_cancel_df.to_excel(writer, sheet_name='Same_Month_Cancellation', index=False)
                        _autofit_ws(writer.sheets['Same_Month_Cancellation'])

                    # Unmatched in ITC: line items with no corresponding 2A/2B entry
                    if download_df is not None and 'Status' in download_df.columns:
                        unmatched_itc_df = download_df[
                            download_df['Status'].isin(['Unmatched', 'Not found in 2B'])
                        ].reset_index(drop=True)
                        unmatched_itc_df.to_excel(writer, sheet_name='Unmatched_in_ITC', index=False)
                        self._apply_status_colors(writer.sheets['Unmatched_in_ITC'], unmatched_itc_df)

                    # Party-wise summary and detail sheets
                    self._write_party_sheets(writer, dl)

                    self._write_report_info_sheet(writer, 'Books (ITC) Results')

                output.seek(0)
                with open(filepath, 'wb') as f:
                    f.write(output.getvalue())
                self.log(f"Books Results saved to: {filepath}")
                messagebox.showinfo("Success", f"Books Results saved successfully!\n\nLocation: {filepath}")
        except Exception as e:
            self.log(f"Error saving report: {str(e)}")
            messagebox.showerror("Error", f"Failed to save report: {str(e)}")

    def download_2a_results(self):
        """Download 2B Results (all GSTR 2B rows with Status) as Excel"""
        if self.gstr_2a_results_df is None or self.gstr_2a_results_df.empty:
            messagebox.showinfo("Info", "No 2B results available!")
            return

        try:
            filepath = filedialog.asksaveasfilename(
                title="Save 2B Results",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile="2B_Results.xlsx"
            )
            if filepath:
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    self.gstr_2a_results_df.to_excel(writer, sheet_name='2B_Results', index=False)
                    self._apply_status_colors(writer.sheets['2B_Results'], self.gstr_2a_results_df)
                    self._write_report_info_sheet(writer, '2B Results')
                output.seek(0)
                with open(filepath, 'wb') as f:
                    f.write(output.getvalue())
                self.log(f"2B Results saved to: {filepath}")
                messagebox.showinfo("Success", f"2B Results saved successfully!\n\nRecords: {len(self.gstr_2a_results_df)}\nLocation: {filepath}")
        except Exception as e:
            self.log(f"Error saving report: {str(e)}")
            messagebox.showerror("Error", f"Failed to save report: {str(e)}")

    # Ordered list of columns to show in party-wise sheets (matches Template updated.xlsx ITC-SR)
    _PARTY_WISE_COLS = [
        'Sr. No.', 'My GSTIN', 'Type of Credit', 'Vendor GSTIN Status',
        'Vendor  GSTN', 'Vendor Name', 'Booking Date',
        'Vendor Inv. No/  External Doc no', 'Invoice Date', 'Type of voucher',
        'Taxable  Value', 'IGST Amount', 'CGST Amount', 'SGST Amount',
        'cess', 'Total tax', 'Description of goods/service',
        'Whether Eligible for ITC\nYes/No/Partially',
        'Type of inward supply: Import, RCM, All other ITC',
        'Month of Availment in GSTR-3B', '3B TABLE', '3B MONTH', 'Date of Payment',
        '2B MONTH', 'Diff Taxable value', 'Diff IGST', 'Diff CGST', 'Diff SGST',
        'GSTR 2B Invoice No\n(After consolidation of all 2Bs)',
        'Common ITC?',
        'Whether Credit is reversed in Subsequent Month (Yes/No)',
        'Month Of reversal', 'Amount of Reversal', 'Reason For Reversal',
        'Whether reavailed?', 'Reavailment GSTR-3B month', 'Reason For Reavailment',
        'Whether reavailed in FY 2024-25 or 2025-26?',
        'Whether Inward supply of 2023-24 but availed in 2024-25 within prescribed time limit\n(Yes/No)',
        'Whether Inward supply of 2023-24 and  Availed  in 2023-24\n(Yes/No)',
        'Whether Inward supply of 2024-25 but Carried Fowarded and Availed in 2025-26\n(Yes/No)',
        # Reconciliation columns added by app
        'Status', 'CGST as per 2B', 'SGST as per 2B', 'IGST as per 2B',
        'Taxable Value as per 2B', 'Type', 'Booking Month as per GSTR-2B',
        'Booking Month as per ITC', '2B Invoice No', '2B GSTIN', 'Invoice Date Match',
        'GSTR-1 Month',
    ]

    def _write_party_sheets(self, writer, df):
        """Write Party_Summary and Party_Details sheets into an open ExcelWriter.
        Returns (n_companies, n_invoices) or None if vendor GSTN column not found."""
        from openpyxl.styles import PatternFill, Font, Alignment

        def _nc(s):
            return ' '.join(str(s).strip().split()).lower()

        _desired_norm = [_nc(c) for c in self._PARTY_WISE_COLS]
        _col_map = {}
        for dc in df.columns:
            n = _nc(dc)
            for i, dn in enumerate(_desired_norm):
                if n == dn and self._PARTY_WISE_COLS[i] not in _col_map:
                    _col_map[self._PARTY_WISE_COLS[i]] = dc
                    break

        display_cols = [_col_map[c] for c in self._PARTY_WISE_COLS if c in _col_map]
        if not display_cols:
            display_cols = [c for c in df.columns if not c.startswith('_')]

        vendor_gstn_col = _col_map.get('Vendor  GSTN')
        vendor_name_col = _col_map.get('Vendor Name')
        taxable_col     = _col_map.get('Taxable  Value')
        igst_col        = _col_map.get('IGST Amount')
        cgst_col        = _col_map.get('CGST Amount')
        sgst_col        = _col_map.get('SGST Amount')

        if not vendor_gstn_col:
            for col in df.columns:
                if 'vendor' in col.lower() and 'gstn' in col.lower():
                    vendor_gstn_col = col; break
        if not vendor_gstn_col:
            return None

        df = df.copy()

        def _num(col):
            return pd.to_numeric(df[col], errors='coerce').fillna(0) if col and col in df.columns else pd.Series(0, index=df.index)

        df['_taxable'] = _num(taxable_col)
        df['_igst']    = _num(igst_col)
        df['_cgst']    = _num(cgst_col)
        df['_sgst']    = _num(sgst_col)

        party_totals = (df.groupby(vendor_gstn_col)
                          .agg(_taxable=('_taxable', 'sum'),
                               _igst=('_igst', 'sum'),
                               _cgst=('_cgst', 'sum'),
                               _sgst=('_sgst', 'sum'),
                               _total=('_taxable', 'sum'),
                               _count=(vendor_gstn_col, 'count'))
                          .sort_values('_taxable', ascending=False))

        wb = writer.book
        num_fmt = '#,##0.00'

        # ── Party_Summary sheet ───────────────────────────────────────────────
        ws_sum = wb.create_sheet("Party_Summary")
        sum_hdr_fill = PatternFill("solid", fgColor="1565C0")
        sum_hdr_font = Font(color="FFFFFF", bold=True, size=11)
        sum_headers = ['#', 'Party Name', 'GSTN', '# Invoices',
                       'Total Taxable', 'Total IGST', 'Total CGST', 'Total SGST', 'Grand Total']
        for ci, h in enumerate(sum_headers, 1):
            c = ws_sum.cell(row=1, column=ci, value=h)
            c.fill = sum_hdr_fill; c.font = sum_hdr_font
            c.alignment = Alignment(horizontal='center', vertical='center')
        ws_sum.row_dimensions[1].height = 22
        ws_sum.freeze_panes = 'A2'
        alt_fills = [PatternFill("solid", fgColor="F5F5F5"), PatternFill("solid", fgColor="FFFFFF")]
        for rank, (gstn, row) in enumerate(party_totals.iterrows(), 1):
            name = ''
            if vendor_name_col and vendor_name_col in df.columns:
                m = df[df[vendor_gstn_col] == gstn][vendor_name_col]
                name = m.iloc[0] if not m.empty else ''
            grand = row['_taxable'] + row['_igst'] + row['_cgst'] + row['_sgst']
            vals = [rank, name, gstn, int(row['_count']),
                    row['_taxable'], row['_igst'], row['_cgst'], row['_sgst'], grand]
            for ci, v in enumerate(vals, 1):
                cell = ws_sum.cell(row=rank + 1, column=ci, value=v)
                cell.fill = alt_fills[rank % 2]
                if ci >= 5:
                    cell.number_format = num_fmt
                    cell.alignment = Alignment(horizontal='right')
        _autofit_ws(ws_sum)

        # ── Party_Details sheet ───────────────────────────────────────────────
        ws_det = wb.create_sheet("Party_Details")
        col_hdr_fill  = PatternFill("solid", fgColor="37474F")
        col_hdr_font  = Font(color="FFFFFF", bold=True, size=10)
        party_hdr_fill = PatternFill("solid", fgColor="1A237E")
        party_hdr_font = Font(color="FFFFFF", bold=True, size=11)
        status_fills = {
            'Matched':          PatternFill("solid", fgColor="C8E6C9"),
            'Matched but invoice number is not accurate': PatternFill("solid", fgColor="FFE0B2"),
            'Higher in 2B':     PatternFill("solid", fgColor="FFF9C4"),
            'Lower in 2B':      PatternFill("solid", fgColor="BBDEFB"),
            'Not found in 2B':  PatternFill("solid", fgColor="E1BEE7"),
            'Unmatched':        PatternFill("solid", fgColor="FFCDD2"),
        }
        for ci, col in enumerate(display_cols, 1):
            c = ws_det.cell(row=1, column=ci, value=col)
            c.fill = col_hdr_fill; c.font = col_hdr_font
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws_det.row_dimensions[1].height = 24
        ws_det.freeze_panes = 'A2'

        row_num = 2
        for rank, (gstn, _) in enumerate(party_totals.iterrows(), 1):
            party_df = df[df[vendor_gstn_col] == gstn]
            name = ''
            if vendor_name_col and vendor_name_col in df.columns:
                m = party_df[vendor_name_col]
                name = m.iloc[0] if not m.empty else ''
            total_tax = party_totals.loc[gstn, '_taxable']
            ws_det.merge_cells(start_row=row_num, start_column=1,
                               end_row=row_num, end_column=len(display_cols))
            ph = ws_det.cell(row=row_num, column=1,
                             value=f"  #{rank}  {name or gstn}   |   GSTN: {gstn}"
                                   f"   |   {len(party_df)} invoices"
                                   f"   |   Taxable Total: {total_tax:,.2f}")
            ph.fill = party_hdr_fill; ph.font = party_hdr_font
            ph.alignment = Alignment(horizontal='left', vertical='center')
            ws_det.row_dimensions[row_num].height = 20
            row_num += 1
            for _, inv_row in party_df.iterrows():
                sfill = status_fills.get(str(inv_row.get('Status', '')))
                for ci, col in enumerate(display_cols, 1):
                    cell = ws_det.cell(row=row_num, column=ci, value=inv_row.get(col, ''))
                    if sfill:
                        cell.fill = sfill
                row_num += 1
            row_num += 1  # blank row between parties
        _autofit_ws(ws_det)

        return len(party_totals), len(df)

    def download_party_wise_report(self):
        """Generate a standalone party/GSTN-wise Excel report."""
        df = self.itc_result_df
        if df is None or df.empty:
            messagebox.showwarning("No Data", "No ITC results available to generate the report.")
            return
        filepath = filedialog.asksaveasfilename(
            title="Save Party-wise Report",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            initialfile="Party_wise_GST_Report.xlsx",
        )
        if not filepath:
            return
        try:
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                result = self._write_party_sheets(writer, df.copy())
                if result is None:
                    messagebox.showerror("Error", "Could not find Vendor GSTN column in results.")
                    return
                self._write_report_info_sheet(writer, 'Party-wise Report')
            n_companies, n_invoices = result
            messagebox.showinfo(
                "Success",
                f"Party-wise report saved!\n\n"
                f"{n_companies} companies  |  {n_invoices} invoices\n"
                f"Location: {filepath}"
            )
        except Exception as e:
            self.log(f"Party-wise report error: {e}")
            messagebox.showerror("Error", f"Failed to generate report: {e}")

    def check_gst_status(self):
        """Look up GST portal status for every vendor GSTIN via gst_status.py automation."""
        import threading, queue, sys, importlib.util

        if self.itc_result_df is None or self.itc_result_df.empty:
            messagebox.showwarning("No Data", "No reconciliation results available.")
            return

        # Locate vendor GSTN column
        vendor_gstn_col = None
        for col in self.itc_result_df.columns:
            if 'vendor' in col.lower() and 'gstn' in col.lower():
                vendor_gstn_col = col; break
        if not vendor_gstn_col:
            messagebox.showerror("Error", "Vendor GSTN column not found in results.")
            return

        _INVALID_GSTINS = {'nan', 'none', 'n/a', '', '(blank)', 'blank', '-', 'na', 'null'}
        unique_gstins = sorted(set(
            g for g in (str(v).strip() for v in self.itc_result_df[vendor_gstn_col].dropna())
            if g and g.lower() not in _INVALID_GSTINS and len(g) == 15
        ))
        if not unique_gstins:
            messagebox.showwarning("No GSTINs", "No vendor GSTINs found in results.")
            return

        # ── Credentials dialog ───────────────────────────────────────────────
        dlg = ctk.CTkToplevel(self)
        dlg.title("GST Status Checker")
        dlg.geometry("500x370")
        dlg.resizable(False, False)
        dlg.attributes('-topmost', True)
        dlg.after(100, lambda: dlg.attributes('-topmost', False))
        dlg.lift(); dlg.focus_force()

        ctk.CTkLabel(dlg, text="GST Status Checker",
                     font=ctk.CTkFont(size=16, weight="bold")).pack(pady=(18, 2))
        ctk.CTkLabel(dlg, text=f"{len(unique_gstins)} unique vendor GSTINs will be checked via GST portal",
                     font=ctk.CTkFont(size=11), text_color="#666666").pack(pady=(0, 14))

        def _row(parent, label, is_pass=False):
            f = ctk.CTkFrame(parent, fg_color="transparent")
            f.pack(fill="x", padx=36, pady=5)
            ctk.CTkLabel(f, text=label, width=170, anchor="w",
                         font=ctk.CTkFont(size=12)).pack(side="left")
            var = tk.StringVar()
            e = ctk.CTkEntry(f, textvariable=var, show="*" if is_pass else "",
                             width=230, height=34)
            e.pack(side="left")
            return var

        user_var = _row(dlg, "GST Portal Username:")
        pass_var = _row(dlg, "Password:", is_pass=True)

        # Output file path
        _app_dir = os.path.dirname(os.path.abspath(__file__))
        path_var = tk.StringVar(value=os.path.join(_app_dir, "gst_data.xlsx"))
        pf = ctk.CTkFrame(dlg, fg_color="transparent")
        pf.pack(fill="x", padx=36, pady=5)
        ctk.CTkLabel(pf, text="Save results to:", width=170, anchor="w",
                     font=ctk.CTkFont(size=12)).pack(side="left")
        ctk.CTkEntry(pf, textvariable=path_var, width=168, height=34).pack(side="left", padx=(0, 6))
        ctk.CTkButton(pf, text="Browse", width=58, height=34,
                      command=lambda: path_var.set(
                          filedialog.asksaveasfilename(
                              title="Save GST data", defaultextension=".xlsx",
                              filetypes=[("Excel", "*.xlsx")], initialfile="gst_data.xlsx"
                          ) or path_var.get()
                      )).pack(side="left")

        def _start():
            uname = user_var.get().strip()
            pwd   = pass_var.get().strip()
            fpath = path_var.get().strip()
            if not uname or not pwd:
                messagebox.showwarning("Missing", "Please enter username and password.", parent=dlg)
                return
            if not fpath:
                messagebox.showwarning("Missing", "Please specify output file path.", parent=dlg)
                return
            dlg.destroy()
            _run_check(uname, pwd, fpath)

        ctk.CTkButton(dlg, text="Start GST Status Check",
                      command=_start, fg_color="#2E7D32", hover_color="#1B5E20",
                      font=ctk.CTkFont(size=13, weight="bold"),
                      height=42, width=230).pack(pady=18)

        # ── Runner + progress window ─────────────────────────────────────────
        def _run_check(username, password, file_path):
            # Write input file
            try:
                pd.DataFrame({'GSTN': unique_gstins}).to_excel(file_path, index=False)
            except Exception as exc:
                messagebox.showerror("Error", f"Could not create input file:\n{exc}")
                return

            # Progress window — force to front
            prog = ctk.CTkToplevel(self)
            prog.title("GST Status Check — Running")
            prog.geometry("720x460")
            prog.attributes('-topmost', True)
            prog.lift()
            prog.focus_force()
            prog.after(300, lambda: prog.attributes('-topmost', False))
            prog.protocol("WM_DELETE_WINDOW", lambda: None)  # block accidental close

            ctk.CTkLabel(prog, text="GST Status Check Running",
                         font=ctk.CTkFont(size=14, weight="bold")).pack(pady=(14, 2))
            ctk.CTkLabel(prog,
                         text="Chrome will open.  Solve the CAPTCHA and click Login — "
                              "the script continues automatically.",
                         font=ctk.CTkFont(size=11), text_color="#E65100",
                         wraplength=660).pack(pady=(0, 8))

            log_box = tk.Text(prog, height=16, font=("Consolas", 9),
                              bg="#1C1C1C", fg="#AAFFAA", state="disabled",
                              relief="flat", bd=0)
            log_box.pack(padx=14, pady=(0, 8), fill="both", expand=True)

            status_lbl = ctk.CTkLabel(prog, text="Status: starting…",
                                      font=ctk.CTkFont(size=11))
            status_lbl.pack(pady=(0, 4))

            apply_btn = ctk.CTkButton(
                prog, text="Apply Results → Update Vendor GSTIN Status",
                fg_color="#1565C0", hover_color="#0D47A1",
                font=ctk.CTkFont(size=12, weight="bold"),
                height=40, width=320, state="disabled",
                command=lambda: _apply_results(file_path, prog))
            apply_btn.pack(pady=(0, 14))

            log_q: queue.Queue = queue.Queue()

            class _Writer:
                def __init__(self, q): self.q = q
                def write(self, m):
                    if m.strip(): self.q.put(m.rstrip())
                def flush(self): pass

            def _append_log(msg):
                log_box.configure(state="normal")
                log_box.insert("end", msg + "\n")
                log_box.see("end")
                log_box.configure(state="disabled")

            def _poll():
                while not log_q.empty():
                    _append_log(log_q.get_nowait())
                if _t.is_alive():
                    prog.after(400, _poll)
                else:
                    while not log_q.empty():
                        _append_log(log_q.get_nowait())
                    status_lbl.configure(
                        text="Done! Click 'Apply Results' to update the template.")
                    apply_btn.configure(state="normal")
                    prog.protocol("WM_DELETE_WINDOW", prog.destroy)

            def _worker():
                orig = sys.stdout
                sys.stdout = _Writer(log_q)
                try:
                    _script = os.path.join(_app_dir, "gst_status.py")
                    if not os.path.exists(_script):
                        log_q.put(f"ERROR: gst_status.py not found at {_script}")
                        return
                    spec = importlib.util.spec_from_file_location("gst_status", _script)
                    mod  = importlib.util.module_from_spec(spec)
                    spec.loader.exec_module(mod)
                    mod.gst_automation(file_path, username, password)
                except Exception as exc:
                    log_q.put(f"ERROR: {exc}")
                finally:
                    sys.stdout = orig

            _t = threading.Thread(target=_worker, daemon=True)
            _t.start()
            prog.after(400, _poll)

        # ── Apply results back to itc_result_df ──────────────────────────────
        def _apply_results(file_path, prog_win):
            import re

            try:
                res_df = pd.read_excel(file_path)
            except Exception as exc:
                messagebox.showerror("Error", f"Could not read results file:\n{exc}",
                                     parent=prog_win)
                return

            if 'GST Status' not in res_df.columns:
                messagebox.showerror("Error",
                                     "'GST Status' column not found in results file.",
                                     parent=prog_win)
                return

            def _is_real(val):
                s = str(val).strip().lower()
                return s and s not in ('nan', '') and 'see row' not in s \
                       and 'no data' not in s

            def _tax_period(filing_str):
                """Extract 'Tax period' value from a filing detail string."""
                m = re.search(r'[Tt]ax\s+period\s*=\s*([^\n,]+)', str(filing_str))
                return m.group(1).strip() if m else ''

            # Build GSTN-keyed maps from results file
            status_map  = {}   # GSTN → GST Status
            gstr3b_map  = {}   # GSTN → GSTR-3B tax period
            gstr1_map   = {}   # GSTN → GSTR-1 tax period

            for _, r in res_df.iterrows():
                gstn = normalize_gstin(str(r.get('GSTN', '')).strip())
                if not gstn:
                    continue

                # GST Status
                st = str(r.get('GST Status', '')).strip()
                if _is_real(st):
                    status_map[gstn] = st

                # GSTR-3B tax period
                fd3b = r.get('Filing details for GSTR3B', '')
                if _is_real(fd3b):
                    tp = _tax_period(fd3b)
                    if tp:
                        gstr3b_map[gstn] = tp

                # GSTR-1 tax period (try GSTR1 first, fall back to GSTR1A)
                fd1  = r.get('Filing details for GSTR1',  '')
                fd1a = r.get('Filing details for GSTR1A', '')
                tp1  = _tax_period(fd1)  if _is_real(fd1)  else ''
                tp1a = _tax_period(fd1a) if _is_real(fd1a) else ''
                tp_final = tp1 or tp1a
                if tp_final:
                    gstr1_map[gstn] = tp_final

            # ── Locate or create target columns in itc_result_df ─────────────
            def _find_col(keywords_must, keywords_exclude=None):
                for col in self.itc_result_df.columns:
                    cl = col.lower().strip()
                    if all(k in cl for k in keywords_must):
                        if not keywords_exclude or not any(k in cl for k in keywords_exclude):
                            return col
                return None

            # "Vendor GSTIN Status"
            vendor_status_col = _find_col(['vendor', 'gstin', 'status'])
            if not vendor_status_col:
                vendor_status_col = 'Vendor GSTIN Status'
                self.itc_result_df[vendor_status_col] = ''

            # "3B MONTH" — use exact phrase to avoid matching 'Month of Availment in GSTR-3B'
            gstr3b_col = _find_col(['3b month'])
            if not gstr3b_col:
                gstr3b_col = '3B MONTH'
                self.itc_result_df[gstr3b_col] = ''

            # "GSTR-1 Month"
            gstr1_col = _find_col(['gstr-1 month']) or _find_col(['gstr-1', 'month'])
            if not gstr1_col:
                gstr1_col = 'GSTR-1 Month'
                self.itc_result_df[gstr1_col] = ''

            # ── Apply to every row keyed by vendor GSTN ───────────────────────
            st_count = tb_count = g1_count = 0
            for idx, row in self.itc_result_df.iterrows():
                key = normalize_gstin(str(row.get(vendor_gstn_col, '')).strip())
                if not key:
                    continue
                if key in status_map:
                    self.itc_result_df.at[idx, vendor_status_col] = status_map[key]
                    st_count += 1
                if key in gstr3b_map:
                    self.itc_result_df.at[idx, gstr3b_col] = gstr3b_map[key]
                    tb_count += 1
                if key in gstr1_map:
                    self.itc_result_df.at[idx, gstr1_col] = gstr1_map[key]
                    g1_count += 1

            self.log(f"GST Status applied: status={st_count}, 3B month={tb_count}, GSTR-1 month={g1_count}")
            messagebox.showinfo(
                "Done",
                f"Results applied from GST portal data:\n\n"
                f"  Vendor GSTIN Status  →  {st_count} rows updated\n"
                f"  3B MONTH             →  {tb_count} rows updated\n"
                f"  GSTR-1 Month         →  {g1_count} rows updated\n\n"
                f"Re-open 'View Results' or download the Books report to see changes.",
                parent=prog_win)
            prog_win.destroy()

    def find_debug_candidate_pairs(self):
        """Find ITC-2A pairs where GSTIN and tax amounts match but invoice numbers only partially match."""
        candidates = []

        if self.itc_result_df is None or self.itc_result_df.empty:
            return candidates
        if self.unmatched_2a_df is None or self.unmatched_2a_df.empty:
            return candidates

        # Find ITC columns
        vendor_gstn_col = None
        vendor_inv_col = None
        for col in self.itc_result_df.columns:
            col_lower = col.lower().strip()
            if 'vendor' in col_lower and 'gstn' in col_lower:
                vendor_gstn_col = col
            elif 'vendor inv' in col_lower or 'external doc' in col_lower:
                vendor_inv_col = col

        itc_cgst_col, itc_sgst_col, itc_igst_col = find_tax_amount_columns(self.itc_result_df)

        if not vendor_gstn_col or not vendor_inv_col:
            return candidates

        # Find taxable value and invoice date columns in ITC
        taxable_col = None
        itc_date_col = None
        for col in self.itc_result_df.columns:
            if 'taxable' in col.lower() and 'value' in col.lower() and not taxable_col:
                taxable_col = col
            if col.lower().strip() == 'invoice date' and not itc_date_col:
                itc_date_col = col

        # Filter to only truly unmatched ITC rows (not found in 2A at all)
        unmatched_itc = self.itc_result_df[
            self.itc_result_df['Status'].isin(['Unmatched', 'Not found in 2B'])
        ]

        if unmatched_itc.empty:
            return candidates

        # Build 2A lookup from unmatched 2A rows only, grouped by normalized GSTIN
        from collections import defaultdict
        twoa_by_gstin = defaultdict(list)
        for idx, row in self.unmatched_2a_df.iterrows():
            twoa_by_gstin[normalize_gstin(str(row.get('GSTN', '')))].append({
                'idx': idx,
                'norm_inv': normalize_invoice(str(row.get('Document_number', ''))),
                'raw_inv': str(row.get('Document_number', '')),
                'gstn': str(row.get('GSTN', '')),
                'cgst': safe_numeric_conversion(row.get('CGST', 0)),
                'sgst': safe_numeric_conversion(row.get('SGST', 0)),
                'igst': safe_numeric_conversion(row.get('IGST', 0)),
                'tax': safe_numeric_conversion(row.get('TAX', 0)),
                'date': str(row.get('Invoice_Date', '')),
                'booking_month': str(row.get('Booking_Month', '')),
                'type': str(row.get('TYPE', '')),
            })

        used_2a = set()

        for itc_idx, itc_row in unmatched_itc.iterrows():
            itc_gstin_norm = normalize_gstin(str(itc_row[vendor_gstn_col]))
            itc_inv_norm = normalize_invoice(str(itc_row[vendor_inv_col]))
            itc_cgst = safe_numeric_conversion(itc_row.get(itc_cgst_col, 0)) if itc_cgst_col else 0
            itc_sgst = safe_numeric_conversion(itc_row.get(itc_sgst_col, 0)) if itc_sgst_col else 0
            itc_igst = safe_numeric_conversion(itc_row.get(itc_igst_col, 0)) if itc_igst_col else 0
            itc_taxable = safe_numeric_conversion(itc_row.get(taxable_col, 0)) if taxable_col else 0

            best = None
            best_sim = 0.0

            for cand in twoa_by_gstin.get(itc_gstin_norm, []):
                if cand['idx'] in used_2a:
                    continue
                # Taxable value, IGST, CGST, SGST must match within Rs 10 tolerance
                if (abs(itc_taxable - cand['tax']) > 10 or
                    abs(itc_igst - cand['igst']) > 10 or
                    abs(itc_cgst - cand['cgst']) > 10 or
                    abs(itc_sgst - cand['sgst']) > 10):
                    continue
                # Invoice similarity >= 10%
                sim = similarity(itc_inv_norm, cand['norm_inv'])
                if sim >= 0.1 and sim > best_sim:
                    best_sim = sim
                    best = cand

            if best is not None:
                used_2a.add(best['idx'])
                candidates.append({
                    'itc_index': itc_idx,
                    'itc_invoice': str(itc_row[vendor_inv_col]),
                    'itc_gstin': str(itc_row[vendor_gstn_col]),
                    'itc_date': str(itc_row[itc_date_col]) if itc_date_col else '',
                    'itc_taxable': itc_taxable,
                    'itc_cgst': itc_cgst,
                    'itc_sgst': itc_sgst,
                    'itc_igst': itc_igst,
                    'twoa_invoice': best['raw_inv'],
                    'twoa_gstin': best['gstn'],
                    'twoa_date': best['date'],
                    'twoa_taxable': best['tax'],
                    'twoa_cgst': best['cgst'],
                    'twoa_sgst': best['sgst'],
                    'twoa_igst': best['igst'],
                    'twoa_booking_month': best['booking_month'],
                    'twoa_type': best['type'],
                    'similarity': best_sim,
                    'itc_inv_norm': itc_inv_norm,
                })

        candidates.sort(key=lambda x: x['similarity'], reverse=True)
        return candidates

    def open_combined_debug_window(self):
        """Open a combined Debug Tools window with Invoice Debug and GSTN Debug tabs."""
        if self.itc_result_df is None or self.itc_result_df.empty:
            messagebox.showwarning("Warning", "Run reconciliation first.")
            return

        dlg = ctk.CTkToplevel(self)
        dlg.title("Debug Tools")
        dlg.geometry("620x320")
        dlg.lift(); dlg.focus_force()
        dlg.after(100, lambda: dlg.attributes('-topmost', False))
        dlg.attributes('-topmost', True)

        hdr = ctk.CTkFrame(dlg, fg_color="#E65100", corner_radius=0)
        hdr.pack(fill="x")
        ctk.CTkLabel(hdr, text="Debug Tools", font=ctk.CTkFont(size=14, weight="bold"),
                     text_color="white").pack(side="left", padx=16, pady=10)
        co = "  |  " + self._company_name if self._company_name else ""
        ctk.CTkLabel(hdr, text=co, font=ctk.CTkFont(size=11),
                     text_color="#FFCCBC").pack(side="left")

        body = ctk.CTkFrame(dlg, fg_color="#F5F5F5")
        body.pack(fill="both", expand=True, padx=12, pady=12)

        # Invoice Debug card
        inv_card = ctk.CTkFrame(body, fg_color="white", corner_radius=8,
                                border_width=1, border_color="#E0E0E0")
        inv_card.pack(fill="x", padx=4, pady=(0, 8))
        ctk.CTkLabel(inv_card, text="Invoice Number Debug",
                     font=ctk.CTkFont(size=13, weight="bold"),
                     text_color="#E65100").pack(anchor="w", padx=14, pady=(10, 2))
        ctk.CTkLabel(inv_card,
                     text="Find Books vs 2B pairs where GSTIN and tax amounts match but invoice numbers "
                          "differ slightly (e.g. formatting, prefix differences).",
                     font=ctk.CTkFont(size=11), text_color="#555555",
                     wraplength=540, justify="left").pack(anchor="w", padx=14, pady=(0, 8))
        ctk.CTkButton(inv_card, text="Run Invoice Debug →",
                      command=lambda: [dlg.destroy(), self.debug_matching()],
                      fg_color="#E65100", hover_color="#BF360C",
                      height=32, width=200).pack(anchor="w", padx=14, pady=(0, 12))

        # GSTN Debug card
        gstn_card = ctk.CTkFrame(body, fg_color="white", corner_radius=8,
                                 border_width=1, border_color="#E0E0E0")
        gstn_card.pack(fill="x", padx=4)
        ctk.CTkLabel(gstn_card, text="GSTIN Number Debug",
                     font=ctk.CTkFont(size=13, weight="bold"),
                     text_color="#F57F17").pack(anchor="w", padx=14, pady=(10, 2))
        ctk.CTkLabel(gstn_card,
                     text="Find Books vs 2B pairs where the invoice number and tax amounts match but the "
                          "GSTIN is different or slightly incorrect.",
                     font=ctk.CTkFont(size=11), text_color="#555555",
                     wraplength=540, justify="left").pack(anchor="w", padx=14, pady=(0, 8))
        ctk.CTkButton(gstn_card, text="Run GSTN Debug →",
                      command=lambda: [dlg.destroy(), self.gstn_debug_matching()],
                      fg_color="#F57F17", hover_color="#E65100",
                      height=32, width=200).pack(anchor="w", padx=14, pady=(0, 12))

    def _open_bulk_debug_panel(self, parent_win, candidates, do_match_fn, title_suffix=''):
        """Bulk-select panel: treeview of all candidates with checkboxes + Match Selected."""
        panel = ctk.CTkToplevel(parent_win)
        panel.title(f"Bulk Match — {title_suffix}  ({len(candidates)} candidates)")
        panel.geometry("1260x650")
        panel.lift(); panel.focus_force()
        panel.after(100, lambda: panel.attributes('-topmost', False))
        panel.attributes('-topmost', True)

        # Header
        hdr = ctk.CTkFrame(panel, fg_color=THEME_DARK, corner_radius=0)
        hdr.pack(fill="x")
        ctk.CTkLabel(hdr, text=f"All Candidates — {len(candidates)} pairs  ·  click a row to check / uncheck",
                     font=ctk.CTkFont(size=13, weight="bold"),
                     text_color="white").pack(side="left", padx=14, pady=8)

        # Use plain tk.Frame — CTkFrame's internal canvas intercepts clicks on ttk widgets
        tv_outer = tk.Frame(panel, bg="#FFFFFF")
        tv_outer.pack(fill="both", expand=True, padx=8, pady=(6, 0))

        _st = ttk.Style()
        _st.configure("Bulk.Treeview", rowheight=34, font=('Helvetica', 12))
        _st.configure("Bulk.Treeview.Heading", font=('Helvetica', 12, 'bold'), padding=(4, 6))
        _st.map("Bulk.Treeview", background=[('selected', '#E3F2FD'), ('!selected', '')])

        vsb = ttk.Scrollbar(tv_outer, orient="vertical")
        hsb = ttk.Scrollbar(tv_outer, orient="horizontal")
        vsb.pack(side="right", fill="y")
        hsb.pack(side="bottom", fill="x")

        _cols = ('chk', 'books_inv', 'books_gstin', 'books_date',
                 'twob_inv', 'twob_gstin', 'sim', 'cgst', 'sgst', 'igst')
        _hdrs = ('✓', 'Books Invoice', 'Books GSTIN', 'Books Date',
                 '2B Invoice', '2B GSTIN', 'Similarity', 'CGST', 'SGST', 'IGST')
        _wids = (52, 185, 170, 100, 185, 170, 88, 90, 90, 90)

        tv = ttk.Treeview(tv_outer, columns=_cols, show='headings',
                          style="Bulk.Treeview", selectmode='none',
                          yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.configure(command=tv.yview)
        hsb.configure(command=tv.xview)
        for col, hdr_txt, w in zip(_cols, _hdrs, _wids):
            anc = 'center' if col in ('chk', 'sim') else 'w'
            tv.heading(col, text=hdr_txt, anchor=anc)
            tv.column(col, width=w, anchor=anc, stretch=False)
        tv.tag_configure('checked',   background='#C8E6C9', foreground='#1B5E20')
        tv.tag_configure('unchecked', background='#F9F9F9', foreground='#212121')

        _iid_to_ci = {}
        _selected  = set()

        for ci, p in enumerate(candidates):
            sim = p.get('similarity', p.get('gstin_similarity', 0))
            iid = tv.insert('', 'end', tags=('unchecked',), values=(
                '☐',
                p.get('itc_invoice', ''),
                p.get('itc_gstin',   ''),
                p.get('itc_date',    ''),
                p.get('twoa_invoice',''),
                p.get('twoa_gstin',  ''),
                f"{sim:.0%}",
                f"{p.get('itc_cgst', 0):,.0f}",
                f"{p.get('itc_sgst', 0):,.0f}",
                f"{p.get('itc_igst', 0):,.0f}",
            ))
            _iid_to_ci[iid] = ci

        tv.pack(fill="both", expand=True)

        # Bottom bar
        bot = ctk.CTkFrame(panel, fg_color="#F5F5F5", corner_radius=0)
        bot.pack(fill="x", padx=0, pady=0)

        sel_lbl = ctk.CTkLabel(bot, text=f"0 of {len(candidates)} selected",
                               font=ctk.CTkFont(size=14, weight="bold"), text_color="#333333")
        sel_lbl.pack(side="left", padx=16, pady=12)

        def _refresh_lbl():
            sel_lbl.configure(text=f"{len(_selected)} of {len(candidates)} selected")

        def _toggle(event):
            iid = tv.identify_row(event.y)
            if not iid:
                return
            ci = _iid_to_ci[iid]
            if ci in _selected:
                _selected.discard(ci)
                tv.set(iid, 'chk', '☐')
                tv.item(iid, tags=('unchecked',))
            else:
                _selected.add(ci)
                tv.set(iid, 'chk', '☑')
                tv.item(iid, tags=('checked',))
            _refresh_lbl()

        tv.bind('<ButtonRelease-1>', _toggle)

        def _select_all():
            for iid, ci in _iid_to_ci.items():
                _selected.add(ci)
                tv.set(iid, 'chk', '☑')
                tv.item(iid, tags=('checked',))
            _refresh_lbl()

        def _deselect_all():
            _selected.clear()
            for iid in _iid_to_ci:
                tv.set(iid, 'chk', '☐')
                tv.item(iid, tags=('unchecked',))
            _refresh_lbl()

        def _match_selected():
            if not _selected:
                messagebox.showwarning("Nothing Selected",
                                       "Check at least one row first.", parent=panel)
                return
            n = len(_selected)
            if not messagebox.askyesno("Confirm",
                    f"Match {n} selected pair(s)?\n\nThis will update the Books Results.",
                    parent=panel):
                return
            for ci in sorted(_selected):
                do_match_fn(candidates[ci])
            messagebox.showinfo("Bulk Match Complete",
                                f"Successfully matched {n} pair(s).\n\n"
                                "The Books Results have been updated.",
                                parent=panel)
            panel.destroy()

        ctk.CTkButton(bot, text="Select All", command=_select_all,
                      fg_color="#1565C0", hover_color="#0D47A1",
                      font=ctk.CTkFont(size=13, weight="bold"),
                      height=42, width=140).pack(side="left", padx=8, pady=10)
        ctk.CTkButton(bot, text="Deselect All", command=_deselect_all,
                      fg_color="#757575", hover_color="#616161",
                      font=ctk.CTkFont(size=13, weight="bold"),
                      height=42, width=140).pack(side="left", padx=6, pady=10)
        ctk.CTkButton(bot, text="✓  Match Selected", command=_match_selected,
                      fg_color="#2E7D32", hover_color="#1B5E20",
                      font=ctk.CTkFont(size=14, weight="bold"),
                      height=42, width=190).pack(side="right", padx=16, pady=10)

    def debug_matching(self):
        """Open debug matching window to review partial invoice matches."""
        if self.itc_result_df is None or self.itc_result_df.empty:
            messagebox.showwarning("Warning", "No Books results available! Run reconciliation first.")
            return
        if self.unmatched_2a_df is None or self.unmatched_2a_df.empty:
            messagebox.showwarning("Warning", "No unmatched 2B data available!")
            return

        candidates = self.find_debug_candidate_pairs()

        if not candidates:
            messagebox.showinfo("Debug Matching",
                                "No candidates found.\n\n"
                                "This means there are no unmatched pairs where:\n"
                                "- GSTIN matches exactly\n"
                                "- Taxable Value, IGST, CGST, SGST match within Rs 10\n"
                                "- Invoice numbers are at least 10% similar")
            return

        self._open_debug_window(candidates)

    def _open_debug_window(self, candidates):
        """Display candidate pairs one at a time for user to match or skip."""
        debug_win = ctk.CTkToplevel(self)
        debug_win.title("Debug Matching - Review Partial Invoice Matches")
        debug_win.geometry("900x500")
        debug_win.lift()
        debug_win.focus_force()
        debug_win.after(100, lambda: debug_win.attributes('-topmost', False))
        debug_win.attributes('-topmost', True)

        state = {'idx': 0, 'matched': 0, 'skipped': 0, 'total': len(candidates),
                 'history': []}

        # Find vendor columns for sibling-row updates
        vendor_gstn_col = None
        vendor_inv_col = None
        for col in self.itc_result_df.columns:
            cl = col.lower().strip()
            if 'vendor' in cl and 'gstn' in cl:
                vendor_gstn_col = col
            elif 'vendor inv' in cl or 'external doc' in cl:
                vendor_inv_col = col

        # Pre-build norm_key -> [row_indices] index to avoid O(n) scans on every click
        from collections import defaultdict
        _norm_key_index = defaultdict(list)
        if vendor_gstn_col and vendor_inv_col:
            for _ri, _row in self.itc_result_df.iterrows():
                _nk = normalize_gstin(str(_row[vendor_gstn_col])) + '|' + normalize_invoice(str(_row[vendor_inv_col]))
                _norm_key_index[_nk].append(_ri)

        # Pre-build 2A results index: norm_key -> [row_indices in gstr_2a_results_df]
        _2a_result_index = defaultdict(list)
        if self.gstr_2a_results_df is not None and not self.gstr_2a_results_df.empty:
            for _ri, _row in self.gstr_2a_results_df.iterrows():
                _nk = normalize_gstin(str(_row.get('GSTN', ''))) + '|' + normalize_invoice(str(_row.get('Document_number', '')))
                _2a_result_index[_nk].append(_ri)

        # Header
        header = ctk.CTkFrame(debug_win, fg_color="#FFF3E0")
        header.pack(fill="x", padx=10, pady=(10, 5))

        progress_lbl = ctk.CTkLabel(header, text=f"Pair 1 of {state['total']}",
                                    font=ctk.CTkFont(size=14, weight="bold"))
        progress_lbl.pack(pady=(5, 0))

        sim_lbl = ctk.CTkLabel(header, text="", font=ctk.CTkFont(size=12), text_color="#E65100")
        sim_lbl.pack(pady=(0, 5))

        # Field labels
        fields = ['Invoice No', 'Invoice Date', 'GSTIN', 'Taxable Value', 'IGST', 'CGST', 'SGST']

        # ITC panel
        itc_frame = ctk.CTkFrame(debug_win)
        itc_frame.pack(fill="x", padx=10, pady=5)
        ctk.CTkLabel(itc_frame, text="Books Row", font=ctk.CTkFont(size=13, weight="bold"),
                     text_color="#E91E63").pack(anchor="w", padx=10, pady=(5, 0))
        itc_grid = ctk.CTkFrame(itc_frame, fg_color="transparent")
        itc_grid.pack(fill="x", padx=10, pady=5)
        itc_vals = {}
        for i, f in enumerate(fields):
            ctk.CTkLabel(itc_grid, text=f"{f}:", font=ctk.CTkFont(size=11, weight="bold")).grid(
                row=0, column=i, padx=8, pady=2, sticky="w")
            lbl = ctk.CTkLabel(itc_grid, text="", font=ctk.CTkFont(size=11))
            lbl.grid(row=1, column=i, padx=8, pady=2, sticky="w")
            itc_vals[f] = lbl

        # 2A panel
        twoa_frame = ctk.CTkFrame(debug_win)
        twoa_frame.pack(fill="x", padx=10, pady=5)
        ctk.CTkLabel(twoa_frame, text="2B Row", font=ctk.CTkFont(size=13, weight="bold"),
                     text_color="#1565C0").pack(anchor="w", padx=10, pady=(5, 0))
        twoa_grid = ctk.CTkFrame(twoa_frame, fg_color="transparent")
        twoa_grid.pack(fill="x", padx=10, pady=5)
        twoa_vals = {}
        for i, f in enumerate(fields):
            ctk.CTkLabel(twoa_grid, text=f"{f}:", font=ctk.CTkFont(size=11, weight="bold")).grid(
                row=0, column=i, padx=8, pady=2, sticky="w")
            lbl = ctk.CTkLabel(twoa_grid, text="", font=ctk.CTkFont(size=11))
            lbl.grid(row=1, column=i, padx=8, pady=2, sticky="w")
            twoa_vals[f] = lbl

        # Remarks
        remarks_frame = ctk.CTkFrame(debug_win, fg_color="transparent")
        remarks_frame.pack(fill="x", padx=10, pady=(5, 0))

        ctk.CTkLabel(remarks_frame, text="Remarks:",
                     font=ctk.CTkFont(size=12, weight="bold")).pack(side="left", padx=(10, 5))

        remarks_entry = ctk.CTkEntry(remarks_frame, placeholder_text="Enter remarks for this invoice...",
                                     font=ctk.CTkFont(size=12), height=32)
        remarks_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))

        # Buttons
        btn_frame = ctk.CTkFrame(debug_win, fg_color="transparent")
        btn_frame.pack(fill="x", padx=10, pady=10)

        back_btn = ctk.CTkButton(btn_frame, text="Back", fg_color="#1565C0",
                                 hover_color="#0D47A1", font=ctk.CTkFont(size=14, weight="bold"),
                                 width=120, height=40, command=lambda: on_back(), state="disabled")
        back_btn.pack(side="left", padx=10)

        match_btn = ctk.CTkButton(btn_frame, text="Match", fg_color="#2E7D32",
                                  hover_color="#1B5E20", font=ctk.CTkFont(size=14, weight="bold"),
                                  width=200, height=40, command=lambda: on_match())
        match_btn.pack(side="left", padx=20)

        skip_btn = ctk.CTkButton(btn_frame, text="Skip", fg_color="#757575",
                                 hover_color="#616161", font=ctk.CTkFont(size=14, weight="bold"),
                                 width=200, height=40, command=lambda: on_skip())
        skip_btn.pack(side="left", padx=20)

        ctk.CTkButton(btn_frame, text="☰ View All / Bulk Match",
                      fg_color="#5C6BC0", hover_color="#3949AB",
                      font=ctk.CTkFont(size=12), height=40, width=190,
                      command=lambda: self._open_bulk_debug_panel(
                          debug_win, candidates, _do_match_data, 'Invoice Debug')
                      ).pack(side="right", padx=10)

        counter_lbl = ctk.CTkLabel(debug_win, text="Matched: 0 | Skipped: 0",
                                   font=ctk.CTkFont(size=11))
        counter_lbl.pack(pady=5)

        def _save_remarks(p, action):
            """Append debug remark to any existing auto-remark for all sibling ITC rows.
            Always writes an auto-remark; user note (if typed) is appended after it."""
            remark_text = remarks_entry.get().strip()
            if action == 'Matched':
                auto = f"Debug matched with {p.get('twoa_invoice', '')}"
            else:
                auto = "Debug skipped"
            new_part = f"[{action}] {auto}" + (f" | {remark_text}" if remark_text else "")
            norm_key = normalize_gstin(p['itc_gstin']) + '|' + p['itc_inv_norm']
            if 'Remarks' not in self.itc_result_df.columns:
                self.itc_result_df['Remarks'] = ''
            if vendor_gstn_col and vendor_inv_col:
                for row_idx in _norm_key_index.get(norm_key, []):
                    existing = str(self.itc_result_df.at[row_idx, 'Remarks']).strip()
                    self.itc_result_df.at[row_idx, 'Remarks'] = (existing + '; ' + new_part) if existing else new_part
            else:
                existing = str(self.itc_result_df.at[p['itc_index'], 'Remarks']).strip()
                self.itc_result_df.at[p['itc_index'], 'Remarks'] = (existing + '; ' + new_part) if existing else new_part

        def display_pair(i):
            if i >= state['total']:
                show_summary()
                return
            # Re-enable buttons in case we came back from summary
            match_btn.configure(state="normal")
            skip_btn.configure(state="normal")
            remarks_entry.configure(state="normal")

            p = candidates[i]
            progress_lbl.configure(text=f"Pair {i + 1} of {state['total']}")
            sim_lbl.configure(text=f"Invoice Similarity: {p['similarity']:.0%}")

            itc_vals['Invoice No'].configure(text=p['itc_invoice'])
            itc_vals['Invoice Date'].configure(text=p.get('itc_date', ''))
            itc_vals['GSTIN'].configure(text=p['itc_gstin'])
            itc_vals['Taxable Value'].configure(text=f"{p['itc_taxable']:,.2f}")
            itc_vals['IGST'].configure(text=f"{p['itc_igst']:,.2f}")
            itc_vals['CGST'].configure(text=f"{p['itc_cgst']:,.2f}")
            itc_vals['SGST'].configure(text=f"{p['itc_sgst']:,.2f}")

            twoa_vals['Invoice No'].configure(text=p['twoa_invoice'])
            twoa_vals['Invoice Date'].configure(text=p.get('twoa_date', ''))
            twoa_vals['GSTIN'].configure(text=p['twoa_gstin'])
            twoa_vals['Taxable Value'].configure(text=f"{p['twoa_taxable']:,.2f}")
            twoa_vals['IGST'].configure(text=f"{p['twoa_igst']:,.2f}")
            twoa_vals['CGST'].configure(text=f"{p['twoa_cgst']:,.2f}")
            twoa_vals['SGST'].configure(text=f"{p['twoa_sgst']:,.2f}")

            # Highlight fields: orange=partial-match, green=matching, red=mismatch
            def _clr(lbl, bg, fg):
                lbl.configure(fg_color=bg, text_color=fg)
            _G_BG, _G_FG = "#E8F5E9", "#1B5E20"
            _R_BG, _R_FG = "#FFEBEE", "#B71C1C"
            _O_BG, _O_FG = "#FFF3E0", "#E65100"
            # Invoice No: partial match by definition in invoice debug → orange
            _clr(itc_vals['Invoice No'],  _O_BG, _O_FG)
            _clr(twoa_vals['Invoice No'], _O_BG, _O_FG)
            # GSTIN: should match in invoice debug
            _gstin_ok = normalize_gstin(p['itc_gstin']) == normalize_gstin(p['twoa_gstin'])
            _clr(itc_vals['GSTIN'],  _G_BG if _gstin_ok else _R_BG, _G_FG if _gstin_ok else _R_FG)
            _clr(twoa_vals['GSTIN'], _G_BG if _gstin_ok else _R_BG, _G_FG if _gstin_ok else _R_FG)
            # Tax amounts and taxable value
            for _f, _iv, _tv in [('IGST', p['itc_igst'], p['twoa_igst']),
                                  ('CGST', p['itc_cgst'], p['twoa_cgst']),
                                  ('SGST', p['itc_sgst'], p['twoa_sgst']),
                                  ('Taxable Value', p['itc_taxable'], p['twoa_taxable'])]:
                try:
                    _ok = abs(float(_iv) - float(_tv)) < 1
                except (TypeError, ValueError):
                    _ok = str(_iv).strip() == str(_tv).strip()
                _clr(itc_vals[_f],  _G_BG if _ok else _R_BG, _G_FG if _ok else _R_FG)
                _clr(twoa_vals[_f], _G_BG if _ok else _R_BG, _G_FG if _ok else _R_FG)
            # Invoice Date
            _d_ok = p.get('itc_date', '').strip() == p.get('twoa_date', '').strip()
            _clr(itc_vals['Invoice Date'],  _G_BG if _d_ok else _R_BG, _G_FG if _d_ok else _R_FG)
            _clr(twoa_vals['Invoice Date'], _G_BG if _d_ok else _R_BG, _G_FG if _d_ok else _R_FG)

            # Clear remarks for new pair
            remarks_entry.delete(0, 'end')

            counter_lbl.configure(text=f"Matched: {state['matched']} | Skipped: {state['skipped']}")
            # Enable/disable back button
            back_btn.configure(state="normal" if state['history'] else "disabled")

        def on_match():
            p = candidates[state['idx']]
            norm_key = normalize_gstin(p['itc_gstin']) + '|' + p['itc_inv_norm']
            df = self.itc_result_df
            target_rows = _norm_key_index.get(norm_key, [p['itc_index']]) if vendor_gstn_col and vendor_inv_col else [p['itc_index']]
            # Columns to save/restore for undo
            tracked_cols = ['Status', 'Remarks', '2B Invoice No', '2B GSTIN',
                            'CGST as per 2B', 'SGST as per 2B', 'IGST as per 2B',
                            'Type', 'Booking Month as per GSTR-2B']
            prev_vals = {col: {} for col in tracked_cols}
            for row_idx in target_rows:
                for col in tracked_cols:
                    prev_vals[col][row_idx] = df.at[row_idx, col] if col in df.columns else ''

            # Save previous 2A result statuses for undo
            twoa_norm_key = normalize_gstin(p['twoa_gstin']) + '|' + normalize_invoice(p['twoa_invoice'])
            prev_2a_statuses = {}
            if self.gstr_2a_results_df is not None and 'Status' in self.gstr_2a_results_df.columns:
                for _ri in _2a_result_index.get(twoa_norm_key, []):
                    prev_2a_statuses[_ri] = self.gstr_2a_results_df.at[_ri, 'Status']

            state['history'].append({'action': 'match', 'idx': state['idx'], 'prev_vals': prev_vals, 'prev_2a_statuses': prev_2a_statuses})

            _save_remarks(p, 'Matched')
            # Ensure all 2A columns exist
            defaults = {'2B Invoice No': 'Not Found', '2B GSTIN': 'Not Found',
                        'CGST as per 2B': 'Not found in 2B', 'SGST as per 2B': 'Not found in 2B',
                        'IGST as per 2B': 'Not found in 2B', 'Type': 'Not found in 2B',
                        'Booking Month as per GSTR-2B': 'Not found in 2B'}
            for col, default in defaults.items():
                if col not in df.columns:
                    df[col] = default
            for row_idx in target_rows:
                df.at[row_idx, 'Status'] = 'Matched but invoice number is not accurate'
                df.at[row_idx, '2B Invoice No'] = p['twoa_invoice']
                df.at[row_idx, '2B GSTIN'] = p['twoa_gstin']
                df.at[row_idx, 'CGST as per 2B'] = p['twoa_cgst']
                df.at[row_idx, 'SGST as per 2B'] = p['twoa_sgst']
                df.at[row_idx, 'IGST as per 2B'] = p['twoa_igst']
                df.at[row_idx, 'Type'] = p.get('twoa_type', '')
                df.at[row_idx, 'Booking Month as per GSTR-2B'] = p.get('twoa_booking_month', '')

            # Update matching 2A result rows to 'Matched'
            if self.gstr_2a_results_df is not None and 'Status' in self.gstr_2a_results_df.columns:
                for _ri in _2a_result_index.get(twoa_norm_key, []):
                    self.gstr_2a_results_df.at[_ri, 'Status'] = 'Matched'

            state['matched'] += 1
            self._debug_final_matches.append({'cand_idx': state['idx'], **p})
            state['idx'] += 1
            self.log(f"Debug Match: '{p['itc_invoice']}' -> '{p['twoa_invoice']}' (similarity: {p['similarity']:.0%})")
            display_pair(state['idx'])

        def on_skip():
            p = candidates[state['idx']]
            # Save previous remarks for undo
            norm_key = normalize_gstin(p['itc_gstin']) + '|' + p['itc_inv_norm']
            prev_remarks = {}
            has_remarks = 'Remarks' in self.itc_result_df.columns
            if vendor_gstn_col and vendor_inv_col:
                for row_idx in _norm_key_index.get(norm_key, []):
                    prev_remarks[row_idx] = self.itc_result_df.at[row_idx, 'Remarks'] if has_remarks else ''
            else:
                prev_remarks[p['itc_index']] = self.itc_result_df.at[p['itc_index'], 'Remarks'] if has_remarks else ''

            state['history'].append({'action': 'skip', 'idx': state['idx'], 'prev_remarks': prev_remarks})

            # Save remarks before moving on
            _save_remarks(p, 'Skipped')
            state['skipped'] += 1
            state['idx'] += 1
            display_pair(state['idx'])

        def on_back():
            if not state['history']:
                return
            entry = state['history'].pop()
            df = self.itc_result_df
            if entry['action'] == 'match':
                state['matched'] -= 1
                reverted = entry['idx']
                self._debug_final_matches = [m for m in self._debug_final_matches if m['cand_idx'] != reverted]
                if 'prev_vals' in entry:
                    for col, row_map in entry['prev_vals'].items():
                        if col in df.columns:
                            for row_idx, old_val in row_map.items():
                                df.at[row_idx, col] = old_val
                else:
                    # legacy undo entries
                    for row_idx, old_status in entry.get('prev_statuses', {}).items():
                        df.at[row_idx, 'Status'] = old_status
                    if '2B Invoice No' in df.columns:
                        for row_idx, old_inv in entry.get('prev_2a_invoices', {}).items():
                            df.at[row_idx, '2B Invoice No'] = old_inv
                    if 'Remarks' in df.columns:
                        for row_idx, old_remark in entry.get('prev_remarks', {}).items():
                            df.at[row_idx, 'Remarks'] = old_remark
                # Revert 2A result statuses
                if self.gstr_2a_results_df is not None and 'Status' in self.gstr_2a_results_df.columns:
                    for _ri, old_status in entry.get('prev_2a_statuses', {}).items():
                        self.gstr_2a_results_df.at[_ri, 'Status'] = old_status
            elif entry['action'] == 'skip':
                state['skipped'] -= 1
                if 'Remarks' in df.columns:
                    for row_idx, old_remark in entry.get('prev_remarks', {}).items():
                        df.at[row_idx, 'Remarks'] = old_remark
            # Go back to that pair
            state['idx'] = entry['idx']
            self.log(f"Debug Back: Reverted pair {entry['idx'] + 1} ({entry['action']})")
            display_pair(state['idx'])

        def show_summary():
            match_btn.configure(state="disabled")
            skip_btn.configure(state="disabled")
            remarks_entry.configure(state="disabled")
            back_btn.configure(state="normal" if state['history'] else "disabled")
            progress_lbl.configure(text="Review Complete!")
            sim_lbl.configure(text="")
            for lbl in itc_vals.values():
                lbl.configure(text="")
            for lbl in twoa_vals.values():
                lbl.configure(text="")
            counter_lbl.configure(text=f"DONE - Matched: {state['matched']} | Skipped: {state['skipped']}")
            self.log(f"Debug Matching complete: {state['matched']} matched, {state['skipped']} skipped out of {state['total']}")
            matched_note = ("\n\nMatched pairs will be saved when you click\n"
                            "'Save to YTD Database' on the dashboard."
                            if self._debug_final_matches else "")
            messagebox.showinfo("Debug Matching Complete",
                                f"Total pairs reviewed: {state['total']}\n"
                                f"Matched by user: {state['matched']}\n"
                                f"Skipped: {state['skipped']}"
                                + matched_note,
                                parent=debug_win)

        def _do_match_data(p, remark=''):
            """Apply Invoice Debug match without touching step-by-step state (used by bulk panel)."""
            norm_key = normalize_gstin(p['itc_gstin']) + '|' + p['itc_inv_norm']
            df = self.itc_result_df
            target_rows = (_norm_key_index.get(norm_key, [p['itc_index']])
                           if vendor_gstn_col and vendor_inv_col else [p['itc_index']])
            defaults = {'2B Invoice No': 'Not Found', '2B GSTIN': 'Not Found',
                        'CGST as per 2B': 'Not found in 2B', 'SGST as per 2B': 'Not found in 2B',
                        'IGST as per 2B': 'Not found in 2B', 'Type': 'Not found in 2B',
                        'Booking Month as per GSTR-2B': 'Not found in 2B'}
            for col, dflt in defaults.items():
                if col not in df.columns:
                    df[col] = dflt
            if 'Remarks' not in df.columns:
                df['Remarks'] = ''
            note = (f"[Bulk Match] Debug matched with {p.get('twoa_invoice', '')}"
                    + (f" | {remark}" if remark else ""))
            for row_idx in target_rows:
                df.at[row_idx, 'Status'] = 'Matched but invoice number is not accurate'
                df.at[row_idx, '2B Invoice No'] = p['twoa_invoice']
                df.at[row_idx, '2B GSTIN'] = p['twoa_gstin']
                df.at[row_idx, 'CGST as per 2B'] = p['twoa_cgst']
                df.at[row_idx, 'SGST as per 2B'] = p['twoa_sgst']
                df.at[row_idx, 'IGST as per 2B'] = p['twoa_igst']
                df.at[row_idx, 'Type'] = p.get('twoa_type', '')
                df.at[row_idx, 'Booking Month as per GSTR-2B'] = p.get('twoa_booking_month', '')
                existing = str(df.at[row_idx, 'Remarks']).strip()
                df.at[row_idx, 'Remarks'] = (existing + '; ' + note) if existing else note
            twoa_nk = normalize_gstin(p['twoa_gstin']) + '|' + normalize_invoice(p['twoa_invoice'])
            if self.gstr_2a_results_df is not None and 'Status' in self.gstr_2a_results_df.columns:
                for _ri in _2a_result_index.get(twoa_nk, []):
                    self.gstr_2a_results_df.at[_ri, 'Status'] = 'Matched'
            self._debug_final_matches.append(p)
            self.log(f"Bulk Invoice Debug: '{p['itc_invoice']}' → '{p['twoa_invoice']}'")

        display_pair(0)

    # ── Summary Report ────────────────────────────────────────────────────────
    # ── Helpers for summary report ────────────────────────────────────────────
    @staticmethod
    def _extract_month_col(df):
        """Return the name of the best available month/date column in df, or None."""
        for candidate in ('2B MONTH', '3B MONTH', 'Booking Date', 'Invoice Date',
                          'GSTR-1/5 Filling Period', 'Period', 'Month', 'Date'):
            if candidate in df.columns:
                return candidate
        for col in df.columns:
            if any(k in col.lower() for k in ('month', 'period', 'date')):
                return col
        return None

    @staticmethod
    def _month_label(raw):
        """Best-effort short month label (e.g. 'Jan 2023') from any date/period cell."""
        if pd.isna(raw) or str(raw).strip() in ('', 'nan'):
            return 'Unknown'
        s = str(raw).strip()
        from datetime import datetime as _dt
        for fmt in ('%b-%Y', '%B-%Y', '%Y-%m', '%m/%Y', '%Y/%m', '%m-%Y',
                    '%m/%d/%Y', '%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d',
                    '%d-%b-%Y', '%b %Y', '%B %Y',
                    '%m/%d/%y', '%d/%m/%y', '%d-%m-%y'):
            try:
                return _dt.strptime(s, fmt).strftime('%b %Y')
            except ValueError:
                pass
        try:
            return pd.to_datetime(s, dayfirst=False).strftime('%b %Y')
        except Exception:
            pass
        if len(s) >= 4 and s[:3].isalpha():
            return s[:7]
        return s[:10]

    def show_summary_report(self):
        """Show a rich multi-tab summary report with KPI cards and graphical analysis."""
        if self.itc_result_df is None or self.itc_result_df.empty:
            messagebox.showwarning("No Data", "Run a reconciliation first.")
            return

        # ── Matplotlib imports ────────────────────────────────────────────
        _has_mpl = False
        try:
            import matplotlib
            matplotlib.use('Agg')
            import matplotlib.pyplot as plt
            import matplotlib.ticker as mticker
            import matplotlib.patches as mpatches
            from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
            _has_mpl = True
        except Exception:
            pass

        df = self._coerce_numeric_cols(self.itc_result_df.copy())
        df2b = self.gstr_2a_results_df
        total = len(df)

        STATUS_COLORS = {
            'Matched':                                    '#4CAF50',
            'Higher in 2B':                              '#FFC107',
            'Lower in 2B':                               '#2196F3',
            'Not found in 2B':                           '#9C27B0',
            'Matched but invoice number is not accurate':'#FF9800',
            'Unmatched':                                 '#F44336',
        }
        MATCHED_STATUSES = {'Matched', 'Matched but invoice number is not accurate'}

        # ── Column resolution ─────────────────────────────────────────────
        _cgst_col, _sgst_col, _igst_col = find_tax_amount_columns(df)

        def _tax_sum(sub):
            c = pd.to_numeric(sub[_cgst_col], errors='coerce').fillna(0).sum() if _cgst_col and _cgst_col in sub.columns else 0
            s = pd.to_numeric(sub[_sgst_col], errors='coerce').fillna(0).sum() if _sgst_col and _sgst_col in sub.columns else 0
            i = pd.to_numeric(sub[_igst_col], errors='coerce').fillna(0).sum() if _igst_col and _igst_col in sub.columns else 0
            return c, s, i

        def _total_tax(sub):
            return sum(_tax_sum(sub))

        def _fmt_amt(v):
            if v >= 1e7:  return f'₹{v/1e7:.2f}Cr'
            if v >= 1e5:  return f'₹{v/1e5:.1f}L'
            return f'₹{v:,.0f}'

        # ── Pre-compute KPIs ──────────────────────────────────────────────
        status_counts  = df['Status'].value_counts()
        matched_df     = df[df['Status'].isin(MATCHED_STATUSES)]
        unmatched_df   = df[~df['Status'].isin(MATCHED_STATUSES)]
        itc_total      = _total_tax(df)
        itc_matched    = _total_tax(matched_df)
        itc_unmatch    = _total_tax(unmatched_df)
        pct_cnt        = matched_df.shape[0] / total * 100 if total else 0
        pct_amt        = itc_matched / itc_total * 100 if itc_total else 0

        # 2B total
        b2b_total = 0.0
        if df2b is not None and not df2b.empty:
            _2b_c, _2b_s, _2b_i = find_tax_amount_columns(df2b)
            for _col in [_2b_c, _2b_s, _2b_i]:
                if _col and _col in df2b.columns:
                    b2b_total += pd.to_numeric(df2b[_col], errors='coerce').fillna(0).sum()

        # ── Window ────────────────────────────────────────────────────────
        win = ctk.CTkToplevel(self)
        win.title("Reconciliation Summary Report")
        win.geometry("1300x820")
        win.lift(); win.focus_force()
        win.after(100, lambda: win.attributes('-topmost', False))
        win.attributes('-topmost', True)

        # ── Header ────────────────────────────────────────────────────────
        hdr = ctk.CTkFrame(win, fg_color=THEME_DARK, corner_radius=0)
        hdr.pack(fill="x")
        ctk.CTkLabel(hdr, text="Reconciliation Summary Report",
                     font=ctk.CTkFont(size=16, weight="bold"),
                     text_color="white").pack(side="left", padx=18, pady=10)
        co_parts = [p for p in [self._company_name, self._company_gst,
                                 getattr(self, '_company_period', '')] if p]
        if co_parts:
            ctk.CTkLabel(hdr, text="  |  " + "  ·  ".join(co_parts),
                         font=ctk.CTkFont(size=11), text_color="#F8BBD0").pack(side="left")

        # ── KPI Strip ─────────────────────────────────────────────────────
        kpi_bar = ctk.CTkFrame(win, fg_color="#ECEFF1", corner_radius=0)
        kpi_bar.pack(fill="x")
        for lbl, val, fg, bg in [
            ("Total Invoices",  f"{total:,}",                                    "#1565C0", "#E3F2FD"),
            ("Matched",         f"{matched_df.shape[0]:,}  ({pct_cnt:.1f}%)",   "#2E7D32", "#E8F5E9"),
            ("Not Matched",     f"{unmatched_df.shape[0]:,}",                   "#B71C1C", "#FFEBEE"),
            ("Total Books Amt",  _fmt_amt(itc_total),                             "#4A148C", "#F3E5F5"),
            ("Matched Amt",     f"{_fmt_amt(itc_matched)}  ({pct_amt:.1f}%)",   "#E65100", "#FFF3E0"),
            ("Unmatched Amt",   _fmt_amt(itc_unmatch),                           "#880E4F", "#FCE4EC"),
            ("2B Total Amt",    _fmt_amt(b2b_total) if b2b_total else "—",      "#37474F", "#ECEFF1"),
        ]:
            card = ctk.CTkFrame(kpi_bar, fg_color=bg, corner_radius=8,
                                border_width=1, border_color=fg)
            card.pack(side="left", padx=5, pady=6, expand=True, fill="x")
            ctk.CTkLabel(card, text=val, font=ctk.CTkFont(size=12, weight="bold"),
                         text_color=fg).pack(pady=(7, 0), padx=8)
            ctk.CTkLabel(card, text=lbl, font=ctk.CTkFont(size=9),
                         text_color="#666").pack(pady=(0, 7), padx=8)

        # ── Shared helpers ────────────────────────────────────────────────
        def _embed_fig(fig, parent):
            try:
                fig_h = int(fig.get_figheight() * fig.dpi)
                cv = FigureCanvasTkAgg(fig, master=parent)
                cv.draw()
                w = cv.get_tk_widget()
                w.configure(height=fig_h)
                w.pack(fill="x", padx=4, pady=(4, 2))
            except Exception as _e:
                ctk.CTkLabel(parent, text=f"Chart error: {_e}",
                             fg_color="#FFF9C4", text_color="#B71C1C",
                             font=ctk.CTkFont(size=10)).pack(fill="x", padx=4, pady=4)

        def _make_table(parent, rows, col_names, widths, height=8):
            style = ttk.Style()
            style.configure("SR4.Treeview", rowheight=22, font=('Helvetica', 9))
            style.configure("SR4.Treeview.Heading", font=('Helvetica', 9, 'bold'))
            outer = ctk.CTkFrame(parent, fg_color="#F8F8F8", corner_radius=6)
            outer.pack(fill="x", padx=6, pady=(2, 6))
            vsb = ttk.Scrollbar(outer, orient="vertical")
            hsb = ttk.Scrollbar(outer, orient="horizontal")
            tree = ttk.Treeview(outer, columns=col_names, show='headings',
                                style="SR4.Treeview",
                                height=min(len(rows), height),
                                yscrollcommand=vsb.set, xscrollcommand=hsb.set)
            vsb.configure(command=tree.yview)
            hsb.configure(command=tree.xview)
            for cn, w in zip(col_names, widths):
                tree.heading(cn, text=cn, anchor='w')
                tree.column(cn, width=w, minwidth=w, anchor='w')
            for r in rows:
                tree.insert('', 'end', values=r)
            vsb.pack(side="right", fill="y")
            hsb.pack(side="bottom", fill="x")
            tree.pack(fill="x", expand=False)
            return outer

        def _ax_clean(ax):
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)
            ax.set_facecolor('#FAFAFA')

        def _rupee_fmt(ax, axis='y'):
            f = mticker.FuncFormatter(
                lambda x, _: f'₹{x/1e7:.1f}Cr' if x >= 1e7
                else (f'₹{x/1e5:.0f}L' if x >= 1e5 else f'₹{x:,.0f}'))
            (ax.yaxis if axis == 'y' else ax.xaxis).set_major_formatter(f)

        # ── Tabs ──────────────────────────────────────────────────────────
        tabs = ctk.CTkTabview(win, fg_color="#F0F2F5")
        tabs.pack(fill="both", expand=True, padx=8, pady=(4, 0))
        for t in ("Overview", "Month Wise", "2B Month Wise", "GSTN Wise", "Amount Analysis"):
            tabs.add(t)

        # ══════════════════════════════════════════════════════════════════
        # TAB 1 — OVERVIEW
        # ══════════════════════════════════════════════════════════════════
        ov = tabs.tab("Overview")

        if _has_mpl:
            try:
                fig_ov, axes_ov = plt.subplots(1, 3, figsize=(14, 3.8))
                fig_ov.patch.set_facecolor('#F8F8F8')

                # [0] Status donut
                lbls = list(status_counts.index)
                szs  = list(status_counts.values)
                cols = [STATUS_COLORS.get(l, '#90A4AE') for l in lbls]
                wedges, _, ats = axes_ov[0].pie(
                    szs, labels=None, colors=cols,
                    autopct=lambda p: f'{p:.0f}%' if p > 3 else '',
                    startangle=90, pctdistance=0.68,
                    wedgeprops=dict(width=0.48, edgecolor='white', linewidth=1.5))
                for at in ats:
                    at.set_fontsize(8); at.set_fontweight('bold')
                axes_ov[0].text(0, 0.10, f'{pct_cnt:.0f}%', ha='center', va='center',
                                fontsize=15, fontweight='bold', color='#2E7D32')
                axes_ov[0].text(0, -0.20, 'Matched', ha='center', va='center',
                                fontsize=9, color='#555')
                axes_ov[0].set_title('Status Distribution', fontsize=11,
                                     fontweight='bold', pad=4)
                axes_ov[0].legend(wedges,
                                  [f'{l} ({v})' for l, v in zip(lbls, szs)],
                                  loc='lower center', bbox_to_anchor=(0.5, -0.32),
                                  ncol=2, fontsize=7, frameon=False)

                # [1] Vertical bar: tax amount by status
                st_amts = {}
                for st in lbls:
                    c2, s2, i2 = _tax_sum(df[df['Status'] == st])
                    st_amts[st] = c2 + s2 + i2
                by_v = list(st_amts.keys())
                bx_v = [st_amts[k] for k in by_v]
                bc_v = [STATUS_COLORS.get(k, '#90A4AE') for k in by_v]
                vb = axes_ov[1].bar(range(len(by_v)), bx_v, color=bc_v,
                                    edgecolor='white', width=0.6, alpha=0.9)
                axes_ov[1].set_title('Tax Amount by Status', fontsize=11,
                                     fontweight='bold', pad=4)
                _rupee_fmt(axes_ov[1], 'y')
                _ax_clean(axes_ov[1])
                axes_ov[1].set_xticks(range(len(by_v)))
                axes_ov[1].set_xticklabels(
                    [k[:14] for k in by_v], rotation=28, ha='right', fontsize=7)
                for bar, val in zip(vb, bx_v):
                    if val > 0:
                        axes_ov[1].text(bar.get_x() + bar.get_width() / 2,
                                        bar.get_height() * 1.01, _fmt_amt(val),
                                        ha='center', va='bottom', fontsize=7,
                                        fontweight='bold')

                # [2] ITC vs 2B grouped vertical bar with per-bar value labels & legend
                tx_lbls, itc_v, b2b_v = [], [], []
                _2b_cc = _2b_ss = _2b_ii = None
                if df2b is not None and not df2b.empty:
                    _2b_cc, _2b_ss, _2b_ii = find_tax_amount_columns(df2b)
                for col, lbl, b2b_col in [
                    (_cgst_col, 'CGST', _2b_cc),
                    (_sgst_col, 'SGST', _2b_ss),
                    (_igst_col, 'IGST', _2b_ii),
                ]:
                    if col and col in df.columns:
                        tx_lbls.append(lbl)
                        itc_v.append(pd.to_numeric(df[col], errors='coerce').fillna(0).sum())
                        b2b_v.append(pd.to_numeric(df2b[b2b_col], errors='coerce').fillna(0).sum()
                                     if b2b_col and df2b is not None and b2b_col in df2b.columns else 0)
                if tx_lbls:
                    xp = range(len(tx_lbls))
                    w  = 0.34
                    ITC_COLOR = '#880E4F'   # dark pink — Books bars
                    B2B_COLOR = '#F48FB1'   # light pink — GSTR-2B bars
                    itc_bars = axes_ov[2].bar(
                        [xi - w/2 for xi in xp], itc_v, width=w,
                        color=ITC_COLOR, label='Books',
                        alpha=0.92, edgecolor='white')
                    b2b_bars = axes_ov[2].bar(
                        [xi + w/2 for xi in xp], b2b_v, width=w,
                        color=B2B_COLOR, label='GSTR-2B',
                        alpha=0.92, edgecolor='white')
                    axes_ov[2].set_xticks(list(xp))
                    axes_ov[2].set_xticklabels(tx_lbls, fontsize=9)
                    axes_ov[2].set_title('Books vs GSTR-2B', fontsize=11,
                                         fontweight='bold', pad=4)
                    _rupee_fmt(axes_ov[2])
                    _ax_clean(axes_ov[2])
                    # Value labels on each bar
                    for bar, val in zip(itc_bars, itc_v):
                        if val > 0:
                            axes_ov[2].text(bar.get_x() + bar.get_width()/2,
                                            bar.get_height() * 1.01, _fmt_amt(val),
                                            ha='center', va='bottom',
                                            fontsize=6.5, color='#880E4F',
                                            fontweight='bold')
                    for bar, val in zip(b2b_bars, b2b_v):
                        if val > 0:
                            axes_ov[2].text(bar.get_x() + bar.get_width()/2,
                                            bar.get_height() * 1.01, _fmt_amt(val),
                                            ha='center', va='bottom',
                                            fontsize=6.5, color='#880E4F',
                                            fontweight='bold')
                    axes_ov[2].legend(
                        handles=[mpatches.Patch(color=ITC_COLOR, label='Books'),
                                 mpatches.Patch(color=B2B_COLOR, label='GSTR-2B')],
                        fontsize=7, frameon=True, framealpha=0.85,
                        loc='upper right')

                fig_ov.tight_layout(pad=1.8)
                _embed_fig(fig_ov, ov)
                plt.close(fig_ov)
            except Exception:
                pass

        ov_rows = []
        for st, cnt in status_counts.items():
            c2, s2, i2 = _tax_sum(df[df['Status'] == st])
            st_tax = c2 + s2 + i2
            amt_pct = f"{st_tax/itc_total*100:.1f}%" if itc_total else "—"
            ov_rows.append((st, cnt, f"{cnt/total*100:.1f}%",
                            f"₹{c2:,.0f}", f"₹{s2:,.0f}", f"₹{i2:,.0f}",
                            f"₹{st_tax:,.0f}", amt_pct))
        _make_table(ov, ov_rows,
                    ('Status', 'Count', '% Count', 'CGST', 'SGST', 'IGST', 'Total Tax', '% Amount'),
                    (240, 60, 70, 110, 110, 110, 125, 85), height=8)

        # ══════════════════════════════════════════════════════════════════
        # TAB 2 — MONTH WISE
        # ══════════════════════════════════════════════════════════════════
        mw = tabs.tab("Month Wise")
        month_col = self._extract_month_col(df)
        if month_col:
            df['_month_lbl'] = df[month_col].apply(self._month_label)
            all_st = list(status_counts.index)
            mw_data = []
            for m in sorted(df['_month_lbl'].unique()):
                sub  = df[df['_month_lbl'] == m]
                mc   = len(sub[sub['Status'].isin(MATCHED_STATUSES)])
                c2, s2, i2 = _tax_sum(sub)
                tt   = c2 + s2 + i2
                mt   = _total_tax(sub[sub['Status'].isin(MATCHED_STATUSES)])
                pct_s = f"{mt/tt*100:.1f}%" if tt else "—"
                mw_data.append({'m': m, 'total': len(sub), 'mc': mc,
                                'uc': len(sub) - mc, 'c': c2, 's': s2, 'i': i2,
                                'tt': tt, 'mt': mt, 'pct': pct_s,
                                'st_c': {st: len(sub[sub['Status'] == st]) for st in all_st}})

            if _has_mpl and mw_data:
                try:
                    fig_mw, (ax_s, ax_l) = plt.subplots(
                        2, 1, figsize=(14, 4.2),
                        gridspec_kw={'height_ratios': [3, 1.5]})
                    fig_mw.patch.set_facecolor('#F8F8F8')
                    n      = len(mw_data)
                    xm     = range(n)
                    xlbls  = [r['m'] for r in mw_data]

                    # How many labels to show — keep ≤ 18 to avoid overlap
                    step = max(1, n // 18)
                    tick_pos  = list(range(0, n, step))
                    tick_lbls = [xlbls[i] for i in tick_pos]

                    bottoms = [0] * n
                    bar_w   = max(0.4, min(0.8, 12.0 / n))
                    for st in all_st:
                        vals = [r['st_c'].get(st, 0) for r in mw_data]
                        ax_s.bar(xm, vals, bottom=bottoms, label=st, width=bar_w,
                                 color=STATUS_COLORS.get(st, '#90A4AE'), alpha=0.9)
                        bottoms = [b + v for b, v in zip(bottoms, vals)]
                    ax_s.set_xticks(tick_pos)
                    ax_s.set_xticklabels(tick_lbls, rotation=40, ha='right', fontsize=8)
                    ax_s.set_title('Month-wise Invoice Status', fontsize=11,
                                   fontweight='bold', pad=4)
                    ax_s.set_ylabel('Invoice Count')
                    ax_s.legend(fontsize=8, loc='upper right', frameon=True,
                                framealpha=0.8, ncol=3)
                    ax_s.set_xlim(-0.5, n - 0.5)
                    _ax_clean(ax_s)

                    pct_line = [r['mt']/r['tt']*100 if r['tt'] else 0 for r in mw_data]
                    mk_size  = max(2, min(5, 60 // n))
                    ax_l.plot(list(xm), pct_line, marker='o', color='#1565C0',
                              linewidth=1.8, markersize=mk_size)
                    ax_l.fill_between(list(xm), pct_line, alpha=0.12, color='#1565C0')
                    ax_l.axhline(y=100, color='#4CAF50', linewidth=1,
                                 linestyle='--', alpha=0.6, label='100%')
                    ax_l.set_xticks(tick_pos)
                    ax_l.set_xticklabels(tick_lbls, rotation=40, ha='right', fontsize=8)
                    ax_l.set_ylim(0, 115)
                    ax_l.yaxis.set_major_formatter(
                        mticker.FuncFormatter(lambda x, _: f'{x:.0f}%'))
                    ax_l.set_title('% Amount Matched — Monthly Trend', fontsize=9)
                    ax_l.set_xlim(-0.5, n - 0.5)
                    _ax_clean(ax_l)
                    # Only annotate when ≤ 24 months — otherwise too crowded
                    if n <= 24:
                        for xi, pv in zip(xm, pct_line):
                            ax_l.annotate(f'{pv:.0f}%', (xi, pv),
                                          textcoords='offset points', xytext=(0, 5),
                                          ha='center', fontsize=7, color='#1565C0',
                                          fontweight='bold')

                    fig_mw.tight_layout(pad=1.8)
                    _embed_fig(fig_mw, mw)
                    plt.close(fig_mw)
                except Exception:
                    pass

            _make_table(mw,
                [(r['m'], r['total'], r['mc'], r['uc'],
                  f"₹{r['c']:,.0f}", f"₹{r['s']:,.0f}", f"₹{r['i']:,.0f}",
                  f"₹{r['tt']:,.0f}", r['pct']) for r in mw_data],
                ('Month', 'Total', 'Matched', 'Unmatched',
                 'CGST', 'SGST', 'IGST', 'Total Tax', '% Amt Matched'),
                (110, 65, 75, 85, 115, 115, 115, 125, 110), height=8)
        else:
            ctk.CTkLabel(mw, text="No date/month column found in Books data.",
                         font=ctk.CTkFont(size=13)).pack(pady=40)

        # ══════════════════════════════════════════════════════════════════
        # TAB 2b — 2B MONTH WISE
        # ══════════════════════════════════════════════════════════════════
        bm_tab = tabs.tab("2B Month Wise")
        _bm_col = 'Booking Month as per GSTR-2B'
        if _bm_col in df.columns:
            _bm_df = df[df[_bm_col].notna()].copy()
            _bm_df['_bm_lbl'] = _bm_df[_bm_col].apply(
                lambda v: self._month_label(v) if str(v).strip() not in ('', 'nan', 'Not found in 2B') else None)
            _bm_df = _bm_df[_bm_df['_bm_lbl'].notna()]
            if not _bm_df.empty:
                _all_st_bm = list(status_counts.index)
                _bm_data = []
                for _bm in sorted(_bm_df['_bm_lbl'].unique()):
                    _sub = _bm_df[_bm_df['_bm_lbl'] == _bm]
                    _mc   = len(_sub[_sub['Status'].isin(MATCHED_STATUSES)])
                    _c2, _s2, _i2 = _tax_sum(_sub)
                    _tt   = _c2 + _s2 + _i2
                    _mt   = _total_tax(_sub[_sub['Status'].isin(MATCHED_STATUSES)])
                    _pct  = f"{min(100.0, _mt/_tt*100):.1f}%" if _tt else "—"
                    _bm_data.append({'m': _bm, 'total': len(_sub), 'mc': _mc,
                                     'uc': len(_sub) - _mc, 'c': _c2, 's': _s2, 'i': _i2,
                                     'tt': _tt, 'mt': _mt, 'pct': _pct,
                                     'st_c': {st: len(_sub[_sub['Status'] == st]) for st in _all_st_bm}})
                if _has_mpl and _bm_data:
                    try:
                        _fig_bm, (_ax_bs, _ax_bl) = plt.subplots(
                            2, 1, figsize=(14, 4.2),
                            gridspec_kw={'height_ratios': [3, 1.5]})
                        _fig_bm.patch.set_facecolor('#F8F8F8')
                        _n_bm   = len(_bm_data)
                        _xm_bm  = range(_n_bm)
                        _xlbls_bm = [r['m'] for r in _bm_data]
                        _step_bm  = max(1, _n_bm // 18)
                        _tick_pos_bm  = list(range(0, _n_bm, _step_bm))
                        _tick_lbls_bm = [_xlbls_bm[i] for i in _tick_pos_bm]
                        _bottoms_bm = [0] * _n_bm
                        _bar_w_bm   = max(0.4, min(0.8, 12.0 / _n_bm))
                        for _st in _all_st_bm:
                            _vals = [r['st_c'].get(_st, 0) for r in _bm_data]
                            _ax_bs.bar(_xm_bm, _vals, bottom=_bottoms_bm, label=_st,
                                       width=_bar_w_bm,
                                       color=STATUS_COLORS.get(_st, '#90A4AE'), alpha=0.9)
                            _bottoms_bm = [b + v for b, v in zip(_bottoms_bm, _vals)]
                        _ax_bs.set_xticks(_tick_pos_bm)
                        _ax_bs.set_xticklabels(_tick_lbls_bm, rotation=40, ha='right', fontsize=8)
                        _ax_bs.set_title('2B Month-wise Invoice Status', fontsize=11,
                                         fontweight='bold', pad=4)
                        _ax_bs.set_ylabel('Invoice Count')
                        _ax_bs.legend(fontsize=8, loc='upper right', frameon=True,
                                      framealpha=0.8, ncol=3)
                        _ax_bs.set_xlim(-0.5, _n_bm - 0.5)
                        _ax_clean(_ax_bs)
                        _pct_line_bm = [r['mt']/r['tt']*100 if r['tt'] else 0 for r in _bm_data]
                        _mk_bm = max(2, min(5, 60 // _n_bm))
                        _ax_bl.plot(list(_xm_bm), _pct_line_bm, marker='o', color='#E65100',
                                    linewidth=1.8, markersize=_mk_bm)
                        _ax_bl.fill_between(list(_xm_bm), _pct_line_bm, alpha=0.12, color='#E65100')
                        _ax_bl.axhline(y=100, color='#4CAF50', linewidth=1,
                                       linestyle='--', alpha=0.6, label='100%')
                        _ax_bl.set_xticks(_tick_pos_bm)
                        _ax_bl.set_xticklabels(_tick_lbls_bm, rotation=40, ha='right', fontsize=8)
                        _ax_bl.set_ylim(0, 115)
                        _ax_bl.yaxis.set_major_formatter(
                            mticker.FuncFormatter(lambda x, _: f'{x:.0f}%'))
                        _ax_bl.set_title('% Amount Matched — 2B Month Trend', fontsize=9)
                        _ax_bl.set_xlim(-0.5, _n_bm - 0.5)
                        _ax_clean(_ax_bl)
                        if _n_bm <= 24:
                            for _xi, _pv in zip(_xm_bm, _pct_line_bm):
                                _ax_bl.annotate(f'{_pv:.0f}%', (_xi, _pv),
                                                textcoords='offset points', xytext=(0, 5),
                                                ha='center', fontsize=7, color='#E65100',
                                                fontweight='bold')
                        _fig_bm.tight_layout(pad=1.8)
                        _embed_fig(_fig_bm, bm_tab)
                        plt.close(_fig_bm)
                    except Exception:
                        pass
                _make_table(bm_tab,
                    [(r['m'], r['total'], r['mc'], r['uc'],
                      f"₹{r['c']:,.0f}", f"₹{r['s']:,.0f}", f"₹{r['i']:,.0f}",
                      f"₹{r['tt']:,.0f}", r['pct']) for r in _bm_data],
                    ('2B Month', 'Total', 'Matched', 'Unmatched',
                     'CGST', 'SGST', 'IGST', 'Total Tax', '% Amt Matched'),
                    (120, 65, 75, 85, 115, 115, 115, 125, 110), height=8)
            else:
                ctk.CTkLabel(bm_tab,
                             text="No 2B booking month data found in matched rows.",
                             font=ctk.CTkFont(size=13)).pack(pady=40)
        else:
            ctk.CTkLabel(bm_tab,
                         text="'Booking Month as per GSTR-2B' column not found.",
                         font=ctk.CTkFont(size=13)).pack(pady=40)

        # ══════════════════════════════════════════════════════════════════
        # TAB 3 — GSTN WISE
        # ══════════════════════════════════════════════════════════════════
        gw = tabs.tab("GSTN Wise")
        gstin_col = next((c for c in df.columns
                          if 'vendor' in c.lower() and 'gstn' in c.lower()), None)
        if not gstin_col:
            gstin_col = next((c for c in df.columns if 'gstin' in c.lower()), None)
        name_col = next((c for c in df.columns
                         if 'vendor' in c.lower() and 'name' in c.lower()), None)
        if gstin_col:
            def _pct_float(ps):
                try:
                    return float(str(ps).rstrip('%'))
                except Exception:
                    return 0.0

            gw_data = []
            for gstin_val, grp in df.groupby(df[gstin_col].fillna('Unknown')):
                mc    = len(grp[grp['Status'].isin(MATCHED_STATUSES)])
                c2, s2, i2 = _tax_sum(grp)
                tt    = c2 + s2 + i2
                mt    = _total_tax(grp[grp['Status'].isin(MATCHED_STATUSES)])
                pct_s = f"{min(100.0, mt/tt*100):.1f}%" if tt else "—"
                vendor = str(grp[name_col].iloc[0]).strip()[:30] if name_col else ''
                gw_data.append({'gstin': str(gstin_val), 'vendor': vendor,
                                'total': len(grp), 'mc': mc,
                                'c': c2, 's': s2, 'i': i2, 'tt': tt, 'mt': mt,
                                'pct': pct_s, 'pf': _pct_float(pct_s)})
            gw_data.sort(key=lambda r: r['tt'], reverse=True)

            if _has_mpl and gw_data:
                try:
                    fig_gw, (ax_hb, ax_gp) = plt.subplots(1, 2, figsize=(14, 3.8))
                    fig_gw.patch.set_facecolor('#F8F8F8')

                    # Vertical bar — top 10 vendors by tax amount
                    top10 = gw_data[:10]
                    gy = [r['gstin'][:12] for r in top10]
                    gx = [r['tt'] for r in top10]
                    gc = ['#4CAF50' if r['pf'] >= 90
                          else '#FFC107' if r['pf'] >= 50
                          else '#F44336' for r in top10]
                    vb_gw = ax_hb.bar(range(len(top10)), gx, color=gc,
                                      edgecolor='white', width=0.6, alpha=0.9)
                    ax_hb.set_xticks(range(len(top10)))
                    ax_hb.set_xticklabels(gy, rotation=30, ha='right', fontsize=7)
                    ax_hb.set_title('Top 10 Vendors — Tax Amount', fontsize=10,
                                    fontweight='bold', pad=4)
                    _rupee_fmt(ax_hb, 'y')
                    _ax_clean(ax_hb)
                    for bar, val in zip(vb_gw, gx):
                        if val > 0:
                            ax_hb.text(bar.get_x() + bar.get_width() / 2,
                                       bar.get_height() * 1.01, _fmt_amt(val),
                                       ha='center', va='bottom', fontsize=6.5,
                                       fontweight='bold')
                    ax_hb.legend(handles=[
                        mpatches.Patch(color='#4CAF50', label='≥90% matched'),
                        mpatches.Patch(color='#FFC107', label='50–90% matched'),
                        mpatches.Patch(color='#F44336', label='<50% matched'),
                    ], fontsize=7, frameon=False)

                    # Pie — top 5 vendor share
                    top5   = gw_data[:5]
                    others = sum(r['tt'] for r in gw_data[5:])
                    pie_lbl = [r['gstin'][:12] for r in top5] + (['Others'] if others else [])
                    pie_val = [r['tt'] for r in top5] + ([others] if others else [])
                    pie_col = ['#1565C0','#C62828','#2E7D32','#E65100','#6A1B9A','#90A4AE']
                    wg, _, atg = ax_gp.pie(
                        pie_val, labels=None, colors=pie_col[:len(pie_val)],
                        autopct='%1.1f%%', startangle=90, pctdistance=0.75,
                        wedgeprops=dict(edgecolor='white', linewidth=1.2))
                    for at in atg:
                        at.set_fontsize(8)
                    ax_gp.legend(wg, pie_lbl, loc='lower center',
                                 bbox_to_anchor=(0.5, -0.24), ncol=2,
                                 fontsize=7.5, frameon=False)
                    ax_gp.set_title('Top 5 Vendors — Tax Share', fontsize=10,
                                    fontweight='bold', pad=4)

                    fig_gw.tight_layout(pad=1.8)
                    _embed_fig(fig_gw, gw)
                    plt.close(fig_gw)
                except Exception:
                    pass

            # ── Sortable GSTN table ──────────────────────────────────────
            gw_col_names = ('GSTIN', 'Vendor Name', 'Total', 'Matched',
                            'CGST', 'SGST', 'IGST', 'Total Tax', '% Matched')
            gw_col_widths = (165, 195, 60, 70, 110, 110, 110, 120, 90)
            _gw_sort = {'key': 'tt', 'asc': False}

            sort_bar = ctk.CTkFrame(gw, fg_color="transparent")
            sort_bar.pack(fill="x", padx=6, pady=(4, 0))
            ctk.CTkLabel(sort_bar, text="Sort by Amount:",
                         font=ctk.CTkFont(size=10)).pack(side="left", padx=(4, 6))

            gw_table_frame = ctk.CTkFrame(gw, fg_color="transparent")
            gw_table_frame.pack(fill="x", padx=0, pady=0)

            def _populate_gw_tree(ascending=False):
                sorted_data = sorted(gw_data, key=lambda r: r['tt'],
                                     reverse=not ascending)
                for w in gw_table_frame.winfo_children():
                    w.destroy()
                style = ttk.Style()
                style.configure("GW.Treeview", rowheight=22, font=('Helvetica', 9))
                style.configure("GW.Treeview.Heading", font=('Helvetica', 9, 'bold'))
                outer = ctk.CTkFrame(gw_table_frame, fg_color="#F8F8F8", corner_radius=6)
                outer.pack(fill="x", padx=6, pady=(2, 6))
                vsb = ttk.Scrollbar(outer, orient="vertical")
                hsb = ttk.Scrollbar(outer, orient="horizontal")
                tree = ttk.Treeview(outer, columns=gw_col_names, show='headings',
                                    style="GW.Treeview",
                                    height=min(len(sorted_data), 10),
                                    yscrollcommand=vsb.set, xscrollcommand=hsb.set)
                vsb.configure(command=tree.yview)
                hsb.configure(command=tree.xview)
                for cn, w in zip(gw_col_names, gw_col_widths):
                    tree.heading(cn, text=cn, anchor='w')
                    tree.column(cn, width=w, minwidth=w, anchor='w')
                for r in sorted_data:
                    tree.insert('', 'end', values=(
                        r['gstin'], r['vendor'] or '—', r['total'], r['mc'],
                        f"₹{r['c']:,.0f}", f"₹{r['s']:,.0f}", f"₹{r['i']:,.0f}",
                        f"₹{r['tt']:,.0f}", r['pct']))
                vsb.pack(side="right", fill="y")
                hsb.pack(side="bottom", fill="x")
                tree.pack(fill="x", expand=False)

            def _sort_desc():
                _gw_sort['asc'] = False
                _populate_gw_tree(ascending=False)
                btn_desc.configure(fg_color=THEME_PRIMARY)
                btn_asc.configure(fg_color="#757575")

            def _sort_asc():
                _gw_sort['asc'] = True
                _populate_gw_tree(ascending=True)
                btn_asc.configure(fg_color=THEME_PRIMARY)
                btn_desc.configure(fg_color="#757575")

            btn_desc = ctk.CTkButton(sort_bar, text="↓ High → Low",
                                     command=_sort_desc, width=110, height=26,
                                     font=ctk.CTkFont(size=10),
                                     fg_color=THEME_PRIMARY)
            btn_desc.pack(side="left", padx=2)
            btn_asc = ctk.CTkButton(sort_bar, text="↑ Low → High",
                                    command=_sort_asc, width=110, height=26,
                                    font=ctk.CTkFont(size=10),
                                    fg_color="#757575")
            btn_asc.pack(side="left", padx=2)
            _populate_gw_tree(ascending=False)  # default: high to low
        else:
            ctk.CTkLabel(gw, text="No GSTIN column found.",
                         font=ctk.CTkFont(size=13)).pack(pady=40)

        # ══════════════════════════════════════════════════════════════════
        # TAB 4 — AMOUNT ANALYSIS
        # ══════════════════════════════════════════════════════════════════
        aa = tabs.tab("Amount Analysis")

        if _has_mpl:
            try:
                fig_aa, axes_aa = plt.subplots(1, 3, figsize=(14, 3.8))
                fig_aa.patch.set_facecolor('#F8F8F8')

                # [0] Gauge donut — amount match rate
                rem = max(0.0, 100.0 - pct_amt)
                axes_aa[0].pie([pct_amt, rem], colors=['#4CAF50', '#EEEEEE'],
                               startangle=90,
                               wedgeprops=dict(width=0.44, edgecolor='white'))
                axes_aa[0].text(0, 0.12, f'{pct_amt:.1f}%', ha='center', va='center',
                                fontsize=14, fontweight='bold', color='#2E7D32')
                axes_aa[0].text(0, -0.22, 'Amount\nMatched', ha='center', va='center',
                                fontsize=9, color='#555')
                axes_aa[0].set_title('Amount Match Rate', fontsize=11,
                                     fontweight='bold', pad=4)

                # [1] Stacked bar — matched vs unmatched by tax type
                tl, mp, up = [], [], []
                for col, lbl in [(_cgst_col,'CGST'), (_sgst_col,'SGST'), (_igst_col,'IGST')]:
                    if col and col in df.columns:
                        tl.append(lbl)
                        mp.append(pd.to_numeric(matched_df[col],  errors='coerce').fillna(0).sum())
                        up.append(pd.to_numeric(unmatched_df[col], errors='coerce').fillna(0).sum())
                if tl:
                    _AA_M_COLOR = '#4CAF50'
                    _AA_U_COLOR = '#EF5350'
                    xb = range(len(tl))
                    mb = axes_aa[1].bar(xb, mp,
                                        color=_AA_M_COLOR, alpha=0.9)
                    axes_aa[1].bar(xb, up, bottom=mp,
                                   color=_AA_U_COLOR, alpha=0.9)
                    axes_aa[1].set_xticks(list(xb))
                    axes_aa[1].set_xticklabels(tl, fontsize=9)
                    axes_aa[1].set_title('Matched vs Unmatched\nby Tax Type',
                                         fontsize=10, fontweight='bold', pad=4)
                    _rupee_fmt(axes_aa[1])
                    _ax_clean(axes_aa[1])
                    _aa1_handles = [
                        mpatches.Patch(color=_AA_M_COLOR, label='Matched'),
                        mpatches.Patch(color=_AA_U_COLOR, label='Unmatched'),
                    ]
                    axes_aa[1].legend(handles=_aa1_handles, fontsize=7,
                                      frameon=True, framealpha=0.85,
                                      loc='upper right', ncol=2)
                    for bar, mv, uv in zip(mb, mp, up):
                        cx = bar.get_x() + bar.get_width() / 2
                        if mv > 0:
                            axes_aa[1].text(cx, mv/2, _fmt_amt(mv),
                                            ha='center', va='center',
                                            fontsize=6.5, color='white', fontweight='bold')
                        if uv > 0:
                            axes_aa[1].text(cx, mv + uv/2, _fmt_amt(uv),
                                            ha='center', va='center',
                                            fontsize=6.5, color='white', fontweight='bold')

                # [2] ITC vs 2B or breakdown
                if b2b_total > 0:
                    rv_lbl = ['Books Amt', '2B Amt', 'Gap']
                    rv_val = [itc_total, b2b_total, abs(itc_total - b2b_total)]
                    rv_col = ['#1565C0', '#42A5F5', '#BBDEFB']
                else:
                    rv_lbl = ['Total Books', 'Matched', 'Unmatched']
                    rv_val = [itc_total, itc_matched, itc_unmatch]
                    rv_col = ['#1565C0', '#42A5F5', '#BBDEFB']
                rb = axes_aa[2].bar(rv_lbl, rv_val, color=rv_col,
                                    edgecolor='white', width=0.5, alpha=0.9)
                for bar, val in zip(rb, rv_val):
                    if val > 0:
                        axes_aa[2].text(bar.get_x() + bar.get_width()/2,
                                        bar.get_height() * 1.01, _fmt_amt(val),
                                        ha='center', va='bottom',
                                        fontsize=8, fontweight='bold')
                axes_aa[2].set_title(
                    'Books vs 2B vs Gap' if b2b_total > 0 else 'Books Amount Breakdown',
                    fontsize=10, fontweight='bold', pad=4)
                _rupee_fmt(axes_aa[2])
                _ax_clean(axes_aa[2])
                axes_aa[2].tick_params(axis='x', labelsize=8)

                fig_aa.tight_layout(pad=1.8)
                _embed_fig(fig_aa, aa)
                plt.close(fig_aa)
            except Exception:
                pass

        _make_table(aa, [
            ('Total Books Records',     f'{total:,}',                                   '—'),
            ('Matched Records',         f'{matched_df.shape[0]:,}',                     f'{pct_cnt:.1f}%'),
            ('Not Matched Records',     f'{unmatched_df.shape[0]:,}',                   f'{100-pct_cnt:.1f}%'),
            ('Total Books Tax Amt',     f'₹{itc_total:,.2f}',                           '100.0%'),
            ('Matched Tax Amt',         f'₹{itc_matched:,.2f}',                         f'{pct_amt:.1f}%'),
            ('Unmatched Tax Amt',       f'₹{itc_unmatch:,.2f}',                         f'{100-pct_amt:.1f}%'),
            ('GSTR-2B Total',           f'₹{b2b_total:,.2f}' if b2b_total else '—',     '—'),
            ('Difference (Books − 2B)', f'₹{itc_total-b2b_total:,.2f}' if b2b_total else '—', '—'),
        ], ('Metric', 'Value', '%'), (300, 220, 90), height=8)

        # ── Higher / Lower Amount Analysis ────────────────────────────────
        _diff_cgst_col = next((c for c in df.columns if c == 'Diff CGST'), None)
        _diff_sgst_col = next((c for c in df.columns if c == 'Diff SGST'), None)
        _diff_igst_col = next((c for c in df.columns if c == 'Diff IGST'), None)
        _has_diff_cols = any([_diff_cgst_col, _diff_sgst_col, _diff_igst_col])
        if _has_diff_cols:
            def _row_diff_total(row):
                d = 0.0
                for _dc in [_diff_cgst_col, _diff_sgst_col, _diff_igst_col]:
                    if _dc:
                        v = row.get(_dc, '')
                        if v != '' and not pd.isna(v):
                            try:
                                d += float(v)
                            except Exception:
                                pass
                return d
            _matched_only = df[df['Status'].isin(MATCHED_STATUSES)].copy()
            if not _matched_only.empty:
                _diffs = _matched_only.apply(_row_diff_total, axis=1)
                _higher_mask = _diffs > 0.005
                _lower_mask  = _diffs < -0.005
                _equal_mask  = ~_higher_mask & ~_lower_mask
                _higher_df = _matched_only[_higher_mask]
                _lower_df  = _matched_only[_lower_mask]
                _equal_df  = _matched_only[_equal_mask]
                _higher_amt = _total_tax(_higher_df)
                _lower_amt  = _total_tax(_lower_df)
                _equal_amt  = _total_tax(_equal_df)
                ctk.CTkLabel(aa, text="Higher / Lower Amount Analysis (Matched Rows)",
                             font=ctk.CTkFont(size=12, weight="bold"),
                             text_color="#4A148C").pack(anchor="w", padx=10, pady=(6, 0))
                _make_table(aa, [
                    ('Books Amount Higher than 2B', f'{len(_higher_df):,}', f'₹{_higher_amt:,.2f}'),
                    ('2B Amount Higher than Books',  f'{len(_lower_df):,}',  f'₹{_lower_amt:,.2f}'),
                    ('Amounts Equal',                f'{len(_equal_df):,}',  f'₹{_equal_amt:,.2f}'),
                ], ('Category', 'Count', 'Total Tax (Books)'), (280, 80, 190), height=4)

        win.update_idletasks()

        # ── Close ─────────────────────────────────────────────────────────
        ctk.CTkButton(win, text="✕  Close", command=win.destroy,
                      fg_color="#757575", hover_color="#616161",
                      height=32, width=120).pack(pady=6)

    # ── GSTN Debug ────────────────────────────────────────────────────────────
    def find_gstn_debug_candidates(self):
        """Find ITC-2A pairs where invoice + tax match but GSTIN is different/similar."""
        candidates = []
        if self.itc_result_df is None or self.itc_result_df.empty:
            return candidates
        if self.unmatched_2a_df is None or self.unmatched_2a_df.empty:
            return candidates

        vendor_gstn_col = vendor_inv_col = None
        for col in self.itc_result_df.columns:
            cl = col.lower().strip()
            if 'vendor' in cl and 'gstn' in cl:
                vendor_gstn_col = col
            elif 'vendor inv' in cl or 'external doc' in cl:
                vendor_inv_col = col

        itc_cgst_col, itc_sgst_col, itc_igst_col = find_tax_amount_columns(self.itc_result_df)
        taxable_col = next((c for c in self.itc_result_df.columns
                            if 'taxable' in c.lower() and 'value' in c.lower()), None)
        itc_date_col = next((c for c in self.itc_result_df.columns
                             if c.lower().strip() == 'invoice date'), None)

        if not vendor_gstn_col or not vendor_inv_col:
            return candidates

        if 'Status' not in self.itc_result_df.columns:
            return candidates
        unmatched_itc = self.itc_result_df[
            self.itc_result_df['Status'].isin(['Unmatched', 'Not found in 2B'])
        ]
        if unmatched_itc.empty:
            return candidates

        # Build 2A lookup by normalized invoice number (cross-GSTIN lookup)
        from collections import defaultdict
        twoa_by_inv = defaultdict(list)
        for idx, row in self.unmatched_2a_df.iterrows():
            norm_inv = normalize_invoice(str(row.get('Document_number', '')))
            if norm_inv:
                twoa_by_inv[norm_inv].append({
                    'idx': idx,
                    'gstn': str(row.get('GSTN', '')),
                    'raw_inv': str(row.get('Document_number', '')),
                    'cgst': safe_numeric_conversion(row.get('CGST', 0)),
                    'sgst': safe_numeric_conversion(row.get('SGST', 0)),
                    'igst': safe_numeric_conversion(row.get('IGST', 0)),
                    'tax': safe_numeric_conversion(row.get('TAX', 0)),
                    'date': str(row.get('Invoice_Date', '')),
                    'booking_month': str(row.get('Booking_Month', '')),
                    'type': str(row.get('TYPE', '')),
                })

        used_2a = set()
        for itc_idx, itc_row in unmatched_itc.iterrows():
            itc_gstin_norm = normalize_gstin(str(itc_row[vendor_gstn_col]))
            itc_inv_norm   = normalize_invoice(str(itc_row[vendor_inv_col]))
            itc_cgst   = safe_numeric_conversion(itc_row.get(itc_cgst_col, 0)) if itc_cgst_col else 0
            itc_sgst   = safe_numeric_conversion(itc_row.get(itc_sgst_col, 0)) if itc_sgst_col else 0
            itc_igst   = safe_numeric_conversion(itc_row.get(itc_igst_col, 0)) if itc_igst_col else 0
            itc_taxable= safe_numeric_conversion(itc_row.get(taxable_col, 0)) if taxable_col else 0

            for cand in twoa_by_inv.get(itc_inv_norm, []):
                if cand['idx'] in used_2a:
                    continue
                cand_gstin_norm = normalize_gstin(cand['gstn'])
                if cand_gstin_norm == itc_gstin_norm:
                    continue  # skip exact GSTIN match (already handled)
                # Tax amounts must match within Rs 50
                if (abs(itc_cgst - cand['cgst']) > 50 or
                    abs(itc_sgst - cand['sgst']) > 50 or
                    abs(itc_igst - cand['igst']) > 50):
                    continue
                gstin_sim = similarity(itc_gstin_norm, cand_gstin_norm)
                used_2a.add(cand['idx'])
                candidates.append({
                    'itc_index':   itc_idx,
                    'itc_invoice': str(itc_row[vendor_inv_col]),
                    'itc_gstin':   str(itc_row[vendor_gstn_col]),
                    'itc_date':    str(itc_row[itc_date_col]) if itc_date_col else '',
                    'itc_taxable': itc_taxable,
                    'itc_cgst': itc_cgst, 'itc_sgst': itc_sgst, 'itc_igst': itc_igst,
                    'twoa_invoice': cand['raw_inv'],
                    'twoa_gstin':   cand['gstn'],
                    'twoa_date':    cand['date'],
                    'twoa_taxable': cand['tax'],
                    'twoa_cgst': cand['cgst'], 'twoa_sgst': cand['sgst'], 'twoa_igst': cand['igst'],
                    'twoa_booking_month': cand['booking_month'],
                    'twoa_type':    cand['type'],
                    'gstin_similarity': gstin_sim,
                    'itc_inv_norm': itc_inv_norm,
                    'similarity': gstin_sim,
                })

        candidates.sort(key=lambda x: x['gstin_similarity'], reverse=True)
        return candidates

    def gstn_debug_matching(self):
        """Open GSTN Debug window — finds pairs where invoice+tax match but GSTIN differs."""
        if self.itc_result_df is None or self.itc_result_df.empty:
            messagebox.showwarning("Warning", "Run reconciliation first.")
            return
        candidates = self.find_gstn_debug_candidates()
        if not candidates:
            messagebox.showinfo("GSTN Debug",
                "No GSTN mismatch candidates found.\n\n"
                "This means all unmatched entries either have no corresponding 2B entry "
                "with the same invoice number, or the tax amounts don't agree.")
            return
        self._open_gstn_debug_window(candidates)

    def _open_gstn_debug_window(self, candidates):
        """Display GSTN mismatch candidates one at a time for user review."""
        debug_win = ctk.CTkToplevel(self)
        debug_win.title("GSTN Debug — Review GSTIN Mismatch Candidates")
        debug_win.geometry("960x520")
        debug_win.lift(); debug_win.focus_force()
        debug_win.after(100, lambda: debug_win.attributes('-topmost', False))
        debug_win.attributes('-topmost', True)

        vendor_gstn_col = vendor_inv_col = None
        for col in self.itc_result_df.columns:
            cl = col.lower().strip()
            if 'vendor' in cl and 'gstn' in cl:
                vendor_gstn_col = col
            elif 'vendor inv' in cl or 'external doc' in cl:
                vendor_inv_col = col

        from collections import defaultdict
        _norm_key_index = defaultdict(list)
        if vendor_gstn_col and vendor_inv_col:
            for _ri, _row in self.itc_result_df.iterrows():
                _nk = normalize_gstin(str(_row[vendor_gstn_col])) + '|' + normalize_invoice(str(_row[vendor_inv_col]))
                _norm_key_index[_nk].append(_ri)

        # Pre-build 2B results index: norm_key -> [row_indices in gstr_2a_results_df]
        _2a_result_index = defaultdict(list)
        if self.gstr_2a_results_df is not None and not self.gstr_2a_results_df.empty:
            for _ri, _row in self.gstr_2a_results_df.iterrows():
                _nk = normalize_gstin(str(_row.get('GSTN', ''))) + '|' + normalize_invoice(str(_row.get('Document_number', '')))
                _2a_result_index[_nk].append(_ri)

        state = {'idx': 0, 'matched': 0, 'skipped': 0, 'total': len(candidates), 'history': []}

        # Header
        hdr = ctk.CTkFrame(debug_win, fg_color="#FFF3E0")
        hdr.pack(fill="x", padx=10, pady=(10, 5))
        ctk.CTkLabel(hdr,
                     text="GSTN Debug: these entries share the same invoice number and tax amounts\n"
                          "but have DIFFERENT GSTIN values — one may be a typo.",
                     font=ctk.CTkFont(size=11), text_color="#E65100",
                     justify="center").pack(pady=(6, 2))
        progress_lbl = ctk.CTkLabel(hdr, text=f"Pair 1 of {state['total']}",
                                    font=ctk.CTkFont(size=14, weight="bold"))
        progress_lbl.pack(pady=(0, 6))

        fields = ['Invoice No', 'Invoice Date', 'GSTIN', 'Taxable Value', 'IGST', 'CGST', 'SGST']

        # ITC panel
        itc_frame = ctk.CTkFrame(debug_win)
        itc_frame.pack(fill="x", padx=10, pady=5)
        ctk.CTkLabel(itc_frame, text="Books Row (Your Register)",
                     font=ctk.CTkFont(size=13, weight="bold"),
                     text_color="#E91E63").pack(anchor="w", padx=10, pady=(5, 0))
        itc_grid = ctk.CTkFrame(itc_frame, fg_color="transparent")
        itc_grid.pack(fill="x", padx=10, pady=5)
        itc_vals = {}
        for i, f in enumerate(fields):
            ctk.CTkLabel(itc_grid, text=f"{f}:", font=ctk.CTkFont(size=11, weight="bold")
                         ).grid(row=0, column=i, padx=8, pady=2, sticky="w")
            lbl = ctk.CTkLabel(itc_grid, text="", font=ctk.CTkFont(size=11))
            lbl.grid(row=1, column=i, padx=8, pady=2, sticky="w")
            itc_vals[f] = lbl

        # 2A panel
        twoa_frame = ctk.CTkFrame(debug_win)
        twoa_frame.pack(fill="x", padx=10, pady=5)
        ctk.CTkLabel(twoa_frame, text="GSTR-2B Row (Portal)",
                     font=ctk.CTkFont(size=13, weight="bold"),
                     text_color="#1565C0").pack(anchor="w", padx=10, pady=(5, 0))
        twoa_grid = ctk.CTkFrame(twoa_frame, fg_color="transparent")
        twoa_grid.pack(fill="x", padx=10, pady=5)
        twoa_vals = {}
        for i, f in enumerate(fields):
            ctk.CTkLabel(twoa_grid, text=f"{f}:", font=ctk.CTkFont(size=11, weight="bold")
                         ).grid(row=0, column=i, padx=8, pady=2, sticky="w")
            lbl = ctk.CTkLabel(twoa_grid, text="", font=ctk.CTkFont(size=11))
            lbl.grid(row=1, column=i, padx=8, pady=2, sticky="w")
            twoa_vals[f] = lbl

        gstin_sim_lbl = ctk.CTkLabel(debug_win, text="",
                                     font=ctk.CTkFont(size=12), text_color="#F57F17")
        gstin_sim_lbl.pack(pady=(2, 0))

        remarks_frame = ctk.CTkFrame(debug_win, fg_color="transparent")
        remarks_frame.pack(fill="x", padx=10, pady=(4, 0))
        ctk.CTkLabel(remarks_frame, text="Remarks:",
                     font=ctk.CTkFont(size=12, weight="bold")).pack(side="left", padx=(10, 5))
        remarks_entry = ctk.CTkEntry(remarks_frame,
                                     placeholder_text="Optional note for this pair…",
                                     font=ctk.CTkFont(size=12), height=32)
        remarks_entry.pack(side="left", fill="x", expand=True, padx=(0, 10))

        btn_frame = ctk.CTkFrame(debug_win, fg_color="transparent")
        btn_frame.pack(fill="x", padx=10, pady=10)

        back_btn = ctk.CTkButton(btn_frame, text="Back", fg_color="#1565C0",
                                 hover_color="#0D47A1", font=ctk.CTkFont(size=14, weight="bold"),
                                 width=120, height=40, command=lambda: on_back(),
                                 state="disabled")
        back_btn.pack(side="left", padx=10)

        match_btn = ctk.CTkButton(btn_frame, text="Accept Match",
                                  fg_color="#2E7D32", hover_color="#1B5E20",
                                  font=ctk.CTkFont(size=14, weight="bold"),
                                  width=200, height=40, command=lambda: on_match())
        match_btn.pack(side="left", padx=20)

        skip_btn = ctk.CTkButton(btn_frame, text="Skip",
                                 fg_color="#757575", hover_color="#616161",
                                 font=ctk.CTkFont(size=14, weight="bold"),
                                 width=160, height=40, command=lambda: on_skip())
        skip_btn.pack(side="left", padx=20)

        ctk.CTkButton(btn_frame, text="☰ View All / Bulk Match",
                      fg_color="#5C6BC0", hover_color="#3949AB",
                      font=ctk.CTkFont(size=12), height=40, width=190,
                      command=lambda: self._open_bulk_debug_panel(
                          debug_win, candidates, _do_gstn_match_data, 'GSTN Debug')
                      ).pack(side="right", padx=10)

        counter_lbl = ctk.CTkLabel(debug_win, text="Matched: 0 | Skipped: 0",
                                   font=ctk.CTkFont(size=11))
        counter_lbl.pack(pady=4)

        def display_pair(i):
            if i >= state['total']:
                show_done()
                return
            match_btn.configure(state="normal")
            skip_btn.configure(state="normal")
            remarks_entry.configure(state="normal")
            p = candidates[i]
            progress_lbl.configure(text=f"Pair {i+1} of {state['total']}")
            gstin_sim_lbl.configure(
                text=f"GSTIN Similarity: {p['gstin_similarity']:.0%}  "
                     f"(ITC: {p['itc_gstin']}  vs  2B: {p['twoa_gstin']})")
            itc_vals['Invoice No'].configure(text=p['itc_invoice'])
            itc_vals['Invoice Date'].configure(text=p.get('itc_date', ''))
            itc_vals['GSTIN'].configure(text=p['itc_gstin'])
            itc_vals['Taxable Value'].configure(text=f"{p['itc_taxable']:,.2f}")
            itc_vals['IGST'].configure(text=f"{p['itc_igst']:,.2f}")
            itc_vals['CGST'].configure(text=f"{p['itc_cgst']:,.2f}")
            itc_vals['SGST'].configure(text=f"{p['itc_sgst']:,.2f}")
            twoa_vals['Invoice No'].configure(text=p['twoa_invoice'])
            twoa_vals['Invoice Date'].configure(text=p.get('twoa_date', ''))
            twoa_vals['GSTIN'].configure(text=p['twoa_gstin'])
            twoa_vals['Taxable Value'].configure(text=f"{p['twoa_taxable']:,.2f}")
            twoa_vals['IGST'].configure(text=f"{p['twoa_igst']:,.2f}")
            twoa_vals['CGST'].configure(text=f"{p['twoa_cgst']:,.2f}")
            twoa_vals['SGST'].configure(text=f"{p['twoa_sgst']:,.2f}")

            # Highlight fields: orange=partial-match, green=matching, red=mismatch
            def _clr(lbl, bg, fg):
                lbl.configure(fg_color=bg, text_color=fg)
            _G_BG, _G_FG = "#E8F5E9", "#1B5E20"
            _R_BG, _R_FG = "#FFEBEE", "#B71C1C"
            _O_BG, _O_FG = "#FFF3E0", "#E65100"
            # GSTIN: mismatch is the point in GSTN debug → red
            _clr(itc_vals['GSTIN'],  _R_BG, _R_FG)
            _clr(twoa_vals['GSTIN'], _R_BG, _R_FG)
            # Invoice No: should match in GSTN debug
            _inv_ok = normalize_invoice(p['itc_invoice']) == normalize_invoice(p['twoa_invoice'])
            _clr(itc_vals['Invoice No'],  _G_BG if _inv_ok else _O_BG, _G_FG if _inv_ok else _O_FG)
            _clr(twoa_vals['Invoice No'], _G_BG if _inv_ok else _O_BG, _G_FG if _inv_ok else _O_FG)
            # Tax amounts and taxable value
            for _f, _iv, _tv in [('IGST', p['itc_igst'], p['twoa_igst']),
                                  ('CGST', p['itc_cgst'], p['twoa_cgst']),
                                  ('SGST', p['itc_sgst'], p['twoa_sgst']),
                                  ('Taxable Value', p['itc_taxable'], p['twoa_taxable'])]:
                try:
                    _ok = abs(float(_iv) - float(_tv)) < 1
                except (TypeError, ValueError):
                    _ok = str(_iv).strip() == str(_tv).strip()
                _clr(itc_vals[_f],  _G_BG if _ok else _R_BG, _G_FG if _ok else _R_FG)
                _clr(twoa_vals[_f], _G_BG if _ok else _R_BG, _G_FG if _ok else _R_FG)
            # Invoice Date
            _d_ok = p.get('itc_date', '').strip() == p.get('twoa_date', '').strip()
            _clr(itc_vals['Invoice Date'],  _G_BG if _d_ok else _R_BG, _G_FG if _d_ok else _R_FG)
            _clr(twoa_vals['Invoice Date'], _G_BG if _d_ok else _R_BG, _G_FG if _d_ok else _R_FG)

            remarks_entry.delete(0, 'end')
            counter_lbl.configure(text=f"Matched: {state['matched']} | Skipped: {state['skipped']}")
            back_btn.configure(state="normal" if state['history'] else "disabled")

        def _apply_match(p):
            norm_key = normalize_gstin(p['itc_gstin']) + '|' + normalize_invoice(p['itc_invoice'])
            remark_text = remarks_entry.get().strip()
            auto = f"GSTN Debug matched with 2B GSTIN {p['twoa_gstin']}"
            note = f"[GSTN Match] {auto}" + (f" | {remark_text}" if remark_text else "")
            df = self.itc_result_df
            if 'Remarks' not in df.columns:
                df['Remarks'] = ''
            targets = _norm_key_index.get(norm_key, [p['itc_index']])
            tracked_cols = ['Status', 'Remarks', '2B Invoice No', '2B GSTIN',
                            'CGST as per 2B', 'SGST as per 2B', 'IGST as per 2B',
                            'Type', 'Booking Month as per GSTR-2B']
            prev_vals = {col: {ri: df.at[ri, col] if col in df.columns else ''
                               for ri in targets} for col in tracked_cols}
            # Save prev 2B result statuses for undo
            twoa_norm_key = normalize_gstin(p['twoa_gstin']) + '|' + normalize_invoice(p['twoa_invoice'])
            prev_2a_statuses = {}
            if self.gstr_2a_results_df is not None and 'Status' in self.gstr_2a_results_df.columns:
                for _ri in _2a_result_index.get(twoa_norm_key, []):
                    prev_2a_statuses[_ri] = self.gstr_2a_results_df.at[_ri, 'Status']
            for col in ['2B Invoice No', '2B GSTIN', 'CGST as per 2B', 'SGST as per 2B',
                        'IGST as per 2B', 'Type', 'Booking Month as per GSTR-2B']:
                if col not in df.columns:
                    df[col] = ''
            for ri in targets:
                df.at[ri, 'Status']              = 'Matched but invoice number is not accurate'
                df.at[ri, '2B Invoice No']        = p['twoa_invoice']
                df.at[ri, '2B GSTIN']             = p['twoa_gstin']
                df.at[ri, 'CGST as per 2B']       = p['twoa_cgst']
                df.at[ri, 'SGST as per 2B']       = p['twoa_sgst']
                df.at[ri, 'IGST as per 2B']       = p['twoa_igst']
                df.at[ri, 'Type']                 = p.get('twoa_type', '')
                df.at[ri, 'Booking Month as per GSTR-2B'] = p.get('twoa_booking_month', '')
                existing = str(df.at[ri, 'Remarks']).strip()
                df.at[ri, 'Remarks'] = (existing + '; ' + note) if existing else note
            # Mark the matching 2B row as Matched
            if self.gstr_2a_results_df is not None and 'Status' in self.gstr_2a_results_df.columns:
                for _ri in _2a_result_index.get(twoa_norm_key, []):
                    self.gstr_2a_results_df.at[_ri, 'Status'] = 'Matched'
            self._debug_final_matches.append({'cand_idx': state['idx'], **p})
            state['history'].append({'action': 'match', 'idx': state['idx'],
                                     'prev_vals': prev_vals, 'prev_2a_statuses': prev_2a_statuses})

        def on_match():
            _apply_match(candidates[state['idx']])
            state['matched'] += 1
            state['idx'] += 1
            display_pair(state['idx'])

        def on_skip():
            state['history'].append({'action': 'skip', 'idx': state['idx']})
            state['skipped'] += 1
            state['idx'] += 1
            display_pair(state['idx'])

        def on_back():
            if not state['history']:
                return
            entry = state['history'].pop()
            if entry['action'] == 'match':
                state['matched'] -= 1
                reverted = entry['idx']
                self._debug_final_matches = [m for m in self._debug_final_matches if m['cand_idx'] != reverted]
                df = self.itc_result_df
                for col, row_map in entry['prev_vals'].items():
                    if col in df.columns:
                        for ri, old in row_map.items():
                            df.at[ri, col] = old
                # Restore 2B result statuses
                if self.gstr_2a_results_df is not None:
                    for _ri, _old_status in entry.get('prev_2a_statuses', {}).items():
                        self.gstr_2a_results_df.at[_ri, 'Status'] = _old_status
            else:
                state['skipped'] -= 1
            state['idx'] = entry['idx']
            display_pair(state['idx'])

        def show_done():
            match_btn.configure(state="disabled")
            skip_btn.configure(state="disabled")
            remarks_entry.configure(state="disabled")
            progress_lbl.configure(text="Review Complete!")
            gstin_sim_lbl.configure(text="")
            for v in itc_vals.values():  v.configure(text="")
            for v in twoa_vals.values(): v.configure(text="")
            counter_lbl.configure(text=f"DONE — Matched: {state['matched']} | Skipped: {state['skipped']}")
            self.log(f"GSTN Debug: {state['matched']} matched, {state['skipped']} skipped")
            note = ("\n\nMatched pairs will be saved when you click\n'Save to YTD Database' on the dashboard."
                    if self._debug_final_matches else "")
            messagebox.showinfo("GSTN Debug Complete",
                                f"Pairs reviewed : {state['total']}\n"
                                f"Accepted       : {state['matched']}\n"
                                f"Skipped        : {state['skipped']}" + note,
                                parent=debug_win)

        def _do_gstn_match_data(p, remark=''):
            """Apply GSTN Debug match without touching step-by-step state (used by bulk panel)."""
            norm_key = normalize_gstin(p['itc_gstin']) + '|' + normalize_invoice(p['itc_invoice'])
            df = self.itc_result_df
            targets = _norm_key_index.get(norm_key, [p['itc_index']])
            for col in ['2B Invoice No', '2B GSTIN', 'CGST as per 2B', 'SGST as per 2B',
                        'IGST as per 2B', 'Type', 'Booking Month as per GSTR-2B']:
                if col not in df.columns:
                    df[col] = ''
            if 'Remarks' not in df.columns:
                df['Remarks'] = ''
            note = (f"[Bulk GSTN Match] GSTIN corrected to {p['twoa_gstin']}"
                    + (f" | {remark}" if remark else ""))
            for ri in targets:
                df.at[ri, 'Status']              = 'Matched but invoice number is not accurate'
                df.at[ri, '2B Invoice No']        = p['twoa_invoice']
                df.at[ri, '2B GSTIN']             = p['twoa_gstin']
                df.at[ri, 'CGST as per 2B']       = p['twoa_cgst']
                df.at[ri, 'SGST as per 2B']       = p['twoa_sgst']
                df.at[ri, 'IGST as per 2B']       = p['twoa_igst']
                df.at[ri, 'Type']                 = p.get('twoa_type', '')
                df.at[ri, 'Booking Month as per GSTR-2B'] = p.get('twoa_booking_month', '')
                existing = str(df.at[ri, 'Remarks']).strip()
                df.at[ri, 'Remarks'] = (existing + '; ' + note) if existing else note
            # Mark the matching 2B row as Matched
            twoa_norm_key = normalize_gstin(p['twoa_gstin']) + '|' + normalize_invoice(p['twoa_invoice'])
            if self.gstr_2a_results_df is not None and 'Status' in self.gstr_2a_results_df.columns:
                for _ri in _2a_result_index.get(twoa_norm_key, []):
                    self.gstr_2a_results_df.at[_ri, 'Status'] = 'Matched'
            self._debug_final_matches.append(p)
            self.log(f"Bulk GSTN Debug: '{p['itc_invoice']}' GSTIN {p['itc_gstin']} → {p['twoa_gstin']}")

        display_pair(0)

    def view_results(self):
        """Open results viewer window"""
        has_itc = self.itc_result_df is not None and not self.itc_result_df.empty
        if not has_itc:
            messagebox.showwarning("Warning", "No data available to view!")
            return

        co_title = f" — {self._company_name}" if self._company_name else ""
        results_window = ctk.CTkToplevel(self)
        results_window.title(f"Books Results{co_title}")
        results_window.geometry("1000x600")
        results_window.lift()
        results_window.focus_force()
        results_window.after(100, lambda: results_window.attributes('-topmost', False))
        results_window.attributes('-topmost', True)

        # Company info sub-header
        if self._company_name or self._company_gst:
            co_hdr = ctk.CTkFrame(results_window, fg_color=THEME_DARK, corner_radius=0)
            co_hdr.pack(fill="x")
            co_parts = [p for p in [self._company_name, self._company_gst,
                                     self._company_period] if p]
            ctk.CTkLabel(co_hdr, text="  |  ".join(co_parts),
                         font=ctk.CTkFont(size=11), text_color="#F8BBD0").pack(
                         side="left", padx=14, pady=6)

        # Notebook for tabs
        tabview = ctk.CTkTabview(results_window)
        tabview.pack(fill="both", expand=True, padx=10, pady=10)

        # Tab 1: Books Results with Status
        tab1 = tabview.add("Books Results")
        self.create_data_table(tab1, self.itc_result_df, status_field='Status',
                               status_values={'Matched': '#c8e6c9', 'Unmatched': '#ffcdd2',
                                              'Higher in 2B': '#fff9c4', 'Lower in 2B': '#bbdefb',
                                              'Not found in 2B': '#e1bee7',
                                              'Matched but invoice number is not accurate': '#ffe0b2'})

        # Tab 2: 2B Results (all GSTR 2B rows with Status)
        tab2 = tabview.add("2B Results")
        if self.gstr_2a_results_df is not None and not self.gstr_2a_results_df.empty:
            self.create_data_table(tab2, self.gstr_2a_results_df, status_field='Status',
                                   status_values={'Matched': '#c8e6c9', 'Unmatched': '#ffcdd2',
                                                  'Higher in 2B': '#fff9c4', 'Lower in 2B': '#bbdefb',
                                                  'Not Found in ITC': '#ffcdd2'})
        else:
            ctk.CTkLabel(tab2, text="No 2B results available!", font=ctk.CTkFont(size=16)).pack(pady=50)

        # Tab 3: Same Month Cancellation
        tab3 = tabview.add("Same Month Cancellation")
        if self.same_month_cancel_df is not None and not self.same_month_cancel_df.empty:
            ctk.CTkLabel(
                tab3,
                text="B2B invoices whose GSTIN and CGST/SGST/IGST match a CDNR entry (invoice number may differ)",
                font=ctk.CTkFont(size=11), text_color="#555555"
            ).pack(pady=(6, 0))
            self.create_data_table(tab3, self.same_month_cancel_df)
        else:
            ctk.CTkLabel(tab3, text="No same month cancellations found!", font=ctk.CTkFont(size=16)).pack(pady=50)

    def show_same_month_cancellations(self):
        """Open a dedicated window showing B2B-CDNR same-month cancellation matches."""
        if self.same_month_cancel_df is None or self.same_month_cancel_df.empty:
            messagebox.showinfo(
                "Same Month Cancellation",
                "No same month cancellations found.\n\n"
                "This means no B2B invoice has a CDNR entry with the same GSTIN "
                "and matching CGST/SGST/IGST amounts."
            )
            return

        win = ctk.CTkToplevel(self)
        win.title("Same Month Cancellation")
        win.geometry("1100x560")
        win.lift()
        win.focus_force()
        win.after(100, lambda: win.attributes('-topmost', False))
        win.attributes('-topmost', True)

        hdr = ctk.CTkFrame(win, fg_color="#5C6BC0", corner_radius=0)
        hdr.pack(fill="x")
        ctk.CTkLabel(
            hdr,
            text=f"Same Month Cancellation  —  {len(self.same_month_cancel_df)} pairs found",
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="white"
        ).pack(side="left", padx=16, pady=10)

        ctk.CTkLabel(
            win,
            text="B2B invoices whose GSTIN and CGST/SGST/IGST match a CDNR entry (invoice number may differ)",
            font=ctk.CTkFont(size=11), text_color="#555555"
        ).pack(pady=(6, 0))

        self.create_data_table(win, self.same_month_cancel_df)

        def _download():
            filepath = filedialog.asksaveasfilename(
                title="Save Same Month Cancellation",
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile="Same_Month_Cancellation.xlsx"
            )
            if filepath:
                try:
                    with pd.ExcelWriter(filepath, engine='openpyxl') as _w:
                        self.same_month_cancel_df.to_excel(_w, sheet_name='Same_Month_Cancellation', index=False)
                        _autofit_ws(_w.sheets['Same_Month_Cancellation'])
                    messagebox.showinfo("Success", f"Saved:\n{filepath}")
                except Exception as e:
                    messagebox.showerror("Error", f"Failed to save: {str(e)}")

        ctk.CTkButton(
            win, text="Download Excel",
            command=_download,
            fg_color=THEME_PRIMARY, hover_color=THEME_HOVER, height=36
        ).pack(pady=10)

    def create_data_table(self, parent, df, status_field=None, status_values=None):
        """Create a data table with treeview.
        status_field: column name to use for row coloring (e.g. 'Status')
        status_values: dict mapping status text to background color (e.g. {'Matched': '#c8e6c9'})
        """
        frame = ctk.CTkFrame(parent)
        frame.pack(fill="both", expand=True, padx=5, pady=5)

        # Fix Windows treeview tag background rendering
        _tv_style = ttk.Style()
        _tv_style.configure("Results.Treeview", rowheight=22, font=('Helvetica', 10))
        _tv_style.configure("Results.Treeview.Heading", font=('Helvetica', 10, 'bold'))
        _tv_style.map("Results.Treeview", background=[('selected', '#1565C0'), ('!selected', '')])

        # Summary counts
        total = len(df)
        if status_field and status_values and status_field in df.columns:
            summary_parts = [f"Total: {total}"]
            for status_name in status_values:
                count = len(df[df[status_field] == status_name])
                summary_parts.append(f"{status_name}: {count}")
            summary_text = " | ".join(summary_parts)
        else:
            matched = len(df[df['Status'] == 'MATCH']) if 'Status' in df.columns else 0
            mismatched = total - matched
            summary_text = f"Total: {total} | Matched: {matched} | Mismatched: {mismatched}"

        ctk.CTkLabel(frame, text=summary_text,
                     font=ctk.CTkFont(size=12, weight="bold")).pack(pady=(6, 2))

        # Color legend (only when status colors are in use)
        if status_field and status_values:
            legend_outer = ctk.CTkFrame(frame, fg_color="transparent")
            legend_outer.pack(fill="x", padx=8, pady=(0, 4))
            for status_name, hex_color in status_values.items():
                count = len(df[df[status_field] == status_name]) if status_field in df.columns else 0
                chip = tk.Frame(legend_outer, bg=hex_color, bd=1, relief="solid")
                chip.pack(side="left", padx=3, pady=2)
                tk.Label(chip, text=f"  {status_name} ({count})  ",
                         bg=hex_color, font=('Helvetica', 9)).pack()

        # Treeview with scrollbars
        tree_frame = ctk.CTkFrame(frame)
        tree_frame.pack(fill="both", expand=True, padx=5, pady=5)

        y_scroll = ttk.Scrollbar(tree_frame)
        y_scroll.pack(side="right", fill="y")
        x_scroll = ttk.Scrollbar(tree_frame, orient="horizontal")
        x_scroll.pack(side="bottom", fill="x")

        columns = list(df.columns)
        tree = ttk.Treeview(tree_frame, columns=columns, show='headings',
                            style="Results.Treeview",
                            yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)
        y_scroll.config(command=tree.yview)
        x_scroll.config(command=tree.xview)

        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, width=100, minwidth=50)

        # 90%+ match highlight on GSTIN/Invoice headings
        highlight_note = None
        if 'Status' in df.columns and total > 0:
            matched = len(df[df['Status'] == 'MATCH']) if 'MATCH' in df['Status'].values else len(df[df['Status'] == 'Matched'])
            if matched == 0 and df['Status'].dtype == 'object':
                matched = len(df[df['Status'].str.lower() == 'match'])
            if matched / total >= 0.9:
                for c in columns:
                    cl = c.lower()
                    if 'gstin' in cl or 'gstn' in cl or 'invoice' in cl or 'inv no' in cl:
                        tree.heading(c, text=f"{c} ✅")
                highlight_note = 'GSTIN & Invoice columns are 90%+ matched'

        # Tag colours
        if status_field and status_values:
            for status_name, color in status_values.items():
                tag_name = status_name.lower().replace(' ', '_')
                tree.tag_configure(tag_name, background=color,
                                   foreground='#1A237E' if 'match' in status_name.lower() else '#212121')
        tree.tag_configure('mismatch', background='#FFCDD2', foreground='#B71C1C')

        # Insert rows (capped at 2000 for performance)
        display_df = df.head(2000)
        for _, row in display_df.iterrows():
            values = [str(v) for v in row.values]
            if status_field and status_values and status_field in df.columns:
                status_val = str(row.get(status_field, ''))
                tag_name = status_val.lower().replace(' ', '_') if status_val in status_values else ''
                tree.insert('', 'end', values=values, tags=(tag_name,))
            else:
                tag = 'mismatch' if row.get('Status', '') != 'MATCH' else ''
                tree.insert('', 'end', values=values, tags=(tag,))

        tree.pack(fill="both", expand=True)

        if highlight_note:
            ctk.CTkLabel(frame, text=highlight_note,
                         font=ctk.CTkFont(size=11, weight='bold'),
                         text_color='#2E7D32').pack(pady=3)

        if len(df) > 2000:
            ctk.CTkLabel(frame,
                         text=f"Showing first 2,000 of {len(df)} rows. Download Excel for full data.",
                         font=ctk.CTkFont(size=11), text_color="orange").pack(pady=5)


if __name__ == "__main__":
    app = GSTReconciliationApp()
    app.mainloop()
